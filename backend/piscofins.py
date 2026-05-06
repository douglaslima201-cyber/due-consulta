"""
Motor de Análise Tributária PIS/COFINS — Regime Não Cumulativo
Leis 10.637/2002 e 10.833/2003 | SC COSIT nº 5/2018 | STJ REsp 1.221.170 (Tema 779)

Critérios de análise:
  - Somente contas analíticas de resultado (DRE): receitas, custos e despesas
  - Contas sintéticas e de balanço (ativo/passivo/PL) são identificadas e excluídas
  - Classificação contábil por CPCs
  - Teste de essencialidade e relevância (SC COSIT 5/2018)
"""
import io
import re
import uuid
from datetime import datetime

import pandas as pd
from flask import Blueprint, jsonify, request, send_file

bp = Blueprint("piscofins", __name__, url_prefix="/api/piscofins")
_analyses: dict[str, dict] = {}

PIS_RATE = 1.65
COFINS_RATE = 7.6
COMBINED_RATE = PIS_RATE + COFINS_RATE  # 9.25 %

# ─── 1. NATUREZA CONTÁBIL (CPC) ───────────────────────────────────────────────

# Padrões de contas de ATIVO e PASSIVO/PL para filtro
_BALANCE_SHEET_KEYWORDS = [
    # Ativo
    "ativo circulante", "ativo não circulante", "ativo imobilizado", "ativo intangível",
    "caixa e equivalentes", "caixa e bancos", "contas a receber", "clientes", "estoques",
    "estoque de", "adiantamentos a fornecedores", "aplicações financeiras",
    "investimentos", "depósitos judiciais", "ativo realizável", "ativo permanente",
    "imóveis", "terrenos", "máquinas e equipamentos", "veículos", "móveis e utensílios",
    "benfeitorias", "software ativado", "direitos de uso", "arrendamento ativo",
    "(-) depreciação acumulada", "amortização acumulada", "provisão para perdas",
    # Passivo
    "passivo circulante", "passivo não circulante", "fornecedores", "salários a pagar",
    "tributos a recolher", "financiamentos", "empréstimos", "debêntures",
    "contas a pagar", "provisão para", "adiantamentos de clientes", "arrendamento passivo",
    "pis a recolher", "cofins a recolher", "icms a recolher", "iss a pagar",
    "irpj a pagar", "csll a pagar", "dividendos a pagar",
    # Patrimônio Líquido
    "patrimônio líquido", "capital social", "capital subscrito", "reserva de capital",
    "reserva legal", "reserva de lucros", "lucros acumulados", "prejuízos acumulados",
    "ajuste de avaliação patrimonial", "ações em tesouraria",
]

_RESULT_KEYWORDS = [
    "receita", "faturamento", "vendas", "custo dos produtos", "custo das mercadorias",
    "custo dos serviços", "cpv", "cmv", "despesa", "despesas", "gastos",
    "resultado", "outras receitas", "outras despesas",
]


def _get_code_prefix(conta: str) -> str:
    """Extrai apenas dígitos e pontos do código contábil."""
    return re.sub(r"[^0-9.]", "", conta.strip())


def _first_digit(conta: str) -> str:
    code = _get_code_prefix(conta)
    return code[0] if code else ""


def detect_account_nature(conta: str, descricao: str) -> dict:
    """
    Classifica a natureza contábil da conta com referência ao CPC.
    Retorna: nature, cpc_ref, analyze (bool), skip_reason
    """
    code = _get_code_prefix(conta)
    fd = code[0] if code else ""
    dl = descricao.lower()

    # ── Balanço Patrimonial ──────────────────────────────────────────────────
    is_bs = any(kw in dl for kw in _BALANCE_SHEET_KEYWORDS)

    if fd == "1" or (is_bs and not any(k in dl for k in _RESULT_KEYWORDS)):
        # Subclassifica ativos para mensagem mais precisa
        if any(k in dl for k in ["imobilizado", "máquinas", "equipamentos", "veículos", "terrenos", "edificações"]):
            cpc = "CPC 27 / IAS 16 — Ativo Imobilizado"
        elif any(k in dl for k in ["intangível", "software ativado", "direito de uso de marca"]):
            cpc = "CPC 04 / IAS 38 — Ativo Intangível"
        elif any(k in dl for k in ["estoque", "mercadorias", "matéria-prima em estoque"]):
            cpc = "CPC 16 / IAS 2 — Estoques"
        elif any(k in dl for k in ["arrendamento", "direito de uso"]):
            cpc = "CPC 06 / IFRS 16 — Arrendamentos"
        else:
            cpc = "CPC 26 / IAS 1 — Balanço Patrimonial (Ativo)"
        return {"nature": "ATIVO", "cpc_ref": cpc, "analyze": False,
                "skip_reason": "Conta de ativo — não integra a DRE e não é base para crédito de PIS/COFINS."}

    if fd == "2":
        return {"nature": "PASSIVO", "cpc_ref": "CPC 26 / IAS 1 — Balanço Patrimonial (Passivo)",
                "analyze": False,
                "skip_reason": "Conta de passivo — não integra a DRE e não é base para crédito de PIS/COFINS."}

    if fd == "3":
        return {"nature": "PATRIMÔNIO LÍQUIDO", "cpc_ref": "CPC 26 / IAS 1 — Patrimônio Líquido",
                "analyze": False,
                "skip_reason": "Conta de patrimônio líquido — não integra a DRE."}

    # ── Resultado (DRE) ──────────────────────────────────────────────────────
    if fd == "4" or any(k in dl for k in ["receita bruta", "receita de vendas", "receita de serviços",
                                           "receita operacional", "receita líquida"]):
        return {"nature": "RECEITA", "cpc_ref": "CPC 47 / IFRS 15 — Receitas de Contratos com Clientes",
                "analyze": True, "skip_reason": None}

    if fd == "5" or any(k in dl for k in ["custo dos produtos", "custo das mercadorias",
                                            "custo dos serviços", "cpv", "cmv", "cpmv"]):
        return {"nature": "CUSTO", "cpc_ref": "CPC 16 / IAS 2 — Estoques / Custos de Produção",
                "analyze": True, "skip_reason": None}

    if fd == "6" or any(k in dl for k in ["despesa", "gastos com", "gasto com"]):
        if any(k in dl for k in ["financeira", "juros", "variação cambial", "iof"]):
            cpc = "CPC 48 / IFRS 9 — Instrumentos Financeiros (Despesas)"
        elif any(k in dl for k in ["vendas", "comercial", "distribuição"]):
            cpc = "CPC 26 / IAS 1 — Despesas de Vendas"
        elif any(k in dl for k in ["administrativ", "gerais e administrativ"]):
            cpc = "CPC 26 / IAS 1 — Despesas Gerais e Administrativas"
        else:
            cpc = "CPC 26 / IAS 1 — Despesas Operacionais"
        return {"nature": "DESPESA", "cpc_ref": cpc, "analyze": True, "skip_reason": None}

    if fd in ("7", "8", "9") or any(k in dl for k in ["outras receitas", "outras despesas", "resultado"]):
        return {"nature": "RESULTADO", "cpc_ref": "CPC 26 / IAS 1 — Outras Receitas/Despesas",
                "analyze": True, "skip_reason": None}

    # Sem código — tenta pelo nome
    if is_bs:
        return {"nature": "BALANÇO", "cpc_ref": "CPC 26", "analyze": False,
                "skip_reason": "Conta de balanço patrimonial — não analisada para PIS/COFINS."}

    return {"nature": "INDETERMINADO", "cpc_ref": "—", "analyze": True, "skip_reason": None}


# ─── 2. DETECÇÃO DE CONTAS SINTÉTICAS ────────────────────────────────────────

def mark_synthetic_accounts(rows: list[dict]) -> set[str]:
    """
    Retorna o conjunto de códigos sintéticos: conta X é sintética se existe
    outra conta cujo código começa com X + '.'.
    """
    codes = set()
    for r in rows:
        c = _get_code_prefix(r.get("_raw_conta", ""))
        if c:
            codes.add(c)

    synthetic = set()
    code_list = list(codes)
    for code in code_list:
        for other in code_list:
            if other != code and other.startswith(code + "."):
                synthetic.add(code)
                break
    return synthetic


# ─── 3. REGRAS TRIBUTÁRIAS PIS/COFINS ────────────────────────────────────────
# Fundamentação: Leis 10.637/2002 e 10.833/2003, Art. 3º
# SC COSIT 5/2018: critérios de essencialidade e relevância para "insumo"
# STJ REsp 1.221.170/PR (Tema 779): conceito amplo de insumo
# Demais SCs COSIT referenciadas individualmente.

RULES = [
    # ════════════════════════════════════════════════════════════════════════
    # CRÉDITOS CERTOS — BASE LEGAL EXPRESSA (Art. 3º)
    # ════════════════════════════════════════════════════════════════════════

    # ── Matéria-prima e insumos diretos ─────────────────────────────────────
    {
        "id": "mp_insumo_direto",
        "keywords": [
            "matéria-prima", "materia-prima", "matérias-primas", "materia prima",
            "material direto", "material de produção", "material produtivo",
            "insumo direto", "insumo de produção", "componente de fabricação",
            "embalagem primária", "embalagem secundária", "material de embalagem",
            "embalagens", "rótulos e embalagens", "material para envase",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 16 (Estoques) — custo de aquisição integra CPV",
        "category": "Insumos Diretos de Produção",
        "credit_type": "Insumo — Art. 3º, II",
        "eligible": "SIM",
        "risk": "BAIXO",
        "legal_basis": "Art. 3º, II, Leis 10.637/2002 e 10.833/2003; SC COSIT 5/2018 (essencialidade direta); STJ Tema 779",
        "comment": (
            "Bens adquiridos para uso ou consumo direto no processo de produção de bens destinados "
            "à venda ou na prestação de serviços. A SC COSIT 5/2018 confirma o critério da essencialidade "
            "direta: sem esses insumos a produção é inviável. O STJ (Tema 779) consolidou o entendimento "
            "de que o conceito de insumo abrange tudo que seja essencial ou relevante ao processo produtivo, "
            "superando a interpretação restritiva anterior da RFB. CRÉDITO PLENO — risco mínimo de autuação."
        ),
    },

    # ── Energia elétrica ────────────────────────────────────────────────────
    {
        "id": "energia_eletrica",
        "keywords": [
            "energia elétrica", "energia eletrica", "fornecimento de energia",
            "concessionária de energia", "conta de luz", "tarifa de energia",
            "cpfl", "celesc", "cemig", "coelba", "eletropaulo", "enel ", "copel",
            "light s.a", "energisa", "equatorial energia", "gasto com energia",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 16 / CPC 27 — compõe custo de fabricação ou despesa operacional",
        "category": "Energia Elétrica",
        "credit_type": "Energia Elétrica — Art. 3º, III",
        "eligible": "SIM",
        "risk": "BAIXO",
        "legal_basis": "Art. 3º, III, Lei 10.833/2003; SC COSIT 5/2018 (relevância ao processo); IN RFB 1.911/2019, art. 172",
        "comment": (
            "Energia elétrica consumida nos estabelecimentos é expressamente listada no Art. 3º, III "
            "como geradora de crédito, independentemente do teste de essencialidade. A SC COSIT 5/2018 "
            "reforça que energia consumida no processo produtivo é insumo por essencialidade, e consumida "
            "nas instalações administrativas gera crédito por previsão expressa. "
            "CRÉDITO PLENO — abrange toda a fatura de energia do CNPJ."
        ),
    },

    # ── Aluguel de prédios e instalações ────────────────────────────────────
    {
        "id": "aluguel_predial",
        "keywords": [
            "aluguel de prédio", "aluguel de galpão", "aluguel de armazém",
            "aluguel de imóvel", "locação de imóvel", "locação predial",
            "aluguel de fábrica", "aluguel de loja", "aluguel de escritório",
            "aluguel comercial", "locação comercial", "aluguel mensal",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 06 / IFRS 16 — contratos de arrendamento / locação",
        "category": "Aluguéis de Imóveis",
        "credit_type": "Aluguel — Art. 3º, IV",
        "eligible": "SIM",
        "risk": "BAIXO",
        "legal_basis": "Art. 3º, IV, Lei 10.833/2003; SC COSIT 5/2018; IN RFB 1.911/2019, art. 173",
        "comment": (
            "Aluguéis de prédios pagos a pessoas jurídicas, utilizados nas atividades da empresa, "
            "geram crédito por previsão expressa no Art. 3º, IV. Atenção: aluguel pago a pessoa física "
            "não gera crédito (vedação do §2º). Aluguel de imóvel que abriga atividade administrativa "
            "também é elegível — a lei não restringe ao ambiente produtivo. CRÉDITO PLENO quando locador é PJ."
        ),
    },

    # ── Aluguel de máquinas e equipamentos ──────────────────────────────────
    {
        "id": "aluguel_maquinas",
        "keywords": [
            "aluguel de máquinas", "aluguel de equipamentos", "locação de equipamentos",
            "locação de máquinas", "aluguel de veículo", "locação de veículos",
            "locação de frota", "aluguel de empilhadeiras", "aluguel de ferramentas",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 06 / IFRS 16 — arrendamento operacional de ativos",
        "category": "Aluguéis de Máquinas e Equipamentos",
        "credit_type": "Aluguel — Art. 3º, IV",
        "eligible": "SIM",
        "risk": "BAIXO",
        "legal_basis": "Art. 3º, IV, Lei 10.833/2003; SC COSIT 5/2018",
        "comment": (
            "Locação de máquinas e equipamentos utilizados nas atividades da empresa gera crédito "
            "por previsão expressa no Art. 3º, IV. Aplica-se tanto a equipamentos produtivos quanto "
            "administrativos. Locador deve ser pessoa jurídica. "
            "Veículos de frota: crédito admitido quando vinculados à atividade operacional."
        ),
    },

    # ── Arrendamento mercantil (leasing financeiro) ──────────────────────────
    {
        "id": "leasing",
        "keywords": [
            "leasing", "arrendamento mercantil", "arrendamento financeiro",
            "lease financeiro", "parcela de leasing", "contraprestação de arrendamento",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 06 / IFRS 16 — passivo de arrendamento",
        "category": "Arrendamento Mercantil (Leasing)",
        "credit_type": "Arrendamento — Art. 3º, V",
        "eligible": "SIM",
        "risk": "BAIXO",
        "legal_basis": "Art. 3º, V, Lei 10.833/2003; IN RFB 1.911/2019, art. 174",
        "comment": (
            "Arrendamento mercantil de máquinas, equipamentos e outros bens incorporados ao ativo "
            "imobilizado gera crédito por previsão expressa no Art. 3º, V. Vedado quando o arrendador "
            "é pessoa física ou empresa optante pelo Simples Nacional. Após o CPC 06/IFRS 16, "
            "contratos contabilizados como direito de uso também são elegíveis."
        ),
    },

    # ── Depreciação de bens do ativo imobilizado ─────────────────────────────
    {
        "id": "depreciacao",
        "keywords": [
            "depreciação de máquinas", "depreciação de equipamentos", "depreciação de veículos",
            "depreciação de edificações", "depreciação de instalações", "depreciação de ferramentas",
            "quota de depreciação", "encargo de depreciação", "depreciação acumulada —",
            "depreciação do período", "despesa de depreciação",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 27 / IAS 16 — vida útil e encargo de depreciação",
        "category": "Depreciação de Bens do Imobilizado",
        "credit_type": "Depreciação — Art. 3º, VI e VII",
        "eligible": "SIM",
        "risk": "BAIXO",
        "legal_basis": "Art. 3º, VI e VII, Lei 10.833/2003; IN RFB 1.911/2019, arts. 175-176",
        "comment": (
            "Encargos de depreciação de máquinas, equipamentos, edificações e outros bens do ativo "
            "imobilizado adquiridos para uso na produção de bens destinados à venda ou na prestação "
            "de serviços. O crédito é calculado sobre o encargo mensal (Art. 3º, VI) ou, se o bem "
            "foi adquirido após 01/05/2004, pode ser apropriar o crédito de forma acelerada "
            "(Art. 3º, VII — 1/48 ao mês). Bens usados na área administrativa também são elegíveis."
        ),
    },

    # ── Amortização de intangíveis ───────────────────────────────────────────
    {
        "id": "amortizacao",
        "keywords": [
            "amortização de intangíveis", "amortização de softwares", "amortização de patentes",
            "amortização de licenças", "amortização de marcas", "encargo de amortização",
            "despesa de amortização", "amortização do período",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 04 / IAS 38 — vida útil de ativo intangível",
        "category": "Amortização de Intangíveis",
        "credit_type": "Amortização — Art. 3º, VII",
        "eligible": "SIM",
        "risk": "MÉDIO",
        "legal_basis": "Art. 3º, VII, Lei 10.833/2003; SC COSIT 5/2018",
        "comment": (
            "Amortização de intangíveis adquiridos de terceiros (ex: softwares, patentes, licenças) "
            "e utilizados nas atividades da empresa pode gerar crédito sob o Art. 3º, VII. "
            "Controvérsia existe para intangíveis desenvolvidos internamente (CAPEX de software). "
            "A SC COSIT 5/2018 admite crédito quando o intangível é essencial ao processo produtivo. "
            "Recomendam-se laudos técnicos para suportar a essencialidade."
        ),
    },

    # ── Frete sobre vendas ───────────────────────────────────────────────────
    {
        "id": "frete_vendas",
        "keywords": [
            "frete sobre vendas", "frete de vendas", "frete sobre faturamento",
            "despesa com frete de entrega", "frete saída", "frete outbound",
            "frete de distribuição", "frete para clientes", "frete de exportação",
            "seguro e frete de saída", "custos com frete",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 47 — custo de obtenção de contrato / entrega ao cliente",
        "category": "Frete sobre Vendas",
        "credit_type": "Frete — Art. 3º, IX",
        "eligible": "SIM",
        "risk": "BAIXO",
        "legal_basis": "Art. 3º, IX, Lei 10.833/2003; SC COSIT 5/2018; IN RFB 1.911/2019, art. 178",
        "comment": (
            "Fretes pagos a PJ na venda de produtos, quando o ônus é suportado pelo vendedor, "
            "geram crédito por previsão expressa no Art. 3º, IX. Aplica-se tanto ao transporte "
            "de produtos acabados quanto de mercadorias revendidas. Frete para exportação também "
            "é elegível. CRÉDITO PLENO — risco mínimo, desde que o transportador seja PJ."
        ),
    },

    # ════════════════════════════════════════════════════════════════════════
    # CRÉDITOS POSSÍVEIS — INSUMO POR ESSENCIALIDADE (SC COSIT 5/2018)
    # ════════════════════════════════════════════════════════════════════════

    # ── Frete sobre compras de insumos ───────────────────────────────────────
    {
        "id": "frete_compras",
        "keywords": [
            "frete sobre compras", "frete de compras", "frete de aquisição",
            "frete inbound", "frete entrada", "fretes pagos", "frete e carretos",
            "frete e seguro de compras",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 16 — custo de aquisição de estoques inclui fretes",
        "category": "Frete sobre Compras",
        "credit_type": "Frete / Insumo — Art. 3º, II e IX",
        "eligible": "POSSÍVEL",
        "risk": "MÉDIO",
        "legal_basis": "Art. 3º, II e IX, Leis 10.637/2002 e 10.833/2003; SC COSIT 23/2020; SC COSIT 5/2018",
        "comment": (
            "Fretes pagos na aquisição de insumos integram o custo dos bens adquiridos (CPC 16). "
            "A SC COSIT 23/2020 admite o crédito de PIS/COFINS sobre frete de compra quando o bem "
            "adquirido é insumo elegível. O teste de essencialidade da SC COSIT 5/2018 aplica-se "
            "ao insumo principal, e o frete é acessório a ele. RECOMENDAÇÃO: segregar fretes de "
            "insumos elegíveis dos demais fretes para maximizar o crédito com segurança jurídica."
        ),
    },

    # ── Manutenção de máquinas e equipamentos produtivos ────────────────────
    {
        "id": "manutencao_produtiva",
        "keywords": [
            "manutenção de máquinas", "manutenção de equipamentos", "manutenção industrial",
            "manutenção preventiva", "manutenção corretiva", "manutenção preditiva",
            "revisão de equipamentos", "reforma de máquinas", "peças de reposição",
            "manutenção de linha de produção", "serviços de manutenção",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 27 — gastos subsequentes no ativo imobilizado",
        "category": "Manutenção de Equipamentos Produtivos",
        "credit_type": "Insumo (Essencialidade) — Art. 3º, II",
        "eligible": "POSSÍVEL",
        "risk": "MÉDIO",
        "legal_basis": "Art. 3º, II, Leis 10.637/2002 e 10.833/2003; SC COSIT 5/2018 (essencialidade); SC COSIT 101/2021",
        "comment": (
            "A SC COSIT 101/2021, fundamentada no critério da essencialidade da SC COSIT 5/2018, "
            "admite crédito sobre serviços de manutenção de máquinas e equipamentos utilizados na "
            "produção: sem manutenção adequada, o processo produtivo seria interrompido ou degradado "
            "(critério da essencialidade). Peças de reposição que não são capitalizadas (CPC 27) "
            "também são elegíveis como insumo. ATENÇÃO: manutenção de bens puramente administrativos "
            "tem posição mais frágil — segregar por finalidade do bem."
        ),
    },

    # ── Combustíveis e lubrificantes ─────────────────────────────────────────
    {
        "id": "combustivel",
        "keywords": [
            "combustível", "combustíveis", "diesel", "gasolina", "etanol", "gnv",
            "óleo diesel", "lubrificantes", "óleo lubrificante", "graxas",
            "combustível e lubrificantes", "abastecimento de frota produtiva",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 16 — insumo consumido no processo produtivo",
        "category": "Combustíveis e Lubrificantes",
        "credit_type": "Insumo (Essencialidade) — Art. 3º, II",
        "eligible": "POSSÍVEL",
        "risk": "MÉDIO",
        "legal_basis": "Art. 3º, II, Leis 10.637/2002 e 10.833/2003; SC COSIT 5/2018; SC COSIT 14/2021",
        "comment": (
            "A SC COSIT 14/2021 reconhece crédito sobre combustíveis utilizados em máquinas, "
            "fornos e geradores diretamente envolvidos no processo produtivo — critério de "
            "essencialidade da SC COSIT 5/2018 (ausência inviabiliza a produção). "
            "Combustível para veículos de entrega: elegível como insumo de distribuição. "
            "RISCO: combustível de frota administrativa ou veículos de diretores não é elegível — "
            "fundamental segregar por finalidade para evitar autuação."
        ),
    },

    # ── Água e utilidades produtivas ─────────────────────────────────────────
    {
        "id": "agua_utilidades",
        "keywords": [
            "água", "fornecimento de água", "concessionária de água", "conta de água",
            "sabesp", "caesb", "embasa", "sanepar", "compesa", "cedae",
            "gás industrial", "gás natural", "vapor industrial", "ar comprimido",
            "utilidades industriais", "gases industriais",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 16 — utilidade consumida no processo produtivo",
        "category": "Água e Utilidades Industriais",
        "credit_type": "Insumo (Essencialidade) — Art. 3º, II",
        "eligible": "POSSÍVEL",
        "risk": "MÉDIO",
        "legal_basis": "Art. 3º, II, Leis 10.637/2002 e 10.833/2003; SC COSIT 5/2018; SC COSIT 218/2019",
        "comment": (
            "A SC COSIT 218/2019 reconhece crédito sobre água consumida diretamente no processo "
            "produtivo (limpeza de equipamentos, resfriamento, componente de produto). "
            "A SC COSIT 5/2018 fundamenta: água é essencial quando sua ausência paralisa a produção. "
            "Uso exclusivamente sanitário/administrativo não gera crédito. "
            "RECOMENDAÇÃO: medir consumo produtivo x administrativo e apropriar proporcionalmente. "
            "Gás industrial tem posição mais sólida — essencialidade direta ao processo térmico/químico."
        ),
    },

    # ── Serviços de terceiros aplicados na produção ──────────────────────────
    {
        "id": "servicos_terceiros_prod",
        "keywords": [
            "serviços de terceiros industriais", "beneficiamento", "industrialização por encomenda",
            "serviços industriais", "serviços de processamento", "serviços de montagem",
            "serviços de embalagem", "serviços de etiquetagem", "serviços de corte",
            "serviços de costura", "serviços de solda", "serviços de pintura industrial",
            "mão de obra terceirizada produção", "terceirização produtiva",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 16 — custo de transformação inclui serviços aplicados ao produto",
        "category": "Serviços de Terceiros (Produção)",
        "credit_type": "Insumo (Essencialidade) — Art. 3º, II",
        "eligible": "POSSÍVEL",
        "risk": "MÉDIO",
        "legal_basis": "Art. 3º, II, Leis 10.637/2002 e 10.833/2003; SC COSIT 5/2018; STJ Tema 779",
        "comment": (
            "O STJ (Tema 779) e a SC COSIT 5/2018 confirmam que serviços de terceiros aplicados "
            "diretamente na produção são insumos por essencialidade ou relevância. "
            "Industrialização por encomenda tem crédito consolidado (base legal expressa no Art. 3º, II). "
            "Serviços de beneficiamento: crédito amplamente aceito. "
            "AÇÃO: identificar contratualmente se o serviço modifica ou é integrado ao produto — "
            "quanto mais direta a vinculação, mais sólida a posição."
        ),
    },

    # ── EPI e segurança do trabalho ──────────────────────────────────────────
    {
        "id": "epi",
        "keywords": [
            "epi", "equipamento de proteção individual", "luvas de proteção",
            "capacetes", "óculos de proteção", "uniforme de proteção",
            "vestimenta de proteção", "botina de segurança", "protetor auricular",
            "segurança do trabalho", "uniforme operacional", "fardamento",
            "equipamentos de segurança do trabalho",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 16 — custo do período vinculado à mão de obra de produção",
        "category": "EPI e Segurança do Trabalho",
        "credit_type": "Insumo (Essencialidade/Relevância) — Art. 3º, II",
        "eligible": "POSSÍVEL",
        "risk": "MÉDIO",
        "legal_basis": "Art. 3º, II, Leis 10.637/2002 e 10.833/2003; SC COSIT 5/2018; SC COSIT 34/2014",
        "comment": (
            "A SC COSIT 34/2014, fundamentada no critério de essencialidade e relevância da "
            "SC COSIT 5/2018, reconhece crédito sobre EPIs obrigatórios (NR-6) utilizados pelos "
            "trabalhadores diretamente na produção. O teste: sem o EPI a atividade produtiva seria "
            "legalmente impedida ou exporia a empresa a sanções (essencialidade indireta). "
            "Uniformes de identificação (sem função protetiva) têm posição mais fraca — "
            "recomendam-se laudos ocupacionais. Crédito restrito ao pessoal de produção."
        ),
    },

    # ── Seguros de bens produtivos ───────────────────────────────────────────
    {
        "id": "seguro_produtivo",
        "keywords": [
            "seguro de máquinas", "seguro de equipamentos", "seguro patrimonial",
            "seguro de instalações produtivas", "seguro de risco de engenharia",
            "seguro de carga", "seguro de transporte de mercadorias",
            "seguro industrial", "prêmio de seguro operacional",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 27 — custo atribuível à manutenção do ativo produtivo",
        "category": "Seguros de Ativos Produtivos",
        "credit_type": "Insumo (Relevância) — Art. 3º, II",
        "eligible": "POSSÍVEL",
        "risk": "MÉDIO",
        "legal_basis": "Art. 3º, II, Leis 10.637/2002 e 10.833/2003; SC COSIT 5/2018 (relevância ao processo)",
        "comment": (
            "Com base no critério de relevância da SC COSIT 5/2018, seguros sobre bens diretamente "
            "utilizados na produção são elegíveis: a proteção dos ativos produtivos é relevante para "
            "garantir a continuidade e qualidade do processo. CARF (Acórdão 9303-012.682) já admitiu "
            "crédito em casos de seguro obrigatório de carga. "
            "ATENÇÃO: seguro de vida de diretores e seguro de saúde de empregados não são elegíveis. "
            "Posição requer suporte técnico — recomenda-se parecer."
        ),
    },

    # ── Telecomunicações essenciais ao processo ──────────────────────────────
    {
        "id": "telecom_operacional",
        "keywords": [
            "comunicação industrial", "sistema de monitoramento remoto",
            "telemetria", "comunicação de dados de produção", "supervisão remota",
            "link dedicado operacional", "comunicação de equipamentos",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 16 — custo de produção quando essencial ao processo",
        "category": "Telecomunicações Operacionais",
        "credit_type": "Insumo (Essencialidade) — Art. 3º, II",
        "eligible": "POSSÍVEL",
        "risk": "ALTO",
        "legal_basis": "Art. 3º, II, Leis 10.637/2002 e 10.833/2003; SC COSIT 5/2018 — requer comprovação de essencialidade direta",
        "comment": (
            "Aplicar o teste da SC COSIT 5/2018: a comunicação é essencial ao processo se sua "
            "ausência impede ou paralisa a produção (ex: telemetria de equipamentos críticos, "
            "supervisão remota de processos contínuos). "
            "A RFB tem autuado créditos sobre telefonia de uso geral — ALTO RISCO sem documentação "
            "técnica robusta provando a essencialidade direta ao processo produtivo específico. "
            "RECOMENDAÇÃO FORTE: laudo técnico identificando os sistemas que dependem dessa comunicação."
        ),
    },

    # ── Serviços de TI essenciais à produção ─────────────────────────────────
    {
        "id": "ti_produtivo",
        "keywords": [
            "sistema mes", "sistema de controle de produção", "automação industrial",
            "sistema scada", "software de controle de qualidade", "sistema de rastreabilidade",
            "software industrial", "sistema de gestão de produção",
            "licença de software produtivo", "sistema erp produção",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 04 — intangível; pode compor custo de produção",
        "category": "TI Aplicada à Produção",
        "credit_type": "Insumo (Essencialidade) — Art. 3º, II",
        "eligible": "POSSÍVEL",
        "risk": "MÉDIO",
        "legal_basis": "Art. 3º, II, Leis 10.637/2002 e 10.833/2003; SC COSIT 5/2018; jurisprudência CARF em consolidação",
        "comment": (
            "Sistemas de TI diretamente vinculados ao controle e execução do processo produtivo "
            "(MES, SCADA, rastreabilidade, controle de qualidade) satisfazem o critério de "
            "essencialidade da SC COSIT 5/2018: sem eles o processo é impossível ou a conformidade "
            "regulatória não é atingida. O CARF tem reconhecido progressivamente esses créditos. "
            "ERP administrativo não é elegível. SEGREGAR claramente por módulo/finalidade."
        ),
    },

    # ════════════════════════════════════════════════════════════════════════
    # NÃO ELEGÍVEIS — VEDAÇÕES EXPRESSAS E FALTA DE PREVISÃO LEGAL
    # ════════════════════════════════════════════════════════════════════════

    # ── Receitas (base de cálculo, não crédito) ──────────────────────────────
    {
        "id": "receitas",
        "keywords": [
            "receita bruta de vendas", "receita de prestação de serviços",
            "receita operacional bruta", "receita líquida de vendas",
            "receita de exportação", "outras receitas operacionais",
            "receita financeira", "receita de dividendos", "receita de aluguéis recebidos",
        ],
        "nature_filter": ["RECEITA", "RESULTADO", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 47 — reconhecimento de receita",
        "category": "Receitas (Base de Cálculo)",
        "credit_type": "N/A — Receita, não despesa",
        "eligible": "NÃO",
        "risk": "BAIXO",
        "legal_basis": "Art. 1º, Leis 10.637/2002 e 10.833/2003 — receita é base de cálculo do débito",
        "comment": (
            "Contas de receita representam a base de cálculo do PIS e da COFINS (débito), não "
            "a base para apuração de créditos. O regime não cumulativo funciona pela subtração "
            "dos créditos do débito apurado sobre a receita bruta. Receitas de exportação "
            "(Art. 5º) são isentas, gerando crédito presumido — verificar separadamente."
        ),
    },

    # ── Pessoal e encargos sociais ────────────────────────────────────────────
    {
        "id": "pessoal",
        "keywords": [
            "salários", "salários e ordenados", "remuneração de pessoal",
            "pro labore", "pró-labore", "13° salário", "décimo terceiro",
            "férias", "adicional de férias", "adicional de insalubridade",
            "adicional de periculosidade", "horas extras", "comissões de vendas",
            "encargos sociais", "inss patronal", "fgts", "contribuição previdenciária",
            "vale refeição", "vale alimentação", "vale transporte", "vale cultura",
            "assistência médica", "plano de saúde", "seguro de vida em grupo",
            "previdência complementar", "ticket refeição", "auxílio alimentação",
            "benefícios a empregados", "participação nos lucros", "plr",
            "rescisão contratual", "aviso prévio indenizado", "multa do fgts",
            "contribuição sindical", "custo de pessoal",
        ],
        "nature_filter": ["CUSTO", "DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 33 / IAS 19 — Benefícios a Empregados",
        "category": "Pessoal e Encargos Sociais",
        "credit_type": "N/A — Vedação expressa",
        "eligible": "NÃO",
        "risk": "BAIXO",
        "legal_basis": "Art. 3º, §2º, I, Leis 10.637/2002 e 10.833/2003 — vedação expressa",
        "comment": (
            "Mão de obra remunerada por empregados é expressamente excluída do direito a crédito "
            "pelo Art. 3º, §2º, I, independentemente do regime não cumulativo. A vedação é "
            "objetiva e não comporta exceção — abrange salários, encargos trabalhistas e "
            "previdenciários, benefícios e verbas rescisórias. "
            "A SC COSIT 5/2018 não alterou essa vedação. Creditamento indevido gera multa de 75%."
        ),
    },

    # ── Tributos sobre o lucro ────────────────────────────────────────────────
    {
        "id": "tributos_lucro",
        "keywords": [
            "irpj", "imposto de renda pessoa jurídica", "provisão para irpj",
            "csll", "contribuição social sobre o lucro", "provisão para csll",
            "irrf sobre serviços", "imposto de renda retido na fonte",
        ],
        "nature_filter": ["DESPESA", "RESULTADO", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 32 / IAS 12 — Tributos sobre o Lucro",
        "category": "Tributos sobre o Lucro",
        "credit_type": "N/A — Vedação expressa",
        "eligible": "NÃO",
        "risk": "BAIXO",
        "legal_basis": "Art. 3º, §2º, II, Leis 10.637/2002 e 10.833/2003 — vedação expressa",
        "comment": (
            "IRPJ e CSLL são expressamente vedados como base de crédito de PIS/COFINS pelo "
            "Art. 3º, §2º, II. Vedação objetiva e não há posição favorável na jurisprudência. "
            "Creditamento sobre IRPJ/CSLL é um dos itens de maior materialidade nas autuações "
            "da Receita Federal — risco de multa qualificada (150%) se houver dolo."
        ),
    },

    # ── Multas, penalidades e indenizações ────────────────────────────────────
    {
        "id": "multas_penalidades",
        "keywords": [
            "multa fiscal", "multa de trânsito", "multa contratual", "penalidade",
            "auto de infração", "indenização por rescisão", "sinistros",
            "multa por atraso", "juros de mora pagos", "multa de mora",
        ],
        "nature_filter": ["DESPESA", "RESULTADO", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 25 / IAS 37 — Provisões e Passivos Contingentes",
        "category": "Multas, Penalidades e Indenizações",
        "credit_type": "N/A — Sem previsão legal",
        "eligible": "NÃO",
        "risk": "ALTO",
        "legal_basis": "Ausência de previsão no Art. 3º, Leis 10.637/2002 e 10.833/2003; SC COSIT 5/2018 não ampara",
        "comment": (
            "Multas e penalidades não satisfazem nenhum dos critérios do Art. 3º nem os critérios "
            "de essencialidade e relevância da SC COSIT 5/2018: não são insumos, não são "
            "depreciação, não são aluguel. Representam saídas decorrentes de descumprimento de "
            "obrigações, não vinculadas ao processo produtivo. "
            "Creditamento indevido expõe a empresa a autuação com multa de 75% a 150%."
        ),
    },

    # ── Despesas financeiras ──────────────────────────────────────────────────
    {
        "id": "despesas_financeiras",
        "keywords": [
            "juros sobre empréstimos", "juros sobre financiamentos", "juros bancários",
            "juros devedores", "despesas financeiras", "encargos financeiros",
            "variação cambial passiva", "iof", "taxa de abertura de crédito",
            "spread bancário", "custo de capital de giro", "juros sobre debêntures",
        ],
        "nature_filter": ["DESPESA", "RESULTADO", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 48 / IFRS 9 — Instrumentos Financeiros (custo amortizado)",
        "category": "Despesas Financeiras",
        "credit_type": "N/A — Sem previsão no regime geral",
        "eligible": "NÃO",
        "risk": "BAIXO",
        "legal_basis": "Ausência de previsão no Art. 3º para regime geral; SC COSIT 5/2018 não ampara despesas financeiras",
        "comment": (
            "Despesas financeiras não estão listadas no Art. 3º das Leis 10.637/2002 e 10.833/2003 "
            "para o regime geral. A SC COSIT 5/2018 não estende o conceito de insumo a instrumentos "
            "financeiros. Exceção aplicável apenas a instituições financeiras sujeitas ao regime "
            "diferenciado (IN RFB 1.285/2012). Para empresas do regime geral: sem crédito."
        ),
    },

    # ── Marketing e publicidade ───────────────────────────────────────────────
    {
        "id": "marketing",
        "keywords": [
            "publicidade e propaganda", "agência de publicidade", "mídia digital",
            "investimento em marketing", "patrocínio", "branding",
            "material promocional", "marketing digital", "gestão de redes sociais",
            "anúncios pagos", "google ads", "meta ads", "influenciadores digitais",
        ],
        "nature_filter": ["DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 26 / IAS 1 — Despesas de Vendas (reconhecimento no período)",
        "category": "Marketing e Publicidade",
        "credit_type": "N/A — Sem previsão legal; não satisfaz SC COSIT 5/2018",
        "eligible": "NÃO",
        "risk": "MÉDIO",
        "legal_basis": "Ausência de previsão no Art. 3º, Lei 10.833/2003; RFB e CARF negam sistematicamente",
        "comment": (
            "Despesas de marketing e publicidade não estão listadas no Art. 3º e não satisfazem "
            "os critérios de essencialidade e relevância da SC COSIT 5/2018: não são bens/serviços "
            "integrados ou consumidos no processo produtivo. "
            "O CARF consolidou o entendimento de que marketing é despesa de venda, não insumo. "
            "Creditamento gera risco de autuação — evitar."
        ),
    },

    # ── Serviços administrativos, jurídicos e contábeis ──────────────────────
    {
        "id": "servicos_adm",
        "keywords": [
            "honorários advocatícios", "serviços jurídicos", "serviços contábeis",
            "auditoria", "consultoria administrativa", "consultoria estratégica",
            "honorários de consultoria", "serviços de assessoria",
            "serviços de recursos humanos", "headhunting", "serviços de ti administrativos",
            "suporte de ti administrativo", "manutenção de sistemas administrativos",
        ],
        "nature_filter": ["DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 26 / IAS 1 — Despesas Gerais e Administrativas",
        "category": "Serviços Administrativos / Consultoria",
        "credit_type": "N/A — Sem previsão; não satisfaz SC COSIT 5/2018",
        "eligible": "NÃO",
        "risk": "MÉDIO",
        "legal_basis": "Ausência de previsão no Art. 3º; SC COSIT 5/2018 — não são insumos por não integrarem o processo produtivo",
        "comment": (
            "Serviços jurídicos, contábeis e consultoria administrativa não integram nem são "
            "consumidos no processo de produção de bens ou prestação de serviços ao cliente final. "
            "A SC COSIT 5/2018 exige vínculo ao processo produtivo — despesas com suporte da "
            "estrutura administrativa da empresa ficam fora do escopo. "
            "CARF nega crédito sistematicamente. Risco de autuação com 75% de multa."
        ),
    },

    # ── Viagens, diárias e representação ─────────────────────────────────────
    {
        "id": "viagens",
        "keywords": [
            "passagens aéreas", "hospedagem", "hotel", "diárias de viagem",
            "reembolso de viagem", "despesas de viagem", "representação comercial",
            "aluguel de veículo para viagem", "transporte de executivos",
        ],
        "nature_filter": ["DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 26 / IAS 1 — Despesas Gerais e Administrativas",
        "category": "Viagens e Representação",
        "credit_type": "N/A — Sem previsão legal",
        "eligible": "NÃO",
        "risk": "MÉDIO",
        "legal_basis": "Ausência de previsão no Art. 3º, Lei 10.833/2003; SC COSIT 5/2018 não ampara",
        "comment": (
            "Despesas de viagem e representação não satisfazem os critérios de essencialidade e "
            "relevância da SC COSIT 5/2018 para insumos. Não são bens/serviços que integram ou "
            "são consumidos no processo produtivo. Vedação implícita pela ausência de previsão no "
            "Art. 3º. Sem jurisprudência favorável relevante."
        ),
    },

    # ── Material de escritório, limpeza e higiene (adm) ──────────────────────
    {
        "id": "mat_adm",
        "keywords": [
            "material de escritório", "material de expediente", "papelaria",
            "material de limpeza administrativa", "produtos de higiene administrativa",
            "café e copa", "material de copa e cozinha", "material de informática administrativo",
        ],
        "nature_filter": ["DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 26 / IAS 1 — Despesas Gerais e Administrativas",
        "category": "Material de Escritório / Administrativo",
        "credit_type": "N/A — Sem previsão; uso administrativo não satisfaz SC COSIT 5/2018",
        "eligible": "NÃO",
        "risk": "MÉDIO",
        "legal_basis": "Ausência de previsão no Art. 3º; SC COSIT 5/2018 — não integra processo produtivo",
        "comment": (
            "Materiais de escritório e limpeza de uso administrativo não satisfazem os critérios "
            "da SC COSIT 5/2018: não são essenciais nem relevantes ao processo produtivo. "
            "Exceção possível para material de limpeza utilizado diretamente no processo produtivo "
            "(ex: limpeza de equipamentos de alimentos) — nesse caso, reclassificar como insumo "
            "e suportar com laudo técnico."
        ),
    },

    # ── Telecomunicações de uso geral ─────────────────────────────────────────
    {
        "id": "telecom_geral",
        "keywords": [
            "telefonia fixa", "telefonia móvel", "telefone corporativo",
            "celular corporativo", "plano de dados", "banda larga administrativa",
            "internet corporativa", "claro empresarial", "vivo empresas",
            "tim empresarial", "oi empresarial", "conta de telefone",
        ],
        "nature_filter": ["DESPESA", "INDETERMINADO"],
        "cpc_tax_ref": "CPC 26 / IAS 1 — Despesas Gerais e Administrativas",
        "category": "Telecomunicações (Uso Geral)",
        "credit_type": "N/A — Alto risco; SC COSIT 5/2018 exige essencialidade direta ao processo",
        "eligible": "NÃO",
        "risk": "ALTO",
        "legal_basis": "Ausência de previsão expressa no Art. 3º; RFB nega; SC COSIT 5/2018 não ampara uso geral",
        "comment": (
            "Telefonia e internet de uso geral administrativo não satisfazem o teste da SC COSIT "
            "5/2018: não são essenciais ao processo produtivo (essencial = ausência impede a produção). "
            "A RFB tem autuado consistentemente créditos sobre telecomunicações sem segregação. "
            "ALTO RISCO. Se houver uso produtivo específico e comprovável, segregar e reclassificar "
            "a parcela produtiva com laudo técnico."
        ),
    },
]


# ─── 4. ENGINE PRINCIPAL ──────────────────────────────────────────────────────

def _parse_value(raw) -> float:
    try:
        s = str(raw).strip().replace("R$", "").replace(" ", "")
        if not s or s.lower() in ("nan", "none", ""):
            return 0.0
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _normalize_cols(df: pd.DataFrame) -> dict:
    col_map: dict[str, str] = {}
    for col in df.columns:
        cl = col.lower().strip()
        if any(k in cl for k in ["conta", "código", "codigo", "cód.", "cod."]):
            col_map.setdefault("conta", col)
        elif any(k in cl for k in ["descriç", "descricao", "description", "nome da conta", "nome"]):
            col_map.setdefault("descricao", col)
        elif any(k in cl for k in ["valor", "saldo", "value", "amount", "montante"]):
            col_map.setdefault("valor", col)
    cols = list(df.columns)
    col_map.setdefault("conta", cols[0] if cols else "")
    col_map.setdefault("descricao", cols[1] if len(cols) > 1 else cols[0] if cols else "")
    col_map.setdefault("valor", cols[2] if len(cols) > 2 else "")
    return col_map


def _find_rule(conta: str, descricao: str, nature: str) -> dict | None:
    text = f"{conta} {descricao}".lower()
    for rule in RULES:
        filters = rule.get("nature_filter", [])
        if filters and nature not in filters:
            continue
        for kw in rule["keywords"]:
            if kw.lower() in text:
                return rule
    return None


def analyze_balancete(df: pd.DataFrame, col_map: dict | None = None, nivel_filtro: int = 5) -> dict:
    if col_map is None:
        col_map = _normalize_cols(df)

    col_nivel = col_map.get("nivel")  # None se não informado
    usar_nivel = col_nivel and col_nivel in df.columns

    raw_rows: list[dict] = []
    skipped_nivel: list[dict] = []

    for _, row in df.iterrows():
        conta     = str(row.get(col_map["conta"],    "")).strip()
        descricao = str(row.get(col_map["descricao"], "")).strip()
        valor     = _parse_value(row.get(col_map["valor"], 0))

        if not conta and not descricao:
            continue

        # Filtro de nível: se a coluna existe, inclui somente o nível solicitado
        if usar_nivel:
            nivel_raw = str(row.get(col_nivel, "")).strip()
            try:
                nivel_val = int(float(nivel_raw))
            except (ValueError, TypeError):
                nivel_val = None

            if nivel_val != nivel_filtro:
                skipped_nivel.append({
                    "conta": conta, "descricao": descricao, "valor": valor,
                    "nivel": nivel_val,
                })
                continue

        raw_rows.append({"_raw_conta": conta, "_raw_desc": descricao, "_raw_valor": valor})

    # Com filtro de nível ativo, toda conta restante já é analítica — skip detecção sintética
    if usar_nivel:
        synthetic_codes: set[str] = set()
    else:
        synthetic_codes = mark_synthetic_accounts(raw_rows)

    rows_out: list[dict] = []
    for r in raw_rows:
        conta = r["_raw_conta"]
        descricao = r["_raw_desc"]
        valor = r["_raw_valor"]

        code = _get_code_prefix(conta)
        is_synthetic = code in synthetic_codes

        nature_info = detect_account_nature(conta, descricao)
        nature = nature_info["nature"]

        # ── Contas sintéticas ────────────────────────────────────────────────
        if is_synthetic:
            rows_out.append({
                "conta": conta, "descricao": descricao, "valor": valor,
                "account_type": "SINTÉTICA",
                "nature": nature,
                "cpc_ref": nature_info["cpc_ref"],
                "category": "Conta Sintética — Agrupamento",
                "credit_type": "N/A",
                "eligible": "N/A",
                "risk": "N/A",
                "legal_basis": "Conta de agrupamento — não possui movimentação própria",
                "comment": (
                    "Conta sintética (grupo): consolida saldos das contas analíticas subordinadas. "
                    "Não deve ser utilizada como base para apuração de crédito — o crédito deve ser "
                    "apurado nas contas analíticas individuais para garantir precisão e auditabilidade."
                ),
                "credit_value": 0.0,
                "cpc_tax_ref": "",
                "law_type": "N/A",
            })
            continue

        # ── Contas de balanço ────────────────────────────────────────────────
        if not nature_info["analyze"]:
            rows_out.append({
                "conta": conta, "descricao": descricao, "valor": valor,
                "account_type": "BALANÇO",
                "nature": nature,
                "cpc_ref": nature_info["cpc_ref"],
                "category": f"Conta de {nature} — Balanço Patrimonial",
                "credit_type": "N/A",
                "eligible": "N/A",
                "risk": "N/A",
                "legal_basis": nature_info["skip_reason"],
                "comment": nature_info["skip_reason"],
                "credit_value": 0.0,
                "cpc_tax_ref": "",
                "law_type": "N/A",
            })
            continue

        # ── Contas analíticas de resultado ───────────────────────────────────
        rule = _find_rule(conta, descricao, nature)

        if rule is None:
            # Conta de resultado não classificada pelo motor — analise manual
            row_out = {
                "conta": conta, "descricao": descricao, "valor": valor,
                "account_type": "ANALÍTICA",
                "nature": nature,
                "cpc_ref": nature_info["cpc_ref"],
                "category": "Não Classificado — Análise Manual Necessária",
                "credit_type": "Indeterminado",
                "eligible": "POSSÍVEL",
                "risk": "MÉDIO",
                "legal_basis": "Requer análise manual — aplicar teste de essencialidade e relevância (SC COSIT 5/2018)",
                "comment": (
                    "Conta de resultado não identificada pelo motor de regras. "
                    "Aplicar manualmente o teste da SC COSIT 5/2018: "
                    "(1) Essencialidade — a ausência desse bem/serviço impede ou paralisa o processo produtivo? "
                    "(2) Relevância — a ausência compromete significativamente a qualidade ou quantidade da produção? "
                    "Se SIM a qualquer dos critérios, há base para crédito sob o Art. 3º, II. "
                    "Documentar a análise com laudo técnico ou parecer jurídico."
                ),
                "credit_value": round(valor * COMBINED_RATE / 100, 2) if valor > 0 else 0.0,
                "cpc_tax_ref": nature_info["cpc_ref"],
                "law_type": "manual",
            }
        else:
            credit_value = 0.0
            if rule["eligible"] in ("SIM", "POSSÍVEL") and valor > 0:
                credit_value = round(valor * COMBINED_RATE / 100, 2)
            row_out = {
                "conta": conta, "descricao": descricao, "valor": valor,
                "account_type": "ANALÍTICA",
                "nature": nature,
                "cpc_ref": nature_info["cpc_ref"],
                "category": rule["category"],
                "credit_type": rule["credit_type"],
                "eligible": rule["eligible"],
                "risk": rule["risk"],
                "legal_basis": rule["legal_basis"],
                "comment": rule["comment"],
                "credit_value": credit_value,
                "cpc_tax_ref": rule.get("cpc_tax_ref", ""),
                "law_type": rule.get("law_type", ""),
            }

        rows_out.append(row_out)

    # ── Agregações (somente contas analíticas de resultado) ──────────────────
    result_rows = [r for r in rows_out if r["account_type"] == "ANALÍTICA" and r["eligible"] != "N/A"]
    bs_rows = [r for r in rows_out if r["account_type"] == "BALANÇO"]
    syn_rows = [r for r in rows_out if r["account_type"] == "SINTÉTICA"]

    eligible_sim = [r for r in result_rows if r["eligible"] == "SIM" and r["valor"] > 0]
    eligible_pos = [r for r in result_rows if r["eligible"] == "POSSÍVEL" and r["valor"] > 0]

    total_analisado = sum(r["valor"] for r in result_rows if r["valor"] > 0)
    total_sim = sum(r["valor"] for r in eligible_sim)
    total_pos = sum(r["valor"] for r in eligible_pos)
    credito_certo = sum(r["credit_value"] for r in eligible_sim)
    credito_possivel = sum(r["credit_value"] for r in eligible_pos)
    pct = round(total_sim / total_analisado * 100, 1) if total_analisado else 0

    by_cat: dict = {}
    for r in result_rows:
        cat = r["category"]
        if cat not in by_cat:
            by_cat[cat] = {"valor": 0.0, "credito": 0.0, "eligible": r["eligible"], "risk": r["risk"]}
        if r["valor"] > 0:
            by_cat[cat]["valor"] += r["valor"]
        by_cat[cat]["credito"] += r["credit_value"]

    by_type: dict = {}
    for r in [r for r in result_rows if r["eligible"] in ("SIM", "POSSÍVEL") and r["valor"] > 0]:
        ct = r["credit_type"]
        if ct not in by_type:
            by_type[ct] = {"valor": 0.0, "credito": 0.0}
        by_type[ct]["valor"] += r["valor"]
        by_type[ct]["credito"] += r["credit_value"]

    risk_dist = {"ALTO": 0.0, "MÉDIO": 0.0, "BAIXO": 0.0}
    for r in result_rows:
        if r["risk"] in risk_dist:
            risk_dist[r["risk"]] += abs(r["valor"])

    top_oportunidades = sorted(
        [r for r in result_rows if r["eligible"] in ("SIM", "POSSÍVEL") and r["credit_value"] > 0],
        key=lambda x: x["credit_value"], reverse=True,
    )[:10]

    top_riscos = sorted(
        [r for r in result_rows if r["risk"] == "ALTO" and abs(r["valor"]) > 0],
        key=lambda x: abs(x["valor"]), reverse=True,
    )[:10]

    return {
        "rows": rows_out,
        "summary": {
            "total_analisado": round(total_analisado, 2),
            "total_elegivel": round(total_sim, 2),
            "total_possivel": round(total_pos, 2),
            "credito_certo": round(credito_certo, 2),
            "credito_possivel": round(credito_possivel, 2),
            "credito_total_potencial": round(credito_certo + credito_possivel, 2),
            "pct_aproveitamento": pct,
            "total_linhas": len(rows_out) + len(skipped_nivel),
            "total_analiticas": len(result_rows),
            "total_balanço": len(bs_rows),
            "total_sinteticas": len(syn_rows),
            "total_outros_niveis": len(skipped_nivel),
            "nivel_filtro": nivel_filtro if usar_nivel else None,
            "pis_rate": PIS_RATE,
            "cofins_rate": COFINS_RATE,
        },
        "by_category": by_cat,
        "by_credit_type": by_type,
        "risk_distribution": {k: round(v, 2) for k, v in risk_dist.items()},
        "top_oportunidades": top_oportunidades,
        "top_riscos": top_riscos,
    }


# ─── 5. ROTAS FLASK ───────────────────────────────────────────────────────────

@bp.route("/preview", methods=["POST"])
def preview():
    """Retorna colunas e primeiras linhas do Excel para o usuário confirmar o mapeamento."""
    if "file" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Apenas arquivos .xlsx são aceitos"}), 400

    try:
        # Tenta ler com e sem cabeçalho para mostrar ao usuário
        file_bytes = f.read()
        import io as _io
        df = pd.read_excel(_io.BytesIO(file_bytes), dtype=str, nrows=8)
    except Exception as e:
        return jsonify({"error": f"Erro ao ler planilha: {e}"}), 400

    if df.empty:
        return jsonify({"error": "Planilha vazia"}), 400

    columns = list(df.columns)
    preview_rows = []
    for _, row in df.head(5).iterrows():
        preview_rows.append([str(v) if pd.notna(v) else "" for v in row])

    # Sugestão automática de mapeamento
    suggestion = {"conta": None, "descricao": None, "valor": None, "nivel": None}
    for i, col in enumerate(columns):
        cl = str(col).lower().strip()
        if suggestion["conta"] is None and any(k in cl for k in ["conta", "código", "codigo", "cód", "cod"]):
            suggestion["conta"] = i
        if suggestion["descricao"] is None and any(k in cl for k in ["descriç", "descricao", "description", "nome", "histórico", "historico"]):
            suggestion["descricao"] = i
        if suggestion["valor"] is None and any(k in cl for k in ["valor", "saldo", "value", "amount", "montante", "débito", "credito", "crédito"]):
            suggestion["valor"] = i
        if suggestion["nivel"] is None and any(k in cl for k in ["nível", "nivel", "level", "grau", "hierarquia"]):
            suggestion["nivel"] = i

    # Fallback: se não achou por nome, sugere por posição
    if suggestion["conta"] is None and len(columns) > 0:
        suggestion["conta"] = 0
    if suggestion["descricao"] is None and len(columns) > 1:
        suggestion["descricao"] = 1
    if suggestion["valor"] is None and len(columns) > 2:
        suggestion["valor"] = 2
    # nivel não tem fallback — campo opcional

    # Salva os bytes para o upload posterior
    preview_id = str(uuid.uuid4())
    _analyses[f"preview_{preview_id}"] = {
        "file_bytes": file_bytes,
        "filename": f.filename,
        "columns": columns,
    }

    return jsonify({
        "preview_id": preview_id,
        "filename": f.filename,
        "columns": columns,
        "preview_rows": preview_rows,
        "suggestion": suggestion,
    })


@bp.route("/upload", methods=["POST"])
def upload():
    """Executa a análise com base no preview_id e mapeamento de colunas confirmado pelo usuário."""
    data = request.get_json(force=True)
    preview_id = data.get("preview_id")
    col_conta    = data.get("col_conta")      # índice (int) ou nome (str)
    col_descricao = data.get("col_descricao")
    col_valor    = data.get("col_valor")
    col_nivel    = data.get("col_nivel")     # opcional
    nivel_filtro = int(data.get("nivel_filtro", 5))

    if not preview_id:
        return jsonify({"error": "preview_id ausente"}), 400

    cache_key = f"preview_{preview_id}"
    if cache_key not in _analyses:
        return jsonify({"error": "Preview expirado. Faça o upload novamente."}), 400

    cached = _analyses.pop(cache_key)
    file_bytes = cached["file_bytes"]
    filename = cached["filename"]
    columns = cached["columns"]

    import io as _io
    try:
        df = pd.read_excel(_io.BytesIO(file_bytes), dtype=str)
    except Exception as e:
        return jsonify({"error": f"Erro ao ler planilha: {e}"}), 400

    # Resolve índice → nome de coluna
    def resolve(ref):
        if ref is None:
            return None
        if isinstance(ref, int) and 0 <= ref < len(columns):
            return columns[ref]
        if isinstance(ref, str) and ref in columns:
            return ref
        try:
            idx = int(ref)
            if 0 <= idx < len(columns):
                return columns[idx]
        except (ValueError, TypeError):
            pass
        return None

    col_map = {
        "conta":    resolve(col_conta)     or columns[0],
        "descricao": resolve(col_descricao) or (columns[1] if len(columns) > 1 else columns[0]),
        "valor":    resolve(col_valor)     or (columns[2] if len(columns) > 2 else columns[0]),
        "nivel":    resolve(col_nivel),   # pode ser None se não informado
    }

    if df.empty:
        return jsonify({"error": "Planilha vazia"}), 400

    try:
        result = analyze_balancete(df, col_map=col_map, nivel_filtro=nivel_filtro)
    except Exception as e:
        return jsonify({"error": f"Erro na análise: {e}"}), 500

    analysis_id = str(uuid.uuid4())
    result.update({"analysis_id": analysis_id, "filename": filename,
                   "created_at": datetime.now().isoformat()})
    _analyses[analysis_id] = result

    return jsonify({
        "analysis_id": analysis_id,
        "filename": filename,
        "summary": result["summary"],
        "by_category": result["by_category"],
        "by_credit_type": result["by_credit_type"],
        "risk_distribution": result["risk_distribution"],
        "top_oportunidades": result["top_oportunidades"],
        "top_riscos": result["top_riscos"],
        "rows": result["rows"],
    })


@bp.route("/download/excel/<analysis_id>")
def download_excel(analysis_id):
    if analysis_id not in _analyses:
        return jsonify({"error": "Análise não encontrada. Refaça o upload."}), 404

    result = _analyses[analysis_id]
    rows = result["rows"]
    summary = result["summary"]
    filename = result.get("filename", "balancete")

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Aba 1 — Apenas analíticas de resultado
        result_rows = [r for r in rows if r["account_type"] == "ANALÍTICA" and r["eligible"] != "N/A"]
        df_det = pd.DataFrame([{
            "Conta Contábil": r["conta"],
            "Descrição": r["descricao"],
            "Natureza (CPC)": r["nature"],
            "CPC de Referência": r["cpc_ref"],
            "Valor (R$)": r["valor"],
            "Categoria Tributária": r["category"],
            "Elegibilidade PIS/COFINS": r["eligible"],
            "Tipo de Crédito": r["credit_type"],
            "Crédito Potencial 9,25% (R$)": r["credit_value"],
            "PIS 1,65% (R$)": round(r["credit_value"] * PIS_RATE / COMBINED_RATE, 2),
            "COFINS 7,6% (R$)": round(r["credit_value"] * COFINS_RATE / COMBINED_RATE, 2),
            "Nível de Risco": r["risk"],
            "Base Legal": r["legal_basis"],
            "Fundamentação Técnica (SC COSIT 5/2018)": r["comment"],
        } for r in result_rows])
        df_det.to_excel(writer, sheet_name="Análise PIS-COFINS", index=False)

        # Aba 2 — Contas excluídas (balanço + sintéticas)
        excl = [r for r in rows if r["account_type"] in ("BALANÇO", "SINTÉTICA")]
        if excl:
            df_excl = pd.DataFrame([{
                "Conta": r["conta"], "Descrição": r["descricao"],
                "Tipo": r["account_type"], "Motivo da Exclusão": r["comment"],
            } for r in excl])
            df_excl.to_excel(writer, sheet_name="Excluídas da Análise", index=False)

        # Aba 3 — Resumo executivo
        ws_sum = writer.book.create_sheet("Resumo Executivo")
        pis_certo = round(summary["credito_certo"] * PIS_RATE / COMBINED_RATE, 2)
        cof_certo = round(summary["credito_certo"] * COFINS_RATE / COMBINED_RATE, 2)
        for rd in [
            ["ANÁLISE TRIBUTÁRIA PIS/COFINS — REGIME NÃO CUMULATIVO"],
            ["Fundamentação: SC COSIT 5/2018 + STJ Tema 779 + Leis 10.637/2002 e 10.833/2003"],
            ["Arquivo:", filename], ["Gerado em:", datetime.now().strftime("%d/%m/%Y %H:%M")], [""],
            ["ESCOPO DA ANÁLISE"],
            ["Total de contas no balancete:", summary["total_linhas"]],
            ["Contas analíticas de resultado analisadas:", summary["total_analiticas"]],
            ["Contas de balanço excluídas:", summary["total_balanço"]],
            ["Contas sintéticas excluídas:", summary["total_sinteticas"]], [""],
            ["RESULTADO"], ["Total da base analisada (R$):", summary["total_analisado"]],
            ["Base confirmada para crédito (R$):", summary["total_elegivel"]],
            ["Base com potencial de crédito (R$):", summary["total_possivel"]],
            ["% de aproveitamento confirmado:", f"{summary['pct_aproveitamento']}%"], [""],
            ["CRÉDITOS POTENCIAIS"],
            ["PIS 1,65% — confirmado (R$):", pis_certo],
            ["COFINS 7,6% — confirmado (R$):", cof_certo],
            ["Total crédito confirmado (R$):", summary["credito_certo"]],
            ["Total crédito potencial (R$):", summary["credito_total_potencial"]], [""],
            ["AVISO: Esta análise é indicativa. Validação por especialista tributário é obrigatória.", ""],
        ]:
            ws_sum.append(rd)
        ws_sum.column_dimensions["A"].width = 55
        ws_sum.column_dimensions["B"].width = 22

        # Formatação aba principal
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        ws = writer.sheets["Análise PIS-COFINS"]
        hf = PatternFill("solid", start_color="1B3A5C")
        hfont = Font(bold=True, color="FFFFFF", size=10)
        for c in range(1, len(df_det.columns) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = hf; cell.font = hfont
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cmap = {"SIM": PatternFill("solid", start_color="C6EFCE"),
                "NÃO": PatternFill("solid", start_color="FFC7CE"),
                "POSSÍVEL": PatternFill("solid", start_color="FFEB9C")}
        for ri, row in enumerate(result_rows, start=2):
            fill = cmap.get(row["eligible"], PatternFill("solid", start_color="D9D9D9"))
            for c in range(1, len(df_det.columns) + 1):
                ws.cell(row=ri, column=c).fill = fill
        widths = [18, 45, 14, 40, 16, 32, 16, 30, 20, 16, 16, 12, 65, 100]
        for i, w in enumerate(widths[:len(df_det.columns)], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    output.seek(0)
    dl_name = f"analise_piscofins_{analysis_id[:8]}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=dl_name,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/health")
def health():
    return jsonify({"status": "ok", "rules": len(RULES), "analyses_cached": len(_analyses)})

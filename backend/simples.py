"""
Simples Nacional — Consulta Histórica em Lote
Registra blueprint Flask: /api/simples/...
"""

import asyncio
import json
import os
import re
import sqlite3
import threading
import uuid
from datetime import date, datetime
from pathlib import Path

import aiohttp
import pandas as pd
from flask import Blueprint, jsonify, request, send_file

bp = Blueprint("simples", __name__, url_prefix="/api/simples")

_BASE = Path(__file__).parent
_DB   = str(_BASE / "simples.db")
_UPL  = _BASE / "uploads"
_RES  = _BASE / "results"
_UPL.mkdir(exist_ok=True)
_RES.mkdir(exist_ok=True)

ANO_FIM    = date.today().year
ANO_INICIO = ANO_FIM - 5   # últimos 60 meses

# ── DB ────────────────────────────────────────────────────────────────────────

def _init_db():
    c = sqlite3.connect(_DB)
    c.executescript("""
        CREATE TABLE IF NOT EXISTS sn_jobs (
            id TEXT PRIMARY KEY,
            status TEXT DEFAULT 'aguardando',
            total INTEGER DEFAULT 0,
            processados INTEGER DEFAULT 0,
            criado_em TEXT,
            arquivo_saida TEXT,
            erro TEXT
        );
        CREATE TABLE IF NOT EXISTS sn_resultados (
            job_id TEXT,
            cnpj TEXT,
            razao_social TEXT,
            situacao_atual TEXT,
            ja_foi_optante TEXT,
            nunca_foi_optante TEXT,
            data_entrada TEXT,
            data_exclusao TEXT,
            mei TEXT,
            observacoes TEXT,
            historico_json TEXT,
            PRIMARY KEY (job_id, cnpj)
        );
        CREATE TABLE IF NOT EXISTS sn_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id TEXT,
            ts TEXT,
            msg TEXT
        );
    """)
    c.commit()
    c.close()

_init_db()

# ── Helpers ───────────────────────────────────────────────────────────────────

def _validar_cnpj(raw: str) -> str | None:
    cnpj = re.sub(r"\D", "", str(raw))
    if len(cnpj) != 14 or len(set(cnpj)) == 1:
        return None
    def dig(seq, pesos):
        s = sum(int(d) * p for d, p in zip(seq, pesos))
        r = s % 11
        return 0 if r < 2 else 11 - r
    if dig(cnpj, [5,4,3,2,9,8,7,6,5,4,3,2])   != int(cnpj[12]): return None
    if dig(cnpj, [6,5,4,3,2,9,8,7,6,5,4,3,2]) != int(cnpj[13]): return None
    return cnpj

def _fmt(cnpj: str) -> str:
    return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"

def _parse_date(v) -> date | None:
    if not v: return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y%m%d"):
        try: return datetime.strptime(str(v)[:10], fmt).date()
        except: pass
    return None

def _historico(entrada: date | None, exclusao: date | None) -> dict:
    hist = {}
    for ano in range(ANO_INICIO, ANO_FIM + 1):
        ini, fim = date(ano, 1, 1), date(ano, 12, 31)
        if entrada and entrada <= fim and (not exclusao or exclusao > ini):
            if entrada.year == ano:
                hist[str(ano)] = "Entrou"
            elif exclusao and exclusao.year == ano:
                hist[str(ano)] = "Saiu"
            else:
                hist[str(ano)] = "Sim"
        else:
            hist[str(ano)] = "Não"
    return hist

def _log(job_id, msg):
    c = sqlite3.connect(_DB)
    c.execute("INSERT INTO sn_logs (job_id,ts,msg) VALUES (?,?,?)",
              (job_id, datetime.now().isoformat(), msg))
    c.commit(); c.close()

def _update_job(job_id, **kw):
    c = sqlite3.connect(_DB)
    sets = ", ".join(f"{k}=?" for k in kw)
    c.execute(f"UPDATE sn_jobs SET {sets} WHERE id=?", [*kw.values(), job_id])
    c.commit(); c.close()

# ── API query ─────────────────────────────────────────────────────────────────

async def _brasil_api(session: aiohttp.ClientSession, cnpj: str) -> dict | None:
    url = f"https://brasilapi.com.br/api/cnpj/v1/{cnpj}"
    try:
        async with session.get(url, timeout=aiohttp.ClientTimeout(total=18)) as r:
            if r.status == 200:
                return await r.json(content_type=None)
    except Exception:
        pass
    return None

async def _receita_ws(session: aiohttp.ClientSession, cnpj: str) -> dict | None:
    url = f"https://www.receitaws.com.br/v1/cnpj/{cnpj}"
    try:
        async with session.get(url, timeout=aiohttp.ClientTimeout(total=15)) as r:
            if r.status == 200:
                d = await r.json(content_type=None)
                if d.get("status") != "ERROR":
                    return d
    except Exception:
        pass
    return None

async def _consultar(session: aiohttp.ClientSession, cnpj: str) -> dict | None:
    d = await _brasil_api(session, cnpj)
    if d:
        sim = d.get("simples") or {}
        mei = d.get("mei") or {}
        return {
            "razao_social":        d.get("razao_social", ""),
            "optante_simples":     bool(sim.get("optante")),
            "data_opcao_simples":  sim.get("data_opcao"),
            "data_exclusao_simples": sim.get("data_exclusao"),
            "optante_mei":         bool(mei.get("optante")),
            "fonte": "BrasilAPI",
        }
    d = await _receita_ws(session, cnpj)
    if d:
        return {
            "razao_social":        d.get("nome", ""),
            "optante_simples":     d.get("simples", "Não") == "Sim",
            "data_opcao_simples":  d.get("data_opcao_pelo_simples"),
            "data_exclusao_simples": d.get("data_exclusao_do_simples"),
            "optante_mei":         d.get("mei", "Não") == "Sim",
            "fonte": "ReceitaWS",
        }
    return None

# ── Processing ────────────────────────────────────────────────────────────────

async def _processar_job(job_id: str, cnpjs: list[str]):
    _update_job(job_id, status="processando", total=len(cnpjs))
    _log(job_id, f"▶ Iniciando — {len(cnpjs)} CNPJs")

    resultados = []
    conn = aiohttp.TCPConnector(limit=4)
    headers = {"User-Agent": "Mozilla/5.0 (compatible; SimpesNacionalConsulta/1.0)"}

    async with aiohttp.ClientSession(connector=conn, headers=headers) as session:
        for i, cnpj in enumerate(cnpjs):
            try:
                if i and i % 5 == 0:
                    await asyncio.sleep(1.2)

                _log(job_id, f"[{i+1}/{len(cnpjs)}] {_fmt(cnpj)}")
                info = await _consultar(session, cnpj)

                if info:
                    entrada  = _parse_date(info.get("data_opcao_simples"))
                    exclusao = _parse_date(info.get("data_exclusao_simples"))
                    optante  = info.get("optante_simples", False)

                    if not entrada and not optante:
                        ja_foi, nunca_foi = "Não", "Sim"
                        hist = {str(a): "Não" for a in range(ANO_INICIO, ANO_FIM + 1)}
                    else:
                        ja_foi, nunca_foi = "Sim", "Não"
                        hist = _historico(entrada, exclusao)

                    if optante:
                        situacao = "Optante"
                    elif ja_foi == "Sim":
                        situacao = "Excluída"
                    else:
                        situacao = "Nunca Optante"

                    obs = []
                    if info.get("optante_mei"):
                        obs.append("Optante MEI")
                    if exclusao and (date.today() - exclusao).days <= 365:
                        obs.append("⚠️ Exclusão recente (< 1 ano)")
                    if not info.get("data_opcao_simples") and optante:
                        obs.append("Data de opção não disponível")

                    res = {
                        "cnpj":            _fmt(cnpj),
                        "razao_social":    info.get("razao_social", ""),
                        "situacao_atual":  situacao,
                        "ja_foi_optante":  ja_foi,
                        "nunca_foi_optante": nunca_foi,
                        "data_entrada":    entrada.strftime("%d/%m/%Y") if entrada else "",
                        "data_exclusao":   exclusao.strftime("%d/%m/%Y") if exclusao else "",
                        "mei":             "Sim" if info.get("optante_mei") else "Não",
                        "observacoes":     "; ".join(obs),
                        "historico":       hist,
                        "fonte":           info.get("fonte", ""),
                    }
                else:
                    hist = {str(a): "—" for a in range(ANO_INICIO, ANO_FIM + 1)}
                    res = {
                        "cnpj": _fmt(cnpj), "razao_social": "",
                        "situacao_atual": "Erro na consulta",
                        "ja_foi_optante": "", "nunca_foi_optante": "",
                        "data_entrada": "", "data_exclusao": "",
                        "mei": "", "observacoes": "CNPJ não encontrado ou API indisponível",
                        "historico": hist, "fonte": "",
                    }

                resultados.append(res)

                db = sqlite3.connect(_DB)
                db.execute(
                    """INSERT OR REPLACE INTO sn_resultados
                    (job_id,cnpj,razao_social,situacao_atual,ja_foi_optante,nunca_foi_optante,
                     data_entrada,data_exclusao,mei,observacoes,historico_json)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                    (job_id, res["cnpj"], res["razao_social"], res["situacao_atual"],
                     res["ja_foi_optante"], res["nunca_foi_optante"],
                     res["data_entrada"], res["data_exclusao"],
                     res["mei"], res["observacoes"], json.dumps(res["historico"])),
                )
                db.commit(); db.close()
                _update_job(job_id, processados=i + 1)

            except Exception as exc:
                _log(job_id, f"⚠ Erro {cnpj}: {exc}")

    try:
        path = _gerar_excel(job_id, resultados)
        _update_job(job_id, status="concluido", arquivo_saida=str(path))
        _log(job_id, f"✅ Concluído — {len(resultados)} CNPJs · Excel disponível para download")
    except Exception as exc:
        _update_job(job_id, status="erro", erro=str(exc))
        _log(job_id, f"❌ Erro ao gerar Excel: {exc}")


def _gerar_excel(job_id: str, resultados: list[dict]) -> Path:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    anos = [str(a) for a in range(ANO_INICIO, ANO_FIM + 1)]
    rows = []
    for r in resultados:
        row = {
            "CNPJ": r["cnpj"], "Razão Social": r["razao_social"],
            "Situação Atual": r["situacao_atual"],
            "Já foi Optante?": r["ja_foi_optante"],
            "Nunca foi Optante?": r["nunca_foi_optante"],
            "Data de Entrada": r["data_entrada"],
            "Data de Exclusão": r["data_exclusao"],
            "Optante MEI": r["mei"], "Observações": r["observacoes"],
        }
        for ano in anos:
            row[ano] = r.get("historico", {}).get(ano, "")
        rows.append(row)

    df = pd.DataFrame(rows)
    total = len(resultados)
    cnt = lambda s: sum(1 for r in resultados if r["situacao_atual"] == s)
    optantes  = cnt("Optante")
    excluidas = cnt("Excluída")
    nunca     = cnt("Nunca Optante")
    erros     = cnt("Erro na consulta")

    pct = lambda n: f"{n/total*100:.1f}%" if total else "0%"
    df_sum = pd.DataFrame({
        "Métrica": ["Total Consultado","Optantes Atuais","Excluídas","Nunca Optantes","Erros na Consulta"],
        "Quantidade": [total, optantes, excluidas, nunca, erros],
        "Percentual": ["100%", pct(optantes), pct(excluidas), pct(nunca), pct(erros)],
    })

    path = _RES / f"simples_{job_id}.xlsx"

    with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Dados Consolidados", index=False)
        df_sum.to_excel(writer, sheet_name="Resumo Executivo", index=False)

        wb = writer.book
        ws = writer.sheets["Dados Consolidados"]

        hdr_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True, size=10)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        fills = {
            "Optante":       PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
            "Excluída":      PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
            "Nunca Optante": PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
        }
        yr_fills = {
            "Sim":    (PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"), Font(color="166534", size=9)),
            "Entrou": (PatternFill(start_color="BBF7D0", end_color="BBF7D0", fill_type="solid"), Font(color="15803D", bold=True, size=9)),
            "Saiu":   (PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"), Font(color="B91C1C", bold=True, size=9)),
            "Não":    (PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"), Font(color="9CA3AF", size=9)),
        }

        for ridx in range(2, len(rows) + 2):
            sc = ws.cell(row=ridx, column=3)
            if sc.value in fills:
                sc.fill = fills[sc.value]
            for cidx, _ in enumerate(anos, start=10):
                cell = ws.cell(row=ridx, column=cidx)
                if cell.value in yr_fills:
                    cell.fill, cell.font = yr_fills[cell.value]
                cell.alignment = Alignment(horizontal="center")

        for i, col in enumerate(df.columns, 1):
            w = max(len(str(col)), 8)
            ws.column_dimensions[get_column_letter(i)].width = min(w + 2, 30) if i <= 9 else 6

        ws.freeze_panes = "J2"

        ws2 = writer.sheets["Resumo Executivo"]
        for cell in ws2[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font

    return path

# ── Flask Routes ──────────────────────────────────────────────────────────────

@bp.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400
    f = request.files["file"]
    ext = (f.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "csv"):
        return jsonify({"error": "Formato inválido. Use .xlsx ou .csv"}), 400

    job_id = str(uuid.uuid4())
    save = _UPL / f"{job_id}.{ext}"
    f.save(str(save))

    try:
        df = pd.read_excel(str(save), dtype=str) if ext == "xlsx" else pd.read_csv(str(save), dtype=str)

        cnpj_col = next((c for c in df.columns if "cnpj" in c.lower()), None)
        if cnpj_col is None:
            return jsonify({"error": "Coluna CNPJ não encontrada. Nomeie a coluna como 'CNPJ'."}), 400

        validos, invalidos, seen = [], [], set()
        for raw in df[cnpj_col].dropna():
            v = _validar_cnpj(str(raw))
            if v and v not in seen:
                validos.append(v); seen.add(v)
            elif v in seen:
                invalidos.append({"valor": raw, "motivo": "Duplicado"})
            else:
                invalidos.append({"valor": raw, "motivo": "CNPJ inválido"})

        if not validos:
            return jsonify({"error": "Nenhum CNPJ válido na planilha"}), 400

        db = sqlite3.connect(_DB)
        db.execute(
            "INSERT INTO sn_jobs (id,status,total,processados,criado_em) VALUES (?,?,?,?,?)",
            (job_id, "aguardando", len(validos), 0, datetime.now().isoformat()),
        )
        db.commit(); db.close()

        with open(str(_UPL / f"{job_id}_cnpjs.json"), "w") as fp:
            json.dump(validos, fp)

        secs = len(validos) * 2
        mins = secs // 60
        tempo = f"~{mins} min" if mins >= 1 else f"~{secs}s"
        return jsonify({
            "job_id": job_id, "total": len(validos),
            "invalidos": len(invalidos), "log_invalidos": invalidos[:20],
            "tempo_estimado": tempo,
        })

    except Exception as exc:
        return jsonify({"error": f"Erro ao ler planilha: {exc}"}), 500


@bp.route("/processar/<job_id>", methods=["POST"])
def processar(job_id):
    p = _UPL / f"{job_id}_cnpjs.json"
    if not p.exists():
        return jsonify({"error": "Job não encontrado"}), 404
    cnpjs = json.loads(p.read_text())

    def _run():
        asyncio.run(_processar_job(job_id, cnpjs))

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True})


@bp.route("/status/<job_id>")
def status(job_id):
    db = sqlite3.connect(_DB)
    row = db.execute("SELECT * FROM sn_jobs WHERE id=?", (job_id,)).fetchone()
    if not row:
        db.close(); return jsonify({"error": "Não encontrado"}), 404
    cols = ["id","status","total","processados","criado_em","arquivo_saida","erro"]
    job  = dict(zip(cols, row))
    logs = db.execute(
        "SELECT ts,msg FROM sn_logs WHERE job_id=? ORDER BY id DESC LIMIT 30", (job_id,)
    ).fetchall()
    db.close()
    job["logs"] = [{"ts": l[0], "msg": l[1]} for l in logs]
    job["pct"]  = round(job["processados"] / job["total"] * 100) if job["total"] else 0
    return jsonify(job)


@bp.route("/download/<job_id>")
def download(job_id):
    db = sqlite3.connect(_DB)
    row = db.execute("SELECT arquivo_saida FROM sn_jobs WHERE id=?", (job_id,)).fetchone()
    db.close()
    if not row or not row[0] or not Path(row[0]).exists():
        return jsonify({"error": "Arquivo não disponível"}), 404
    nome = f"simples_nacional_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(row[0], as_attachment=True, download_name=nome)


@bp.route("/rows/<job_id>")
def rows(job_id):
    db = sqlite3.connect(_DB)
    rs = db.execute(
        "SELECT cnpj,razao_social,situacao_atual,ja_foi_optante,nunca_foi_optante,"
        "data_entrada,data_exclusao,mei,observacoes,historico_json "
        "FROM sn_resultados WHERE job_id=?", (job_id,)
    ).fetchall()
    db.close()
    cols = ["cnpj","razao_social","situacao_atual","ja_foi_optante","nunca_foi_optante",
            "data_entrada","data_exclusao","mei","observacoes","historico_json"]
    resultado = []
    for r in rs:
        d = dict(zip(cols, r))
        try:
            d["historico"] = json.loads(d.pop("historico_json") or "{}")
        except Exception:
            d["historico"] = {}
        resultado.append(d)
    return jsonify(resultado)


@bp.route("/jobs")
def listar_jobs():
    db = sqlite3.connect(_DB)
    rows = db.execute(
        "SELECT id,status,total,processados,criado_em FROM sn_jobs ORDER BY criado_em DESC LIMIT 20"
    ).fetchall()
    db.close()
    return jsonify([{"id":r[0],"status":r[1],"total":r[2],"processados":r[3],"criado_em":r[4]} for r in rows])

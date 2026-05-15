"""
DUE Consulta Backend - Servidor Flask com automação Playwright
Execução: python main.py
"""

import asyncio
import json
import math
import os
import re
import sqlite3
import threading
import time
import uuid
from collections import deque
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=False)

@app.after_request
def add_cors_headers(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    response.headers["Access-Control-Allow-Private-Network"] = "true"
    return response

_BASE_DIR = Path(__file__).parent
DB_PATH = str(_BASE_DIR / "consultas.db")
UPLOAD_DIR = _BASE_DIR / "uploads"
RESULTS_DIR = _BASE_DIR / "results"
UPLOAD_DIR.mkdir(exist_ok=True)
RESULTS_DIR.mkdir(exist_ok=True)

ANTICAPTCHA_KEY = "6d73ae3890ea23b5d54c6240355586c2"
PORTAL_URL = "https://portalunico.siscomex.gov.br/due/x/#/consulta/consulta-filtro?perfil=publico"
PROXY_FILE = _BASE_DIR / "proxies.txt"


def _proxy_para_playwright(proxy_url: str | None) -> dict | None:
    """Converte 'http://user:pass@host:port' para o formato de proxy do Playwright."""
    if not proxy_url:
        return None
    from urllib.parse import urlparse
    p = urlparse(proxy_url)
    cfg: dict = {"server": f"{p.scheme}://{p.hostname}:{p.port}"}
    if p.username:
        cfg["username"] = p.username
    if p.password:
        cfg["password"] = p.password
    return cfg

# ─── Banco de dados ─────────────────────────────────────────────────────────

def init_db():
    conn = sqlite3.connect(DB_PATH, timeout=15)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS jobs (
            id TEXT PRIMARY KEY,
            status TEXT,
            total INTEGER,
            processed INTEGER,
            created_at TEXT,
            finished_at TEXT,
            input_file TEXT,
            output_file TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS resultados (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id TEXT,
            chave_nfe TEXT,
            status_nfe TEXT,
            numero_due TEXT,
            data_due TEXT,
            status_due TEXT,
            observacao TEXT,
            consultado_em TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id TEXT,
            nivel TEXT,
            mensagem TEXT,
            criado_em TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS perdcomp_projetos (
            id TEXT PRIMARY KEY,
            nome TEXT,
            empresa TEXT,
            cnpj TEXT,
            competencia_ini TEXT,
            competencia_fim TEXT,
            pasta_pdfs TEXT,
            total_docs INTEGER,
            total_pers INTEGER,
            total_dcomps INTEGER,
            criado_em TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS proxy_limites (
            proxy TEXT PRIMARY KEY,
            reset_em TEXT,
            ultimo_uso TEXT
        )
    """)
    conn.commit()
    conn.close()


def _proxy_key(proxy: str | None) -> str:
    return proxy or "direto"


def registrar_rate_limit_proxy(proxy: str | None, reset_em: str):
    """Persiste o horário de reset do rate limit de um proxy no banco."""
    conn = sqlite3.connect(DB_PATH, timeout=15)
    conn.execute(
        "INSERT OR REPLACE INTO proxy_limites (proxy, reset_em, ultimo_uso) VALUES (?, ?, ?)",
        (_proxy_key(proxy), reset_em, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()


def registrar_uso_proxy(proxy: str | None):
    """Atualiza o timestamp de último uso de um proxy."""
    conn = sqlite3.connect(DB_PATH, timeout=15)
    conn.execute(
        "INSERT OR REPLACE INTO proxy_limites (proxy, reset_em, ultimo_uso) VALUES (?, COALESCE((SELECT reset_em FROM proxy_limites WHERE proxy=?), NULL), ?)",
        (_proxy_key(proxy), _proxy_key(proxy), datetime.now().isoformat())
    )
    conn.commit()
    conn.close()


def ordenar_proxies_por_disponibilidade(proxies: list[str | None]) -> list[str | None]:
    """
    Ordena proxies pelo estado de rate limit persistido:
      0 — nunca usados (mais frescos)
      1 — limite já expirou (disponíveis)
      2 — ainda em cooldown (evitar)
    Dentro de cada grupo, ordena pelo reset_em mais antigo primeiro.
    """
    conn = sqlite3.connect(DB_PATH, timeout=15)
    rows = {row[0]: row[1] for row in conn.execute(
        "SELECT proxy, reset_em FROM proxy_limites"
    ).fetchall()}
    conn.close()

    agora = datetime.now().isoformat()

    def prioridade(proxy):
        key = _proxy_key(proxy)
        reset = rows.get(key)
        if not reset:
            return (0, "")        # nunca usou — melhor
        if reset <= agora:
            return (1, reset)     # limite expirado — bom
        return (2, reset)         # ainda em cooldown — pior

    return sorted(proxies, key=prioridade)

def log(job_id, nivel, msg):
    conn = sqlite3.connect(DB_PATH, timeout=15)
    conn.execute(
        "INSERT INTO logs (job_id, nivel, mensagem, criado_em) VALUES (?, ?, ?, ?)",
        (job_id, nivel, msg, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()
    print(f"[{nivel}] [{job_id[:8]}] {msg}")

def update_job(job_id, **kwargs):
    conn = sqlite3.connect(DB_PATH, timeout=15)
    sets = ", ".join(f"{k} = ?" for k in kwargs)
    vals = list(kwargs.values()) + [job_id]
    conn.execute(f"UPDATE jobs SET {sets} WHERE id = ?", vals)
    conn.commit()
    conn.close()

def salvar_resultado(job_id, chave, status_nfe, numero_due="", data_due="", status_due="", obs=""):
    conn = sqlite3.connect(DB_PATH, timeout=15)
    conn.execute(
        """INSERT INTO resultados 
           (job_id, chave_nfe, status_nfe, numero_due, data_due, status_due, observacao, consultado_em)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (job_id, chave, status_nfe, numero_due, data_due, status_due, obs, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()

# ─── Validação de chave NF-e ─────────────────────────────────────────────────

def validar_chave(chave: str) -> bool:
    chave = re.sub(r'\D', '', str(chave))
    return len(chave) == 44 and chave.isdigit()

def normalizar_chave(chave: str) -> str:
    return re.sub(r'\D', '', str(chave))

# ─── Resolução de CAPTCHA via 2captcha/anticaptcha ──────────────────────────

async def resolver_captcha_anticaptcha(site_key: str, page_url: str, captcha_type: str = "hcaptcha") -> str | None:
    """Resolve hCaptcha ou reCAPTCHA via anti-captcha.com API."""
    task_type = "HCaptchaTaskProxyless" if captcha_type == "hcaptcha" else "NoCaptchaTaskProxyless"
    import aiohttp
    try:
        async with aiohttp.ClientSession() as session:
            payload = {
                "clientKey": ANTICAPTCHA_KEY,
                "task": {
                    "type": task_type,
                    "websiteURL": page_url,
                    "websiteKey": site_key
                }
            }
            async with session.post("https://api.anti-captcha.com/createTask", json=payload) as resp:
                data = await resp.json()
                if data.get("errorId") != 0:
                    print(f"Anti-captcha createTask erro: {data.get('errorDescription')}")
                    return None
                task_id = data["taskId"]

            for _ in range(60):
                await asyncio.sleep(3)
                async with session.post("https://api.anti-captcha.com/getTaskResult",
                                        json={"clientKey": ANTICAPTCHA_KEY, "taskId": task_id}) as resp:
                    result = await resp.json()
                    if result.get("status") == "ready":
                        return result["solution"]["gRecaptchaResponse"]
                    if result.get("errorId") != 0:
                        print(f"Anti-captcha getTaskResult erro: {result.get('errorDescription')}")
                        return None
    except Exception as e:
        print(f"Erro anticaptcha: {e}")
    return None


async def _obter_sitekey(page) -> str | None:
    """Extrai data-sitekey do reCAPTCHA buscando na página principal, frames e objeto JS interno."""
    # 1. Página principal (timeout curto — não aguardar se não existir)
    try:
        sitekey = await page.locator('[data-sitekey]').first.get_attribute("data-sitekey", timeout=2000)
        if sitekey:
            return sitekey
    except Exception:
        pass

    # 2. Frames carregados (reCAPTCHA vive dentro de iframes)
    for frame in page.frames:
        if not frame.url or frame.url == "about:blank":
            continue
        try:
            sitekey = await frame.locator('[data-sitekey]').first.get_attribute("data-sitekey", timeout=1000)
            if sitekey:
                return sitekey
        except Exception:
            continue

    # 3. FrameLocator para iframes não acessíveis diretamente
    for iframe_sel in [
        'iframe[src*="recaptcha"]',
        'iframe[src*="captcha"]',
        'iframe[title*="reCAPTCHA"]',
        'iframe[title*="recaptcha"]',
    ]:
        try:
            sitekey = await page.frame_locator(iframe_sel).locator('[data-sitekey]').first.get_attribute(
                "data-sitekey", timeout=1000
            )
            if sitekey:
                return sitekey
        except Exception:
            continue

    # 4. DOM via JavaScript (captura elementos ocultos ou renderizados por frameworks)
    try:
        sitekey = await page.evaluate(
            "() => { const el = document.querySelector('[data-sitekey]'); return el ? el.getAttribute('data-sitekey') : null; }"
        )
        if sitekey:
            return sitekey
    except Exception:
        pass

    # 5. Extrair sitekey da URL do iframe hCaptcha — o portal Siscomex usa Angular
    #    que não coloca data-sitekey no DOM; o sitekey fica no hash da URL do iframe
    #    ex: https://newassets.hcaptcha.com/.../hcaptcha.html#anchor?sitekey=UUID&...
    for frame in page.frames:
        url = frame.url or ""
        if "hcaptcha.com" in url.lower():
            match = re.search(r'[?&#]sitekey=([0-9a-f-]{20,})', url)
            if match:
                return match.group(1)

    # 6. Sitekey no src do iframe hCaptcha via DOM (antes do frame carregar)
    try:
        sitekey = await page.evaluate("""
            () => {
                for (const f of document.querySelectorAll('iframe')) {
                    const src = f.src || f.getAttribute('src') || '';
                    if (!src.includes('hcaptcha')) continue;
                    const m = src.match(/[?&#]sitekey=([0-9a-f-]{20,})/);
                    if (m) return m[1];
                }
                return null;
            }
        """)
        if sitekey:
            return sitekey
    except Exception:
        pass

    # 7. Objeto interno ___grecaptcha_cfg (reCAPTCHA invisível/v3)
    try:
        sitekey = await page.evaluate("""
            () => {
                try {
                    const cfg = window.___grecaptcha_cfg;
                    if (!cfg || !cfg.clients) return null;
                    for (const client of Object.values(cfg.clients)) {
                        for (const val of Object.values(client)) {
                            if (!val || typeof val !== 'object') continue;
                            if (val.sitekey) return val.sitekey;
                            for (const v2 of Object.values(val)) {
                                if (v2 && typeof v2 === 'object' && v2.sitekey) return v2.sitekey;
                            }
                        }
                    }
                } catch(e) {}
                return null;
            }
        """)
        if sitekey:
            return sitekey
    except Exception:
        pass

    # 8. Sitekey em tags <script> ou atributos da página (fallback final)
    try:
        sitekey = await page.evaluate(r"""
            () => {
                for (const s of document.querySelectorAll('script, [sitekey], [data-sitekey]')) {
                    const attr = s.getAttribute('sitekey') || s.getAttribute('data-sitekey');
                    if (attr && attr.length > 10) return attr;
                    const m = (s.textContent || '').match(/"sitekey"\s*:\s*"([^"]{20,})"/);
                    if (m) return m[1];
                }
                return null;
            }
        """)
        if sitekey:
            return sitekey
    except Exception:
        pass

    return None


async def _injetar_token_captcha(page, token: str, captcha_type: str = "hcaptcha") -> bool:
    """Injeta token de hCaptcha ou reCAPTCHA e aciona callback do formulário."""
    try:
        if captcha_type == "hcaptcha":
            await page.evaluate("""
                (token) => {
                    // hCaptcha: usar native setter para compatibilidade com Angular/React
                    const hc = document.querySelector('textarea[name="h-captcha-response"]');
                    if (hc) {
                        try {
                            const setter = Object.getOwnPropertyDescriptor(
                                window.HTMLTextAreaElement.prototype, 'value'
                            ).set;
                            setter.call(hc, token);
                        } catch(e) {
                            hc.value = token;
                        }
                        hc.dispatchEvent(new Event('input', { bubbles: true }));
                        hc.dispatchEvent(new Event('change', { bubbles: true }));
                    }
                    // Fallback g-recaptcha-response (alguns portais leem os dois)
                    const gc = document.getElementById('g-recaptcha-response');
                    if (gc) { gc.value = token; }
                    // Acionar data-callback se definido
                    try {
                        const el = document.querySelector('[data-callback]');
                        if (el) {
                            const fn = window[el.getAttribute('data-callback')];
                            if (typeof fn === 'function') fn(token);
                        }
                    } catch(e) {}
                }
            """, token)
        else:
            await page.evaluate("""
                (token) => {
                    const resp = document.getElementById('g-recaptcha-response');
                    if (resp) {
                        resp.style.display = 'block';
                        resp.value = token;
                        resp.dispatchEvent(new Event('change', { bubbles: true }));
                    }
                    try {
                        const cfg = window.___grecaptcha_cfg;
                        if (cfg && cfg.clients) {
                            Object.values(cfg.clients).forEach(client => {
                                const entry = Object.values(client).find(
                                    v => v && typeof v === 'object' && typeof v.callback === 'function'
                                );
                                if (entry) entry.callback(token);
                            });
                        }
                    } catch(e) {}
                    try {
                        const el = document.querySelector('[data-callback]');
                        if (el) {
                            const fn = window[el.getAttribute('data-callback')];
                            if (typeof fn === 'function') fn(token);
                        }
                    } catch(e) {}
                }
            """, token)
        return True
    except Exception as e:
        print(f"Erro ao injetar token captcha: {e}")
        return False


async def verificar_e_resolver_captcha(page, job_id: str) -> bool:
    """
    Aguarda captcha inicializar (até 10s), detecta tipo (hCaptcha/reCAPTCHA),
    extrai sitekey, resolve via anti-captcha e injeta token.
    Retorna True se não há captcha ou se foi resolvido. Retorna False se falhou.
    """
    sitekey = None
    tem_indicador = False
    captcha_type = "hcaptcha"  # portal Siscomex usa hCaptcha

    for tentativa in range(10):
        # Detectar tipo pelo URL dos frames
        for frame in page.frames:
            url = (frame.url or "").lower()
            if "hcaptcha.com" in url:
                captcha_type = "hcaptcha"
                tem_indicador = True
                break
            if "recaptcha" in url:
                captcha_type = "recaptcha"
                tem_indicador = True
                break

        # Checar elementos DOM se frames não indicaram
        if not tem_indicador:
            try:
                count = await page.locator(
                    '.h-captcha, .g-recaptcha, [data-sitekey], '
                    'iframe[src*="hcaptcha"], iframe[src*="recaptcha"]'
                ).count()
                if count > 0:
                    tem_indicador = True
            except Exception:
                pass

        if tem_indicador:
            sitekey = await _obter_sitekey(page)
            if sitekey:
                break
            log(job_id, "INFO", f"CAPTCHA carregando... ({tentativa + 1}/10)")
            await asyncio.sleep(1)
        else:
            await asyncio.sleep(1)

    if not tem_indicador and not sitekey:
        return True  # Sem CAPTCHA na página

    if not sitekey:
        log(job_id, "ERROR", "CAPTCHA detectado mas sitekey inacessível após 10s")
        return False

    log(job_id, "INFO", f"CAPTCHA tipo={captcha_type} sitekey={sitekey[:20]}... — resolvendo")
    token = await resolver_captcha_anticaptcha(sitekey, PORTAL_URL, captcha_type)
    if not token:
        log(job_id, "ERROR", "Anti-captcha não retornou token")
        return False

    sucesso = await _injetar_token_captcha(page, token, captcha_type)
    if not sucesso:
        log(job_id, "WARN", "Falha ao injetar token CAPTCHA")
        return False

    log(job_id, "INFO", "Token CAPTCHA injetado — aguardando callback")
    await asyncio.sleep(2)
    return True

# ─── Automação Playwright ─────────────────────────────────────────────────────

async def consultar_chave(page, chave: str, job_id: str) -> dict:
    """Consulta uma chave NF-e no portal Siscomex"""
    resultado = {
        "chave": chave,
        "status_nfe": "Erro",
        "numero_due": "",
        "data_due": "",
        "status_due": "",
        "obs": ""
    }

    try:
        log(job_id, "INFO", f"Consultando chave: {chave[:10]}...{chave[-6:]}")

        # Navegar para o portal
        await page.goto(PORTAL_URL, wait_until="networkidle", timeout=30000)
        await asyncio.sleep(2)

        # Selecionar tipo "NF-e" — o portal inicia em DU-E, precisa trocar o radio
        nfe_clicado = False
        for sel in ['label:has-text("NF-e")', 'label:has-text("NF-E")',
                    'input[value="NFE"]', 'input[value="NF-e"]', 'input[value="NF_E"]']:
            try:
                el = page.locator(sel).first
                if await el.is_visible(timeout=2000):
                    await el.click()
                    nfe_clicado = True
                    log(job_id, "INFO", f"Radio NF-e clicado via: {sel}")
                    await asyncio.sleep(1)
                    break
            except Exception:
                continue
        if not nfe_clicado:
            log(job_id, "WARN", "Radio NF-e não encontrado — tentando continuar no tipo atual")

        # Localizar campo de entrada da chave NF-e
        # IMPORTANTE: excluir o input da barra de busca do navbar usando seletores específicos
        # O campo correto fica no conteúdo principal da página, não no nav
        input_selectors = [
            'input[placeholder*="chave" i]',
            'input[placeholder*="NF-e" i]',
            'input[placeholder*="acesso" i]',
            'input[maxlength="44"]',
            'input[id*="chave" i]',
            'input[name*="chave" i]',
            # Fallback: input de texto dentro do formulário/conteúdo principal (excluindo navbar)
            'main input[type="text"]:not([placeholder*="Buscar" i])',
            'section input[type="text"]:not([placeholder*="Buscar" i])',
            'form input[type="text"]',
            '.filtros input[type="text"]',
        ]

        campo = None
        for sel in input_selectors:
            try:
                el = page.locator(sel).first
                if await el.is_visible(timeout=2000):
                    campo = el
                    log(job_id, "INFO", f"Campo chave encontrado: {sel}")
                    break
            except Exception:
                continue

        if not campo:
            resultado["obs"] = "Campo de entrada não localizado"
            resultado["status_nfe"] = "Erro Layout"
            return resultado

        await campo.click()
        await campo.fill("")
        await campo.type(chave, delay=30)
        await asyncio.sleep(1)

        # Verificar e resolver CAPTCHA se presente
        captcha_ok = await verificar_e_resolver_captcha(page, job_id)
        if not captcha_ok:
            resultado["obs"] = "Falha ao resolver CAPTCHA"
            resultado["status_nfe"] = "Erro CAPTCHA"
            return resultado

        # Submeter consulta
        submit_selectors = [
            'button[type="submit"]',
            'button:has-text("Consultar")',
            'button:has-text("Pesquisar")',
            'button:has-text("Buscar")',
            'input[type="submit"]',
        ]
        submitted = False
        for sel in submit_selectors:
            try:
                btn = page.locator(sel).first
                if await btn.is_visible(timeout=2000):
                    await btn.click()
                    submitted = True
                    break
            except Exception:
                continue

        if not submitted:
            await campo.press("Enter")

        # Aguardar resultado
        await asyncio.sleep(3)
        try:
            await page.wait_for_load_state("networkidle", timeout=15000)
        except PlaywrightTimeout:
            pass

        # Extrair resultado da página
        page_content = await page.content()
        page_text = await page.evaluate("document.body.innerText")

        # Analisar resultado
        texto_lower = page_text.lower()

        if "averbada" in texto_lower and "não averbada" not in texto_lower:
            resultado["status_nfe"] = "Averbada"

            # Tentar extrair número da DUE
            due_patterns = [
                r'DUE[:\s]+(\d{2}/\d+/\d{4}-\d)',
                r'(\d{2}/\d{7,}/\d{4}-\d)',
                r'Número da DUE[:\s]+([^\s\n]+)',
                r'DUE\s*N[°º]?\s*[:\s]+([^\s\n]+)',
            ]
            for pattern in due_patterns:
                match = re.search(pattern, page_text, re.IGNORECASE)
                if match:
                    resultado["numero_due"] = match.group(1).strip()
                    break

            # Tentar extrair data
            date_patterns = [
                r'(\d{2}/\d{2}/\d{4})',
                r'(\d{4}-\d{2}-\d{2})',
            ]
            for pattern in date_patterns:
                match = re.search(pattern, page_text)
                if match:
                    resultado["data_due"] = match.group(1)
                    break

            # Status da DUE
            if "desembaraçada" in texto_lower or "desembaracada" in texto_lower:
                resultado["status_due"] = "Desembaraçada"
            elif "registrada" in texto_lower:
                resultado["status_due"] = "Registrada"
            elif "ativa" in texto_lower:
                resultado["status_due"] = "Ativa"

        elif "não averbada" in texto_lower or "nao averbada" in texto_lower:
            resultado["status_nfe"] = "Não Averbada"

        elif "não encontrada" in texto_lower or "nao encontrada" in texto_lower \
                or "nenhum resultado" in texto_lower or "sem resultado" in texto_lower:
            resultado["status_nfe"] = "Não Encontrada"

        elif any(kw in texto_lower for kw in [
            "pucx-er0204", "desafio do captcha", "você deve responder", "responder o desafio"
        ]):
            resultado["status_nfe"] = "Erro CAPTCHA"
            resultado["obs"] = "Portal retornou erro de CAPTCHA — será retentado"
            log(job_id, "WARN", "Portal exigiu CAPTCHA na resposta — marcando para retry")

        else:
            screenshot_path = f"debug_{chave[:10]}_{int(time.time())}.png"
            try:
                await page.screenshot(path=screenshot_path)
                resultado["obs"] = f"Resultado indefinido — screenshot: {screenshot_path}"
            except Exception:
                resultado["obs"] = "Resultado indefinido"
            resultado["status_nfe"] = "Indefinido"
            log(job_id, "WARN", f"Resultado indefinido para {chave[:10]}... — screenshot: {screenshot_path}")

    except PlaywrightTimeout:
        resultado["status_nfe"] = "Timeout"
        resultado["obs"] = "Timeout na consulta"
        log(job_id, "ERROR", f"Timeout consultando {chave[:10]}...")

    except Exception as e:
        resultado["status_nfe"] = "Erro"
        resultado["obs"] = str(e)[:200]
        log(job_id, "ERROR", f"Erro ao consultar {chave[:10]}...: {e}")

    return resultado

# ─── Worker assíncrono — nova arquitetura: captcha uma vez + API direta ──────

API_BASE = "https://portalunico.siscomex.gov.br/due"
HEADERS_BASE = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Referer": "https://portalunico.siscomex.gov.br/due/x/",
    "X-Requested-With": "XMLHttpRequest",
    "Accept": "application/json, text/plain, */*",
}


async def obter_sessao_com_captcha(job_id: str, proxy: str | None = None) -> tuple[dict, str] | None:
    """
    Abre Chrome (roteado pelo proxy, se fornecido), aguarda resolução do hCaptcha.
    Retorna (cookies, csrf_token) prontos para chamadas HTTP diretas.
    Retorna None em qualquer falha (proxy ruim, timeout, erro de navegação).
    """
    proxy_pw = _proxy_para_playwright(proxy)
    label = proxy or "IP direto"
    csrf_holder = {"token": None}
    captcha_ok = asyncio.Event()

    try:
        async with async_playwright() as p:
            browser = None
            _extra_args = [
                "--disable-blink-features=AutomationControlled",
                "--disable-error-dialogs",
                "--no-first-run",
                "--no-default-browser-check",
                "--disable-infobars",
            ]
            for _channel in ("chrome", "msedge"):
                try:
                    browser = await p.chromium.launch(channel=_channel, headless=False, args=_extra_args)
                    log(job_id, "INFO", f"{'Chrome' if _channel=='chrome' else 'Edge'} aberto [{label}] — resolva o CAPTCHA para iniciar")
                    break
                except Exception:
                    pass
            if browser is None:
                browser = await p.chromium.launch(headless=False, args=_extra_args)
                log(job_id, "INFO", f"Chromium aberto [{label}] — resolva o CAPTCHA para iniciar")

            ctx_kwargs: dict = {"viewport": {"width": 1280, "height": 800}, "user_agent": HEADERS_BASE["User-Agent"]}
            if proxy_pw:
                ctx_kwargs["proxy"] = proxy_pw

            context = await browser.new_context(**ctx_kwargs)
            page = await context.new_page()
            await page.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                Object.defineProperty(navigator, 'languages', { get: () => ['pt-BR', 'pt', 'en-US', 'en'] });
                window.chrome = { runtime: {} };
            """)

            async def rastrear_resposta(response):
                token = response.headers.get("x-csrf-token")
                if token:
                    csrf_holder["token"] = token
                if "portal/proxy/captcha" in response.url and response.status in (200, 204):
                    log(job_id, "INFO", f"CAPTCHA validado! [{label}]")
                    captcha_ok.set()

            page.on("response", rastrear_resposta)

            try:
                await page.goto(PORTAL_URL, wait_until="networkidle", timeout=30000)
            except Exception as nav_err:
                log(job_id, "WARN", f"Falha ao navegar [{label}]: {nav_err} — pulando este proxy")
                await browser.close()
                return None

            log(job_id, "INFO", f"Aguardando resolução do CAPTCHA [{label}] (até 5 minutos)...")
            try:
                await asyncio.wait_for(captcha_ok.wait(), timeout=300)
            except asyncio.TimeoutError:
                log(job_id, "ERROR", f"Timeout: CAPTCHA não resolvido em 5 minutos [{label}]")
                await browser.close()
                return None

            cookies_list = await context.cookies()
            cookies = {c["name"]: c["value"] for c in cookies_list}
            csrf = csrf_holder["token"]

            await asyncio.sleep(3)
            await browser.close()

        return cookies, csrf

    except Exception as ex:
        log(job_id, "ERROR", f"Erro inesperado na sessão [{label}]: {ex}")
        return None


async def inicializar_sessoes(job_id: str, proxies: list[str | None]) -> list[dict]:
    """
    Abre uma sessão autenticada por proxy (um CAPTCHA por proxy), sequencialmente.
    Retorna lista de dicts com proxy, cookies, csrf e limiter próprios.
    """
    sessoes = []
    total = len(proxies)
    for i, proxy in enumerate(proxies):
        label = proxy or "IP direto"
        log(job_id, "INFO", f"Sessão {i+1}/{total} — aguardando CAPTCHA [{label}]")
        update_job(job_id, status=f"iniciando_sessoes:{i+1}/{total}")
        resultado = await obter_sessao_com_captcha(job_id, proxy=proxy)
        if resultado:
            cookies, csrf = resultado
            sessoes.append({
                "proxy":   proxy,
                "cookies": cookies,
                "csrf":    csrf,
                "limiter": RateLimiter(),
                "label":   label,
            })
            log(job_id, "INFO", f"Sessão {i+1}/{total} pronta [{label}]")
        else:
            log(job_id, "WARN", f"Sessão {i+1}/{total} falhou [{label}] — ignorando")
    return sessoes


async def consultar_via_api(
    session,  # aiohttp.ClientSession
    chave: str,
    csrf_token: str,
    job_id: str,
    proxy: str | None = None,
) -> tuple[dict, str]:
    """
    Consulta uma chave NF-e diretamente via HTTP (sem browser).
    Retorna (resultado_dict, novo_csrf_token).
    """
    import aiohttp as _aio

    resultado = {
        "chave": chave,
        "status_nfe": "Erro",
        "numero_due": "",
        "data_due": "",
        "status_due": "",
        "obs": "",
    }

    headers = {**HEADERS_BASE, "X-CSRF-Token": csrf_token}

    try:
        import aiohttp as _aio_timeout
        _timeout = _aio_timeout.ClientTimeout(total=30, connect=10)
        # Parâmetro correto: chaveNfe (descoberto via análise do JS Angular do portal)
        url = f"{API_BASE}/api/due/listar-due-consulta"
        async with session.get(url, headers=headers, params={"chaveNfe": chave}, proxy=proxy, timeout=_timeout) as resp:
            novo_csrf = resp.headers.get("x-csrf-token", csrf_token)
            try:
                body = await resp.json(content_type=None)
            except Exception:
                body = await resp.text()

            if resp.status == 403:
                msg = body.get("message", "") if isinstance(body, dict) else str(body)
                if "CAPTCHA" in msg.upper():
                    resultado["status_nfe"] = "Erro CAPTCHA"
                    resultado["obs"] = "Sessão expirou — CAPTCHA precisa ser renovado"
                else:
                    resultado["status_nfe"] = "Erro"
                    resultado["obs"] = f"HTTP 403: {msg[:150]}"
                return resultado, novo_csrf

            if resp.status == 422:
                msg = body.get("message", "") if isinstance(body, dict) else str(body)
                msg_lower = msg.lower()
                # Distinguir rate limit de "não encontrado"
                is_rate_limit = any(kw in msg_lower for kw in [
                    "limite de", "após as", "limit", "rate"
                ])
                if is_rate_limit:
                    resultado["status_nfe"] = "Erro Rate Limit"
                    resultado["obs"] = msg[:400]
                    log(job_id, "WARN", f"HTTP 422 rate limit para {chave[:10]}...: {msg[:200]}")
                else:
                    resultado["status_nfe"] = "Não Encontrada"
                    resultado["obs"] = msg[:400]
                    log(job_id, "INFO", f"HTTP 422 sem resultado para {chave[:10]}...: {msg[:200]}")
                return resultado, novo_csrf

            if resp.status != 200:
                detalhe = str(body)[:300] if body else ""
                log(job_id, "WARN", f"HTTP {resp.status} para {chave[:10]}...: {detalhe}")
                resultado["obs"] = f"HTTP {resp.status}: {detalhe}"
                return resultado, novo_csrf

            # A resposta pode ter envelope listaDueCover ou ser lista direta
            if isinstance(body, dict):
                itens = body.get("listaDueCover") or body.get("listaDue") or [body]
            elif isinstance(body, list):
                itens = body
            else:
                itens = []

            log(job_id, "INFO", f"Resposta API ({len(itens)} item(ns)): {json.dumps(body, ensure_ascii=False)[:400]}")

            if not itens:
                resultado["status_nfe"] = "Não Encontrada"
                return resultado, novo_csrf

            item = itens[0]

            # Campos confirmados pelo JS: nrDue, dataRegistro, situacao
            def campo(*nomes):
                for n in nomes:
                    v = item.get(n)
                    if v is not None:
                        return str(v)
                return ""

            numero_due = campo("nrDue", "numeroDue", "numero")
            data_due   = campo("dataRegistro", "dataInclusao", "dataDue", "data")
            status_due = campo("situacao", "statusDue", "status")

            resultado["numero_due"] = numero_due
            resultado["data_due"]   = data_due
            resultado["status_due"] = status_due

            # Determinar averbação: se tem nrDue e situação indica exportação concluída
            situacao_lower = status_due.lower()
            if numero_due and ("averbad" in situacao_lower or "desembarac" in situacao_lower
                               or "encerrad" in situacao_lower or "registrad" in situacao_lower):
                resultado["status_nfe"] = "Averbada"
            elif "nao averbad" in situacao_lower or "não averbad" in situacao_lower:
                resultado["status_nfe"] = "Não Averbada"
            elif numero_due:
                resultado["status_nfe"] = "Averbada"
            else:
                resultado["status_nfe"] = "Não Encontrada"

    except asyncio.TimeoutError:
        resultado["status_nfe"] = "Timeout"
        resultado["obs"] = "Timeout de 30s — proxy lento ou bloqueado"
        log(job_id, "WARN", f"Timeout na API para {chave[:10]}... (proxy: {proxy or 'direto'})")
        return resultado, csrf_token

    except Exception as e:
        resultado["obs"] = str(e)[:200]
        log(job_id, "ERROR", f"Erro na API para {chave[:10]}...: {e}")
        return resultado, csrf_token

    return resultado, novo_csrf


class ProxyRotator:
    """Rotaciona proxies automaticamente quando cada um atinge o rate limit."""

    def __init__(self, proxy_file: Path):
        self.proxies: list[str | None] = self._load(proxy_file)
        self._blocked: dict[str, float] = {}  # chave -> timestamp de liberação
        self._idx = 0
        self._limiters: dict = {}

    def _load(self, path: Path) -> list[str | None]:
        if not path.exists():
            return [None]
        proxies: list[str | None] = []
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                proxies.append(line)
        return proxies if proxies else [None]

    def _key(self, proxy: str | None) -> str:
        return proxy or "direto"

    def current(self) -> str | None:
        """Retorna próximo proxy disponível (round-robin). Retorna 'TODOS_BLOQUEADOS' se nenhum disponível."""
        now = time.time()
        for _ in range(len(self.proxies)):
            p = self.proxies[self._idx % len(self.proxies)]
            self._idx += 1  # avança sempre — garante round-robin real
            if self._blocked.get(self._key(p), 0) <= now:
                return p
        return "TODOS_BLOQUEADOS"

    def mark_limited(self, proxy: str | None, reset_timestamp: float):
        self._blocked[self._key(proxy)] = reset_timestamp
        self._idx += 1

    def seconds_to_next(self) -> float:
        now = time.time()
        if not self._blocked:
            return 0.0
        return max(0.0, min(self._blocked.values()) - now) + 5

    def _limiter(self, proxy: str | None):
        k = self._key(proxy)
        if k not in self._limiters:
            self._limiters[k] = RateLimiter()
        return self._limiters[k]

    def record(self, proxy: str | None):
        self._limiter(proxy).record()

    def near_limit(self, proxy: str | None) -> bool:
        return self._limiter(proxy).near_limit()

    def status(self) -> list[dict]:
        now = time.time()
        return [
            {
                "proxy": p or "IP direto",
                "disponivel": self._blocked.get(self._key(p), 0) <= now,
                "libera_em": max(0, int(self._blocked.get(self._key(p), 0) - now)),
            }
            for p in self.proxies
        ]

    @property
    def count(self) -> int:
        return len([p for p in self.proxies if p])


def _calcular_espera_rate_limit(msg: str) -> int:
    """Extrai o horário de liberação da mensagem 422 e retorna segundos a aguardar."""
    m = re.search(r'após as (\d{2}:\d{2}:\d{2})', msg)
    if m:
        h, mi, s = map(int, m.group(1).split(':'))
        agora = datetime.now()
        target = agora.replace(hour=h, minute=mi, second=s, microsecond=0)
        if target <= agora:
            target = target.replace(day=agora.day + 1)
        return max(int((target - agora).total_seconds()) + 10, 0)
    return 3600  # fallback: aguardar 1 hora


class RateLimiter:
    """Janela deslizante: sinaliza quando o limite de req/hora está próximo."""

    def __init__(self, max_requests: int = 950, window_seconds: int = 3600):
        self.max_requests = max_requests
        self.window_seconds = window_seconds
        self._timestamps: deque[float] = deque()

    def _clean(self):
        now = time.time()
        while self._timestamps and now - self._timestamps[0] >= self.window_seconds:
            self._timestamps.popleft()

    def near_limit(self) -> bool:
        self._clean()
        return len(self._timestamps) >= self.max_requests

    def record(self):
        self._timestamps.append(time.time())

    def reset(self):
        self._timestamps.clear()


async def processar_job_async(job_id: str, chaves: list[str]):
    import aiohttp

    # Filtrar chaves já processadas — suporte a retomada de onde parou
    conn = sqlite3.connect(DB_PATH, timeout=15)
    ja_processadas = {row[0] for row in conn.execute(
        "SELECT chave_nfe FROM resultados WHERE job_id=?", (job_id,)
    ).fetchall()}
    conn.close()

    chaves_pendentes = [c for c in chaves if c not in ja_processadas]
    base_processadas = len(ja_processadas)

    if ja_processadas:
        log(job_id, "INFO", f"{base_processadas} chaves já processadas — retomando as {len(chaves_pendentes)} restantes")

    log(job_id, "INFO", f"Iniciando processamento de {len(chaves_pendentes)} chaves")
    update_job(job_id, status="running", total=len(chaves), processed=base_processadas)

    if not chaves_pendentes:
        output_file = gerar_relatorio(job_id)
        update_job(job_id, status="done", finished_at=datetime.now().isoformat(), output_file=str(output_file))
        log(job_id, "INFO", "Todas as chaves já estavam processadas. Relatório gerado.")
        return

    # Fase 1 — calcular quantos proxies são necessários e ordenar pelos mais frescos
    rotator = ProxyRotator(PROXY_FILE)
    todos_proxies = rotator.proxies  # [None] se sem proxies, ou lista de URLs

    # Proxies necessários = ceil(chaves / 1000), mínimo 1, máximo disponível
    necessarios = max(1, min(math.ceil(len(chaves_pendentes) / 1000), len(todos_proxies)))

    # Ordenar pelos mais frescos (sem cooldown) e usar só os necessários
    proxies = ordenar_proxies_por_disponibilidade(todos_proxies)[:necessarios]

    n_proxies = len(proxies)
    log(job_id, "INFO",
        f"{len(chaves_pendentes)} chaves → {necessarios} proxy(s) necessário(s) "
        f"(de {len(todos_proxies)} disponíveis) — abrindo sessões pelos mais frescos")
    update_job(job_id, status=f"iniciando_sessoes:0/{n_proxies}")

    sessoes = await inicializar_sessoes(job_id, proxies)
    if not sessoes:
        log(job_id, "ERROR", "Nenhuma sessão obtida — abortando")
        update_job(job_id, status="error", finished_at=datetime.now().isoformat())
        return

    capacidade = len(sessoes) * 1000
    log(job_id, "INFO", f"{len(sessoes)} sessão(ões) prontas — capacidade: ~{capacidade} req/hora")
    update_job(job_id, status="running")

    # Fase 2 — fila compartilhada, cada sessão puxa e processa independentemente
    queue: asyncio.Queue = asyncio.Queue()
    for chave in chaves_pendentes:
        await queue.put(chave)

    _processed = [base_processadas]
    _stop = [False]

    async def worker_sessao(sessao: dict):
        """Worker dedicado a uma sessão/proxy — puxa da fila até esvaziar."""
        async with aiohttp.ClientSession(cookies=sessao["cookies"]) as http:
            while not _stop[0]:

                # Verificar perto do limite ANTES de pegar chave da fila
                # (assim outras sessões ativas pegam enquanto esta dorme)
                if sessao["limiter"].near_limit():
                    espera = 3610
                    reset_em = (datetime.now() + timedelta(seconds=espera)).isoformat()
                    registrar_rate_limit_proxy(sessao["proxy"], reset_em)
                    log(job_id, "WARN", f"[{sessao['label']}] Perto do limite — aguardando reset (~1h, libera: {reset_em[:19]})")
                    update_job(job_id, status="aguardando_rate_limit")
                    await asyncio.sleep(espera)
                    sessao["limiter"].reset()
                    update_job(job_id, status="running")

                try:
                    chave = queue.get_nowait()
                except asyncio.QueueEmpty:
                    break

                # Verificar cancelamento
                conn = sqlite3.connect(DB_PATH, timeout=15)
                row = conn.execute("SELECT status FROM jobs WHERE id=?", (job_id,)).fetchone()
                conn.close()
                if row and row[0] == "cancelled":
                    log(job_id, "INFO", "Job cancelado pelo usuário")
                    _stop[0] = True
                    break

                proxy = sessao["proxy"]
                total_idx = _processed[0] + 1
                log(job_id, "INFO", f"[{sessao['label']}] Consultando {total_idx}/{len(chaves)}: {chave[:10]}...{chave[-6:]}")
                resultado, novo_csrf = await consultar_via_api(http, chave, sessao["csrf"], job_id, proxy=proxy)
                sessao["csrf"] = novo_csrf
                sessao["limiter"].record()
                registrar_uso_proxy(proxy)

                # Rate limit — persiste reset, devolve chave à fila para outra sessão pegar
                if resultado["status_nfe"] == "Erro Rate Limit":
                    espera = _calcular_espera_rate_limit(resultado["obs"])
                    reset_em = (datetime.now() + timedelta(seconds=espera)).isoformat()
                    registrar_rate_limit_proxy(proxy, reset_em)
                    log(job_id, "WARN", f"[{sessao['label']}] Rate limit — devolvendo chave à fila, aguardando {espera}s (reset: {reset_em[:19]})")
                    await queue.put(chave)
                    update_job(job_id, status="aguardando_rate_limit")
                    await asyncio.sleep(espera)
                    sessao["limiter"].reset()
                    update_job(job_id, status="running")
                    continue

                # CAPTCHA expirou — devolve chave e renova apenas esta sessão
                if resultado["status_nfe"] == "Erro CAPTCHA":
                    log(job_id, "WARN", f"[{sessao['label']}] Sessão expirou — devolvendo chave à fila, renovando CAPTCHA...")
                    await queue.put(chave)  # outra sessão pega enquanto renova
                    nova = await obter_sessao_com_captcha(job_id, proxy=proxy)
                    if nova:
                        sessao["cookies"], sessao["csrf"] = nova
                        http.cookie_jar.update_cookies(sessao["cookies"])
                        sessao["limiter"].reset()
                        log(job_id, "INFO", f"[{sessao['label']}] Sessão renovada — retomando")
                    else:
                        log(job_id, "ERROR", f"[{sessao['label']}] Falha ao renovar — encerrando worker")
                        _stop[0] = True
                        break
                    continue  # volta ao topo sem salvar resultado

                salvar_resultado(
                    job_id, chave,
                    resultado["status_nfe"],
                    resultado["numero_due"],
                    resultado["data_due"],
                    resultado["status_due"],
                    resultado["obs"],
                )
                _processed[0] += 1
                update_job(job_id, processed=_processed[0])
                queue.task_done()

    await asyncio.gather(
        *[worker_sessao(s) for s in sessoes],
        return_exceptions=True,
    )

    output_file = gerar_relatorio(job_id)
    # Marcar como done se todas as chaves foram processadas, independente de status intermediário
    conn = sqlite3.connect(DB_PATH, timeout=15)
    row = conn.execute("SELECT processed, total FROM jobs WHERE id=?", (job_id,)).fetchone()
    conn.close()
    status_final = "done" if (row and row[0] >= row[1]) else "cancelled"
    update_job(job_id, status=status_final, finished_at=datetime.now().isoformat(), output_file=str(output_file))
    log(job_id, "INFO", f"Job {status_final}. Relatório: {output_file}")


def processar_job_thread(job_id: str, chaves: list[str]):
    asyncio.run(processar_job_async(job_id, chaves))

# ─── Geração de relatório Excel ───────────────────────────────────────────────

def gerar_relatorio(job_id: str) -> Path:
    conn = sqlite3.connect(DB_PATH, timeout=15)
    rows = conn.execute(
        """SELECT chave_nfe, status_nfe, numero_due, data_due, status_due, observacao, consultado_em
           FROM resultados WHERE job_id=? ORDER BY id""",
        (job_id,)
    ).fetchall()
    conn.close()

    df = pd.DataFrame(rows, columns=[
        "Chave NF-e", "Status", "Número DUE", "Data DUE",
        "Status DUE", "Observações", "Consultado em"
    ])

    output_path = RESULTS_DIR / f"resultado_{job_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Resultados", index=False)
        ws = writer.sheets["Resultados"]

        # Formatação
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", start_color="1B3A5C")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Colorir linhas por status
        green = PatternFill("solid", start_color="C6EFCE")
        yellow = PatternFill("solid", start_color="FFEB9C")
        red = PatternFill("solid", start_color="FFC7CE")
        gray = PatternFill("solid", start_color="D9D9D9")

        status_colors = {
            "Averbada": green,
            "Não Averbada": yellow,
            "Não Encontrada": gray,
        }

        for row_idx, row in enumerate(df.itertuples(), start=2):
            status = row.Status
            fill = status_colors.get(status, red)
            for col in range(1, len(df.columns) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

        # Largura das colunas
        col_widths = [50, 18, 30, 18, 20, 40, 22]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Aba de resumo
        ws_sum = writer.book.create_sheet("Resumo")
        total = len(df)
        averbadas = (df["Status"] == "Averbada").sum()
        nao_averbadas = (df["Status"] == "Não Averbada").sum()
        nao_encontradas = (df["Status"] == "Não Encontrada").sum()
        erros = total - averbadas - nao_averbadas - nao_encontradas

        summary_data = [
            ["Métrica", "Quantidade", "%"],
            ["Total consultado", total, "100%"],
            ["Averbadas", averbadas, f"{averbadas/total*100:.1f}%" if total else "0%"],
            ["Não Averbadas", nao_averbadas, f"{nao_averbadas/total*100:.1f}%" if total else "0%"],
            ["Não Encontradas", nao_encontradas, f"{nao_encontradas/total*100:.1f}%" if total else "0%"],
            ["Erros", erros, f"{erros/total*100:.1f}%" if total else "0%"],
            ["", "", ""],
            ["Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S"), ""],
        ]
        for r_data in summary_data:
            ws_sum.append(r_data)

    return output_path

# ─── Rotas da API ─────────────────────────────────────────────────────────────

@app.route("/api/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400

    f = request.files["file"]
    if not f.filename.endswith((".xlsx", ".csv")):
        return jsonify({"error": "Formato inválido. Use .xlsx ou .csv"}), 400

    job_id = str(uuid.uuid4())
    safe_name = re.sub(r'[^\w.\-]', '_', f.filename)
    file_path = UPLOAD_DIR / f"{job_id}_{safe_name}"

    try:
        f.save(str(file_path))

        if str(file_path).endswith(".csv"):
            df = pd.read_csv(str(file_path), dtype=str)
        else:
            df = pd.read_excel(str(file_path), dtype=str)

        # Localizar coluna com chaves NF-e
        chave_col = None
        for col in df.columns:
            if any(kw in col.lower() for kw in ["chave", "nfe", "nf-e", "key", "nota"]):
                chave_col = col
                break
        if not chave_col:
            for col in df.columns:
                sample = df[col].dropna().astype(str).str.replace(r'\D', '', regex=True)
                if sample.str.len().eq(44).any():
                    chave_col = col
                    break

        if not chave_col:
            return jsonify({"error": "Coluna com chaves NF-e não encontrada"}), 400

        chaves_raw = df[chave_col].dropna().astype(str).tolist()
        chaves_valid = []
        chaves_invalidas = []

        for c in chaves_raw:
            norm = normalizar_chave(c)
            if validar_chave(norm):
                if norm not in chaves_valid:
                    chaves_valid.append(norm)
            else:
                chaves_invalidas.append(c)

        if not chaves_valid:
            return jsonify({"error": "Nenhuma chave NF-e válida encontrada"}), 400

        conn = sqlite3.connect(DB_PATH, timeout=15)
        conn.execute(
            "INSERT INTO jobs (id, status, total, processed, created_at, input_file) VALUES (?, ?, ?, ?, ?, ?)",
            (job_id, "pending", len(chaves_valid), 0, datetime.now().isoformat(), str(file_path))
        )
        conn.commit()
        conn.close()

        chaves_path = UPLOAD_DIR / f"{job_id}_chaves.json"
        chaves_path.write_text(json.dumps(chaves_valid))

        log(job_id, "INFO", f"Upload processado: {len(chaves_valid)} válidas, {len(chaves_invalidas)} inválidas")

        return jsonify({
            "job_id": job_id,
            "total_validas": len(chaves_valid),
            "total_invalidas": len(chaves_invalidas),
            "invalidas_preview": chaves_invalidas[:10],
            "coluna_detectada": chave_col
        })

    except Exception as ex:
        import traceback
        print(f"[UPLOAD ERROR] {traceback.format_exc()}")
        return jsonify({"error": f"Erro ao processar arquivo: {ex}"}), 400

@app.route("/api/iniciar/<job_id>", methods=["POST"])
def iniciar(job_id):
    conn = sqlite3.connect(DB_PATH, timeout=15)
    row = conn.execute("SELECT status FROM jobs WHERE id=?", (job_id,)).fetchone()
    conn.close()

    if not row:
        return jsonify({"error": "Job não encontrado"}), 404
    if row[0] not in ("pending",):
        return jsonify({"error": f"Job já está em status: {row[0]}"}), 400

    chaves_path = UPLOAD_DIR / f"{job_id}_chaves.json"
    if not chaves_path.exists():
        return jsonify({"error": "Arquivo de chaves não encontrado"}), 400

    chaves = json.loads(chaves_path.read_text())

    t = threading.Thread(target=processar_job_thread, args=(job_id, chaves), daemon=True)
    t.start()

    return jsonify({"message": "Processamento iniciado", "job_id": job_id})

@app.route("/api/status/<job_id>")
def status(job_id):
    conn = sqlite3.connect(DB_PATH, timeout=15)
    row = conn.execute(
        "SELECT status, total, processed, created_at, finished_at, output_file FROM jobs WHERE id=?",
        (job_id,)
    ).fetchone()

    logs = conn.execute(
        "SELECT nivel, mensagem, criado_em FROM logs WHERE job_id=? ORDER BY id DESC LIMIT 50",
        (job_id,)
    ).fetchall()

    # Últimos resultados
    ultimos = conn.execute(
        """SELECT chave_nfe, status_nfe, numero_due, consultado_em 
           FROM resultados WHERE job_id=? ORDER BY id DESC LIMIT 20""",
        (job_id,)
    ).fetchall()

    conn.close()

    if not row:
        return jsonify({"error": "Job não encontrado"}), 404

    status_val, total, processed, created_at, finished_at, output_file = row
    pct = round(processed / total * 100, 1) if total else 0

    return jsonify({
        "job_id": job_id,
        "status": status_val,
        "total": total,
        "processed": processed,
        "percent": pct,
        "created_at": created_at,
        "finished_at": finished_at,
        "has_output": bool(output_file and (
            Path(output_file).exists() if Path(output_file).is_absolute()
            else (_BASE_DIR / output_file).exists() or Path(output_file).resolve().exists()
        )),
        "logs": [{"nivel": l[0], "mensagem": l[1], "criado_em": l[2]} for l in logs],
        "ultimos_resultados": [
            {"chave": r[0][:10] + "..." + r[0][-6:], "status": r[1], "due": r[2], "em": r[3]}
            for r in ultimos
        ]
    })

@app.route("/api/cancelar/<job_id>", methods=["POST"])
def cancelar(job_id):
    update_job(job_id, status="cancelled")
    return jsonify({"message": "Solicitação de cancelamento enviada"})

@app.route("/api/pausar/<job_id>", methods=["POST"])
def pausar(job_id):
    try:
        conn = sqlite3.connect(DB_PATH, timeout=15)
        row = conn.execute("SELECT status, processed FROM jobs WHERE id=?", (job_id,)).fetchone()
        conn.close()
        if not row:
            return jsonify({"error": "Job não encontrado"}), 404
        update_job(job_id, status="cancelled")
        output_file = gerar_relatorio(job_id)
        update_job(job_id, output_file=str(output_file), finished_at=datetime.now().isoformat())
        return jsonify({"message": "Job pausado — relatório parcial gerado", "job_id": job_id})
    except Exception as ex:
        return jsonify({"error": str(ex)}), 500

@app.route("/api/retomar/<job_id>", methods=["POST"])
def retomar(job_id):
    conn = sqlite3.connect(DB_PATH, timeout=15)
    row = conn.execute("SELECT status FROM jobs WHERE id=?", (job_id,)).fetchone()
    conn.close()

    if not row:
        return jsonify({"error": "Job não encontrado"}), 404
    if row[0] == "running":
        return jsonify({"error": "Job já está em execução"}), 400

    chaves_path = UPLOAD_DIR / f"{job_id}_chaves.json"
    if not chaves_path.exists():
        return jsonify({"error": "Arquivo de chaves não encontrado"}), 400

    chaves = json.loads(chaves_path.read_text())
    update_job(job_id, status="pending", finished_at=None)

    t = threading.Thread(target=processar_job_thread, args=(job_id, chaves), daemon=True)
    t.start()

    return jsonify({"message": "Job retomado com sucesso", "job_id": job_id})

@app.route("/api/download/<job_id>")
def download(job_id):
    conn = sqlite3.connect(DB_PATH, timeout=15)
    row = conn.execute("SELECT output_file FROM jobs WHERE id=?", (job_id,)).fetchone()
    conn.close()

    if not row or not row[0]:
        return jsonify({"error": "Relatório não disponível"}), 404

    output_path = Path(row[0])
    if not output_path.is_absolute():
        # Try backend dir first, then resolve against CWD (legacy paths from root runs)
        candidate = _BASE_DIR / output_path
        output_path = candidate if candidate.exists() else output_path.resolve()

    if not output_path.exists():
        output_path = gerar_relatorio(job_id)

    return send_file(str(output_path), as_attachment=True, download_name=f"resultado_due_{job_id[:8]}.xlsx")

@app.route("/api/jobs")
def listar_jobs():
    conn = sqlite3.connect(DB_PATH, timeout=15)
    rows = conn.execute(
        "SELECT id, status, total, processed, created_at, finished_at FROM jobs ORDER BY created_at DESC LIMIT 20"
    ).fetchall()
    conn.close()

    return jsonify([
        {"id": r[0], "status": r[1], "total": r[2], "processed": r[3],
         "created_at": r[4], "finished_at": r[5]}
        for r in rows
    ])

@app.route("/api/health")
def health():
    return jsonify({"status": "ok", "version": "1.0.0"})

@app.route("/debug-pdf-texto", methods=["GET"])
def debug_pdf_texto():
    """Retorna texto bruto do primeiro PDF na pasta de entrada."""
    from perdcomp import _ECAC_DIR, extrair_texto
    pasta = _ECAC_DIR / "entrada"
    pdfs = list(pasta.glob("*.pdf"))
    if not pdfs:
        return jsonify({"error": "Nenhum PDF na pasta. Faça upload primeiro."}), 400
    nome = request.args.get("arquivo")
    pdf = next((p for p in pdfs if p.name == nome), pdfs[0])
    texto = extrair_texto(pdf.read_bytes())
    return f"<pre style='font-family:monospace;white-space:pre-wrap'><b>{pdf.name}</b>\n\n{texto[:4000]}</pre>"

@app.route("/upload-pdfs-ecac", methods=["GET", "POST"])
def ecac_upload_pdfs_main():
    if request.method == "GET":
        return jsonify({"status": "ok", "endpoint": "upload-pdfs-ecac"})
    from perdcomp import _ECAC_DIR
    from datetime import datetime as _dt
    ts = _dt.now().strftime("%Y-%m-%d_%H-%M-%S")
    pasta = _ECAC_DIR / "entrada" / ts
    pasta.mkdir(parents=True, exist_ok=True)
    files = request.files.getlist("files")
    salvos, erros = [], []
    for f in files:
        if not f.filename.lower().endswith(".pdf"):
            continue
        try:
            dest = pasta / Path(f.filename).name
            dest.write_bytes(f.read())
            salvos.append(f.filename)
        except Exception as exc:
            erros.append({"arquivo": f.filename, "erro": str(exc)})
    return jsonify({
        "salvos": len(salvos),
        "total_pasta": len(list(pasta.glob("*.pdf"))),
        "pasta_pdfs": str(pasta.resolve()),
        "erros": erros
    })

@app.route("/api/proxies")
def listar_proxies():
    rotator = ProxyRotator(PROXY_FILE)
    return jsonify({
        "total": len(rotator.proxies),
        "configurados": rotator.count,
        "capacidade_hora": rotator.count * 1000 if rotator.count else 1000,
        "proxies": rotator.status()
    })

from piscofins import bp as piscofins_bp
app.register_blueprint(piscofins_bp)

from perdcomp import bp as perdcomp_bp
app.register_blueprint(perdcomp_bp)

from simples import bp as simples_bp
app.register_blueprint(simples_bp)

from societario import bp as societario_bp
app.register_blueprint(societario_bp)

# Erros sempre retornam JSON para rotas /api, HTML para o resto
@app.errorhandler(404)
def handle_404(e):
    from flask import request as req
    if req.path.startswith("/api/"):
        return jsonify({"error": "Rota não encontrada", "path": req.path}), 404
    return str(e), 404

@app.errorhandler(405)
def handle_405(e):
    from flask import request as req
    if req.path.startswith("/api/"):
        return jsonify({"error": "Método não permitido"}), 405
    return str(e), 405

@app.errorhandler(500)
def handle_500(e):
    from flask import request as req
    import traceback
    tb = traceback.format_exc()
    print(f"[500] {req.path}\n{tb}")
    if req.path.startswith("/api/"):
        return jsonify({"error": "Erro interno do servidor", "detalhe": str(e)}), 500
    return str(e), 500

# Serve static docs so browsers don't hit file:// CORS restrictions
_DOCS_DIR = str(_BASE_DIR.parent / "docs")

@app.route("/")
def serve_index():
    return send_file(str(Path(_DOCS_DIR) / "portal.html"))

@app.route("/<path:filename>")
def serve_docs(filename):
    from flask import abort
    filepath = Path(_DOCS_DIR) / filename
    if filepath.suffix in (".html", ".css", ".js", ".png", ".svg", ".ico") and filepath.exists():
        return send_file(str(filepath))
    abort(404)

if __name__ == "__main__":
    init_db()
    print("=" * 60)
    print("  DUE Consulta Backend — Rodando em http://localhost:5000")
    print("  Portal                — http://localhost:5000/portal.html")
    print("  PIS/COFINS            — http://localhost:5000/piscofins.html")
    print("  Simples Nacional      — http://localhost:5000/simples-nacional.html")
    print("=" * 60)
    app.run(host="0.0.0.0", port=5000, debug=False)

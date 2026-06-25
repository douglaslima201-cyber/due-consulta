"""
Módulo de Gestão Patrimonial e Simulação Tributária
Ativos de transportadoras: cavalos mecânicos, carretas, implementos, máquinas.
"""

import calendar
import io
import json
import math
import sqlite3
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from flask import Blueprint, jsonify, request, send_file

bp = Blueprint("patrimonio", __name__, url_prefix="/patrimonio")

_BASE = Path(__file__).parent
DB = str(_BASE / "patrimonio.db")


# ─── helpers ─────────────────────────────────────────────────────────────────

def add_months(dt: datetime, n: int) -> datetime:
    month = dt.month - 1 + n
    year = dt.year + month // 12
    month = month % 12 + 1
    day = min(dt.day, calendar.monthrange(year, month)[1])
    return dt.replace(year=year, month=month, day=day)


def db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def init_db():
    c = db()
    c.executescript("""
    CREATE TABLE IF NOT EXISTS patrimonio_ativos (
        id TEXT PRIMARY KEY,
        codigo TEXT UNIQUE NOT NULL,
        descricao TEXT NOT NULL,
        tipo_ativo TEXT DEFAULT 'veiculo',
        categoria TEXT,
        placa TEXT,
        chassi TEXT,
        data_aquisicao TEXT,
        data_entrada_operacao TEXT,
        vida_util_fiscal_meses INTEGER DEFAULT 60,
        vida_util_societaria_meses INTEGER DEFAULT 60,
        taxa_depreciacao_societaria_pct_ano REAL DEFAULT 0,
        valor_aquisicao REAL DEFAULT 0,
        valor_residual_estimado REAL DEFAULT 0,
        estado_conservacao TEXT DEFAULT 'bom',
        centro_custo TEXT,
        filial TEXT,
        unidade_negocio TEXT,
        regime_tributario TEXT DEFAULT 'lucro_real',
        aliquota_icms REAL DEFAULT 0,
        valor_icms_aquisicao REAL DEFAULT 0,
        numero_turnos INTEGER DEFAULT 1,
        metodo_credito_pis_cofins TEXT DEFAULT 'depreciacao',
        status TEXT DEFAULT 'ativo',
        observacoes TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS patrimonio_financiamentos (
        id TEXT PRIMARY KEY,
        ativo_id TEXT,
        banco TEXT,
        numero_contrato TEXT,
        valor_financiado REAL DEFAULT 0,
        entrada REAL DEFAULT 0,
        taxa_juros_mensal REAL DEFAULT 0,
        sistema_amortizacao TEXT DEFAULT 'price',
        prazo_meses INTEGER DEFAULT 48,
        valor_residual_balao REAL DEFAULT 0,
        data_inicio TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS patrimonio_vendas (
        id TEXT PRIMARY KEY,
        ativo_id TEXT,
        data_venda TEXT,
        valor_venda REAL DEFAULT 0,
        comprador TEXT,
        resultado_societario REAL,
        resultado_fiscal REAL,
        tributos_totais REAL,
        observacoes TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );
    """)
    c.commit()
    c.close()


init_db()

# Migração: adiciona coluna se DB antigo não a tiver
try:
    _c = db()
    _c.execute("ALTER TABLE patrimonio_ativos ADD COLUMN taxa_depreciacao_societaria_pct_ano REAL DEFAULT 0")
    _c.commit()
    _c.close()
except Exception:
    pass  # coluna já existe


# ─── cálculos ─────────────────────────────────────────────────────────────────

def _resolver_dep_societaria(ativo: dict):
    """
    Retorna (dep_mensal, vida_meses, taxa_anual_pct, metodo).
    Prioridade: taxa_depreciacao_societaria_pct_ano > vida_util_societaria_meses.
    """
    valor = float(ativo.get("valor_aquisicao") or 0)
    residual = float(ativo.get("valor_residual_estimado") or 0)
    depreciavel = max(valor - residual, 0)
    taxa_anual = float(ativo.get("taxa_depreciacao_societaria_pct_ano") or 0)

    if taxa_anual > 0:
        dep_m = depreciavel * (taxa_anual / 100) / 12
        vida = int(depreciavel / dep_m) if dep_m > 0 else 0
        metodo = "aliquota"
    else:
        vida = int(ativo.get("vida_util_societaria_meses") or 60)
        dep_m = depreciavel / vida if vida > 0 else 0
        taxa_anual = (dep_m * 12 / valor * 100) if valor > 0 else 0
        metodo = "vida_util"

    return dep_m, vida, round(taxa_anual, 4), metodo


def dep_societaria(ativo: dict) -> list:
    valor = float(ativo.get("valor_aquisicao") or 0)
    residual = float(ativo.get("valor_residual_estimado") or 0)
    entrada_str = ativo.get("data_entrada_operacao") or ativo.get("data_aquisicao")
    if not entrada_str:
        return []
    try:
        dt0 = datetime.strptime(entrada_str[:10], "%Y-%m-%d")
    except ValueError:
        return []

    dep_m, vida, _, _ = _resolver_dep_societaria(ativo)
    if vida <= 0 or dep_m <= 0:
        return []

    schedule, acum = [], 0.0
    for i in range(vida):
        prev = acum
        acum = min(acum + dep_m, max(valor - residual, 0))
        vlc = max(valor - acum, residual)
        schedule.append({
            "mes": i + 1,
            "competencia": add_months(dt0, i).strftime("%Y-%m"),
            "depreciacao_mensal": round(acum - prev, 2),
            "depreciacao_acumulada": round(acum, 2),
            "valor_liquido_contabil": round(vlc, 2),
        })
    return schedule


def dep_fiscal(ativo: dict) -> list:
    valor = float(ativo.get("valor_aquisicao") or 0)
    vida = int(ativo.get("vida_util_fiscal_meses") or 60)
    turnos = int(ativo.get("numero_turnos") or 1)
    entrada_str = ativo.get("data_entrada_operacao") or ativo.get("data_aquisicao")
    if not entrada_str or vida <= 0:
        return []
    try:
        dt0 = datetime.strptime(entrada_str[:10], "%Y-%m-%d")
    except ValueError:
        return []
    acelerador = {1: 1.0, 2: 1.5, 3: 2.0}.get(turnos, 1.0)
    vida_efetiva = max(int(round(vida / acelerador)), 1)
    dep_m = valor / vida_efetiva
    schedule, acum = [], 0.0
    for i in range(vida_efetiva):
        prev = acum
        acum = min(acum + dep_m, valor)
        schedule.append({
            "mes": i + 1,
            "competencia": add_months(dt0, i).strftime("%Y-%m"),
            "depreciacao_mensal": round(acum - prev, 2),
            "depreciacao_acumulada": round(acum, 2),
            "saldo_fiscal": round(max(valor - acum, 0), 2),
        })
    return schedule


def creditos_pis_cofins(ativo: dict, sched_soc: list) -> dict:
    regime = ativo.get("regime_tributario", "lucro_real")
    metodo = ativo.get("metodo_credito_pis_cofins", "depreciacao")
    valor = float(ativo.get("valor_aquisicao") or 0)
    PIS, COF = 0.0165, 0.076
    if regime != "lucro_real":
        return {"creditos": [], "total_pis": 0, "total_cofins": 0,
                "total_credito": 0, "regime": regime, "metodo": metodo}
    creditos = []
    for i, d in enumerate(sched_soc):
        if metodo == "aquisicao_12":
            base = valor / 12 if i < 12 else 0
        elif metodo == "aquisicao_48":
            base = valor / 48 if i < 48 else 0
        else:
            base = d["depreciacao_mensal"]
        creditos.append({
            "mes": d["mes"],
            "competencia": d["competencia"],
            "base": round(base, 2),
            "credito_pis": round(base * PIS, 2),
            "credito_cofins": round(base * COF, 2),
            "credito_total": round(base * (PIS + COF), 2),
        })
    tp = sum(c["credito_pis"] for c in creditos)
    tc = sum(c["credito_cofins"] for c in creditos)
    return {"creditos": creditos, "total_pis": round(tp, 2),
            "total_cofins": round(tc, 2), "total_credito": round(tp + tc, 2),
            "regime": regime, "metodo": metodo}


def cronograma_financiamento(fin: dict) -> list:
    PV = float(fin.get("valor_financiado") or 0)
    entrada = float(fin.get("entrada") or 0)
    i = float(fin.get("taxa_juros_mensal") or 0) / 100
    n = int(fin.get("prazo_meses") or 48)
    sistema = fin.get("sistema_amortizacao", "price")
    balao = float(fin.get("valor_residual_balao") or 0)
    ds = fin.get("data_inicio")
    saldo0 = PV - entrada
    if saldo0 <= 0 or n <= 0:
        return []
    try:
        dt0 = datetime.strptime(ds[:10], "%Y-%m-%d")
    except Exception:
        dt0 = datetime.now()

    cron, saldo = [], saldo0
    if sistema == "price":
        saldo_pmt = saldo0 - (balao / (1 + i) ** n if balao > 0 and i > 0 else 0)
        pmt = (saldo_pmt * i * (1 + i) ** n / ((1 + i) ** n - 1)) if i > 0 else saldo_pmt / n
        for k in range(1, n + 1):
            j = saldo * i
            am = pmt - j
            saldo = max(saldo - am, 0)
            par = pmt + (balao if k == n else 0)
            if k == n:
                saldo = 0
            cron.append({
                "parcela": k,
                "vencimento": add_months(dt0, k).strftime("%Y-%m-%d"),
                "amortizacao": round(am, 2),
                "juros": round(j, 2),
                "parcela_total": round(par, 2),
                "saldo_devedor": round(saldo, 2),
            })
    else:  # SAC
        am_fixa = saldo0 / n
        for k in range(1, n + 1):
            j = saldo * i
            saldo = max(saldo - am_fixa, 0)
            par = am_fixa + j + (balao if k == n else 0)
            if k == n:
                saldo = 0
            cron.append({
                "parcela": k,
                "vencimento": add_months(dt0, k).strftime("%Y-%m-%d"),
                "amortizacao": round(am_fixa, 2),
                "juros": round(j, 2),
                "parcela_total": round(par, 2),
                "saldo_devedor": round(saldo, 2),
            })
    return cron


def resultado_venda(ativo: dict, valor_venda: float, data_venda_str: str) -> dict:
    soc = dep_societaria(ativo)
    fisc = dep_fiscal(ativo)
    entrada_str = ativo.get("data_entrada_operacao") or ativo.get("data_aquisicao") or ""
    try:
        dv = datetime.strptime(data_venda_str[:10], "%Y-%m-%d")
        de = datetime.strptime(entrada_str[:10], "%Y-%m-%d")
        meses = (dv.year - de.year) * 12 + (dv.month - de.month)
        anos = meses / 12
    except Exception:
        meses, anos = 0, 0.0

    meses = max(meses, 0)
    valor = float(ativo.get("valor_aquisicao") or 0)
    residual = float(ativo.get("valor_residual_estimado") or 0)

    dep_soc_acum = sum(d["depreciacao_mensal"] for d in soc[:meses])
    dep_fisc_acum = sum(d["depreciacao_mensal"] for d in fisc[:meses])
    vlc = max(valor - dep_soc_acum, residual)
    saldo_fisc = max(valor - dep_fisc_acum, 0)

    res_soc = valor_venda - vlc
    res_fisc = valor_venda - saldo_fisc
    ganho = max(res_fisc, 0)

    irpj = ganho * 0.15
    adic = max(ganho - 20000, 0) * 0.10
    csll = ganho * 0.09

    valor_icms = float(ativo.get("valor_icms_aquisicao") or 0)
    icms_estorno = (valor_icms * max(1 - anos / 5, 0)) if anos < 5 and valor_icms > 0 else 0

    tributos = irpj + adic + csll + icms_estorno
    return {
        "valor_venda": round(valor_venda, 2),
        "valor_aquisicao": round(valor, 2),
        "meses_em_uso": meses,
        "anos_em_uso": round(anos, 1),
        "valor_liquido_contabil": round(vlc, 2),
        "saldo_fiscal": round(saldo_fisc, 2),
        "dep_societaria_acumulada": round(dep_soc_acum, 2),
        "dep_fiscal_acumulada": round(dep_fisc_acum, 2),
        "resultado_societario": round(res_soc, 2),
        "resultado_fiscal": round(res_fisc, 2),
        "ganho_capital": round(ganho, 2),
        "irpj": round(irpj, 2),
        "adicional_irpj": round(adic, 2),
        "csll": round(csll, 2),
        "icms_estorno": round(icms_estorno, 2),
        "tributos_totais": round(tributos, 2),
        "resultado_liquido": round(valor_venda - vlc - tributos, 2),
    }


# ─── rotas CRUD ativos ────────────────────────────────────────────────────────

@bp.route("/ativos", methods=["GET"])
def listar_ativos():
    c = db()
    rows = c.execute("SELECT * FROM patrimonio_ativos ORDER BY created_at DESC").fetchall()
    c.close()
    return jsonify([dict(r) for r in rows])


@bp.route("/ativos", methods=["POST"])
def criar_ativo():
    d = request.json or {}
    aid = str(uuid.uuid4())
    now = datetime.now().isoformat()
    c = db()
    try:
        c.execute("""
            INSERT INTO patrimonio_ativos
            (id,codigo,descricao,tipo_ativo,categoria,placa,chassi,
             data_aquisicao,data_entrada_operacao,vida_util_fiscal_meses,
             vida_util_societaria_meses,taxa_depreciacao_societaria_pct_ano,
             valor_aquisicao,valor_residual_estimado,
             estado_conservacao,centro_custo,filial,unidade_negocio,
             regime_tributario,aliquota_icms,valor_icms_aquisicao,
             numero_turnos,metodo_credito_pis_cofins,status,observacoes,
             created_at,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (aid, d["codigo"], d["descricao"],
              d.get("tipo_ativo", "veiculo"), d.get("categoria"),
              d.get("placa"), d.get("chassi"),
              d.get("data_aquisicao"), d.get("data_entrada_operacao"),
              d.get("vida_util_fiscal_meses", 60), d.get("vida_util_societaria_meses", 60),
              d.get("taxa_depreciacao_societaria_pct_ano", 0),
              d.get("valor_aquisicao", 0), d.get("valor_residual_estimado", 0),
              d.get("estado_conservacao", "bom"),
              d.get("centro_custo"), d.get("filial"), d.get("unidade_negocio"),
              d.get("regime_tributario", "lucro_real"),
              d.get("aliquota_icms", 0), d.get("valor_icms_aquisicao", 0),
              d.get("numero_turnos", 1), d.get("metodo_credito_pis_cofins", "depreciacao"),
              d.get("status", "ativo"), d.get("observacoes"),
              now, now))
        c.commit()
        row = c.execute("SELECT * FROM patrimonio_ativos WHERE id=?", (aid,)).fetchone()
        c.close()
        return jsonify(dict(row)), 201
    except Exception as e:
        c.close()
        return jsonify({"error": str(e)}), 400


@bp.route("/ativos/<aid>", methods=["GET"])
def obter_ativo(aid):
    c = db()
    row = c.execute("SELECT * FROM patrimonio_ativos WHERE id=?", (aid,)).fetchone()
    c.close()
    return (jsonify(dict(row)) if row else (jsonify({"error": "Não encontrado"}), 404))


@bp.route("/ativos/<aid>", methods=["PUT"])
def atualizar_ativo(aid):
    d = request.json or {}
    now = datetime.now().isoformat()
    c = db()
    e = c.execute("SELECT * FROM patrimonio_ativos WHERE id=?", (aid,)).fetchone()
    if not e:
        c.close()
        return jsonify({"error": "Não encontrado"}), 404
    e = dict(e)
    c.execute("""
        UPDATE patrimonio_ativos SET
        codigo=?,descricao=?,tipo_ativo=?,categoria=?,placa=?,chassi=?,
        data_aquisicao=?,data_entrada_operacao=?,vida_util_fiscal_meses=?,
        vida_util_societaria_meses=?,taxa_depreciacao_societaria_pct_ano=?,
        valor_aquisicao=?,valor_residual_estimado=?,
        estado_conservacao=?,centro_custo=?,filial=?,unidade_negocio=?,
        regime_tributario=?,aliquota_icms=?,valor_icms_aquisicao=?,
        numero_turnos=?,metodo_credito_pis_cofins=?,status=?,observacoes=?,
        updated_at=? WHERE id=?
    """, (d.get("codigo", e["codigo"]), d.get("descricao", e["descricao"]),
          d.get("tipo_ativo", e["tipo_ativo"]), d.get("categoria", e["categoria"]),
          d.get("placa", e["placa"]), d.get("chassi", e["chassi"]),
          d.get("data_aquisicao", e["data_aquisicao"]),
          d.get("data_entrada_operacao", e["data_entrada_operacao"]),
          d.get("vida_util_fiscal_meses", e["vida_util_fiscal_meses"]),
          d.get("vida_util_societaria_meses", e["vida_util_societaria_meses"]),
          d.get("taxa_depreciacao_societaria_pct_ano", e.get("taxa_depreciacao_societaria_pct_ano", 0)),
          d.get("valor_aquisicao", e["valor_aquisicao"]),
          d.get("valor_residual_estimado", e["valor_residual_estimado"]),
          d.get("estado_conservacao", e["estado_conservacao"]),
          d.get("centro_custo", e["centro_custo"]),
          d.get("filial", e["filial"]), d.get("unidade_negocio", e["unidade_negocio"]),
          d.get("regime_tributario", e["regime_tributario"]),
          d.get("aliquota_icms", e["aliquota_icms"]),
          d.get("valor_icms_aquisicao", e["valor_icms_aquisicao"]),
          d.get("numero_turnos", e["numero_turnos"]),
          d.get("metodo_credito_pis_cofins", e["metodo_credito_pis_cofins"]),
          d.get("status", e["status"]), d.get("observacoes", e["observacoes"]),
          now, aid))
    c.commit()
    row = c.execute("SELECT * FROM patrimonio_ativos WHERE id=?", (aid,)).fetchone()
    c.close()
    return jsonify(dict(row))


@bp.route("/ativos/<aid>", methods=["DELETE"])
def excluir_ativo(aid):
    c = db()
    c.execute("DELETE FROM patrimonio_ativos WHERE id=?", (aid,))
    c.commit()
    c.close()
    return jsonify({"success": True})


# ─── análises por ativo ───────────────────────────────────────────────────────

@bp.route("/ativos/<aid>/depreciacao", methods=["GET"])
def api_dep_soc(aid):
    c = db()
    row = c.execute("SELECT * FROM patrimonio_ativos WHERE id=?", (aid,)).fetchone()
    c.close()
    if not row:
        return jsonify({"error": "Não encontrado"}), 404
    ativo = dict(row)
    sched = dep_societaria(ativo)
    dep_m, vida, taxa_anual, metodo = _resolver_dep_societaria(ativo)
    return jsonify({
        "ativo": ativo,
        "schedule": sched,
        "resumo": {
            "valor_aquisicao": ativo["valor_aquisicao"],
            "valor_residual": ativo["valor_residual_estimado"],
            "depreciacao_mensal": round(dep_m, 2),
            "depreciacao_anual": round(dep_m * 12, 2),
            "vida_util_meses": vida,
            "taxa_anual_pct": taxa_anual,
            "metodo": metodo,
        },
    })


@bp.route("/ativos/<aid>/depreciacao-fiscal", methods=["GET"])
def api_dep_fisc(aid):
    c = db()
    row = c.execute("SELECT * FROM patrimonio_ativos WHERE id=?", (aid,)).fetchone()
    c.close()
    if not row:
        return jsonify({"error": "Não encontrado"}), 404
    ativo = dict(row)
    soc = dep_societaria(ativo)
    fisc = dep_fiscal(ativo)

    soc_map = {d["competencia"]: d for d in soc}
    fisc_map = {d["competencia"]: d for d in fisc}
    comps = sorted(set(list(soc_map) + list(fisc_map)))

    diferencas = []
    for comp in comps:
        s = soc_map.get(comp, {"depreciacao_mensal": 0, "valor_liquido_contabil": 0})
        f = fisc_map.get(comp, {"depreciacao_mensal": 0, "saldo_fiscal": 0})
        dif = f["depreciacao_mensal"] - s["depreciacao_mensal"]
        diferencas.append({
            "competencia": comp,
            "dep_societaria": round(s["depreciacao_mensal"], 2),
            "dep_fiscal": round(f["depreciacao_mensal"], 2),
            "diferenca_temporaria": round(dif, 2),
            "vlc": round(s.get("valor_liquido_contabil", 0), 2),
            "saldo_fiscal": round(f.get("saldo_fiscal", 0), 2),
            "ajuste_irpj": round(dif * 0.15, 2),
            "ajuste_csll": round(dif * 0.09, 2),
        })

    return jsonify({
        "ativo": ativo,
        "societaria": soc,
        "fiscal": fisc,
        "diferencas": diferencas,
        "totais": {
            "dep_soc_total": round(sum(d["depreciacao_mensal"] for d in soc), 2),
            "dep_fisc_total": round(sum(d["depreciacao_mensal"] for d in fisc), 2),
            "diferenca_total": round(sum(d["diferenca_temporaria"] for d in diferencas), 2),
            "economia_irpj_csll_acelerado": round(
                sum(d["diferenca_temporaria"] for d in diferencas if d["diferenca_temporaria"] > 0) * 0.24, 2
            ),
        },
    })


@bp.route("/ativos/<aid>/creditos-pis-cofins", methods=["GET"])
def api_creditos(aid):
    c = db()
    row = c.execute("SELECT * FROM patrimonio_ativos WHERE id=?", (aid,)).fetchone()
    c.close()
    if not row:
        return jsonify({"error": "Não encontrado"}), 404
    ativo = dict(row)
    sched = dep_societaria(ativo)
    return jsonify(creditos_pis_cofins(ativo, sched))


@bp.route("/ativos/<aid>/icms", methods=["GET"])
def api_icms(aid):
    c = db()
    row = c.execute("SELECT * FROM patrimonio_ativos WHERE id=?", (aid,)).fetchone()
    c.close()
    if not row:
        return jsonify({"error": "Não encontrado"}), 404
    ativo = dict(row)
    valor_icms = float(ativo.get("valor_icms_aquisicao") or 0)
    prazo = min(48, int(ativo.get("vida_util_societaria_meses") or 48))
    credito_m = valor_icms / prazo if prazo > 0 else 0
    try:
        dt0 = datetime.strptime((ativo.get("data_aquisicao") or "")[:10], "%Y-%m-%d")
    except Exception:
        dt0 = datetime.now()
    cron, acum = [], 0.0
    for i in range(prazo):
        acum += credito_m
        cron.append({
            "mes": i + 1,
            "competencia": add_months(dt0, i).strftime("%Y-%m"),
            "credito_mensal": round(credito_m, 2),
            "credito_acumulado": round(acum, 2),
        })
    return jsonify({
        "valor_icms_total": valor_icms,
        "aliquota_pct": ativo.get("aliquota_icms", 0),
        "prazo_meses": prazo,
        "credito_mensal": round(credito_m, 2),
        "cronograma": cron,
    })


@bp.route("/ativos/<aid>/simular-venda", methods=["POST"])
def api_simular_venda(aid):
    d = request.json or {}
    c = db()
    row = c.execute("SELECT * FROM patrimonio_ativos WHERE id=?", (aid,)).fetchone()
    c.close()
    if not row:
        return jsonify({"error": "Não encontrado"}), 404
    return jsonify(resultado_venda(
        dict(row),
        float(d.get("valor_venda", 0)),
        d.get("data_venda", datetime.now().strftime("%Y-%m-%d")),
    ))


# ─── financiamentos ───────────────────────────────────────────────────────────

@bp.route("/financiamentos", methods=["GET"])
def listar_fins():
    c = db()
    rows = c.execute("""
        SELECT f.*, a.descricao as ativo_descricao, a.codigo as ativo_codigo
        FROM patrimonio_financiamentos f
        LEFT JOIN patrimonio_ativos a ON f.ativo_id=a.id
        ORDER BY f.created_at DESC
    """).fetchall()
    c.close()
    return jsonify([dict(r) for r in rows])


@bp.route("/financiamentos", methods=["POST"])
def criar_fin():
    d = request.json or {}
    fid = str(uuid.uuid4())
    now = datetime.now().isoformat()
    c = db()
    c.execute("""
        INSERT INTO patrimonio_financiamentos
        (id,ativo_id,banco,numero_contrato,valor_financiado,entrada,
         taxa_juros_mensal,sistema_amortizacao,prazo_meses,valor_residual_balao,
         data_inicio,created_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
    """, (fid, d.get("ativo_id"), d.get("banco"), d.get("numero_contrato"),
          d.get("valor_financiado", 0), d.get("entrada", 0),
          d.get("taxa_juros_mensal", 0), d.get("sistema_amortizacao", "price"),
          d.get("prazo_meses", 48), d.get("valor_residual_balao", 0),
          d.get("data_inicio"), now))
    c.commit()
    row = c.execute("SELECT * FROM patrimonio_financiamentos WHERE id=?", (fid,)).fetchone()
    c.close()
    return jsonify(dict(row)), 201


@bp.route("/financiamentos/<fid>", methods=["DELETE"])
def excluir_fin(fid):
    c = db()
    c.execute("DELETE FROM patrimonio_financiamentos WHERE id=?", (fid,))
    c.commit()
    c.close()
    return jsonify({"success": True})


@bp.route("/financiamentos/<fid>/cronograma", methods=["GET"])
def api_cronograma(fid):
    c = db()
    row = c.execute("SELECT * FROM patrimonio_financiamentos WHERE id=?", (fid,)).fetchone()
    c.close()
    if not row:
        return jsonify({"error": "Não encontrado"}), 404
    fin = dict(row)
    cron = cronograma_financiamento(fin)
    total_pago = sum(x["parcela_total"] for x in cron)
    total_juros = sum(x["juros"] for x in cron)
    return jsonify({
        "financiamento": fin,
        "cronograma": cron,
        "totais": {
            "total_pago": round(total_pago, 2),
            "total_juros": round(total_juros, 2),
            "total_amortizado": round(sum(x["amortizacao"] for x in cron), 2),
            "custo_efetivo_pct": round(
                total_juros / float(fin.get("valor_financiado") or 1) * 100, 2
            ),
        },
    })


# ─── simuladores ─────────────────────────────────────────────────────────────

@bp.route("/simular/compra", methods=["POST"])
def api_sim_compra():
    d = request.json or {}
    valor = float(d.get("valor_aquisicao", 0))
    vida = int(d.get("vida_util_meses", 60))
    residual = float(d.get("valor_residual", 0))
    regime = d.get("regime_tributario", "lucro_real")
    icms = float(d.get("valor_icms", 0))
    dep_m = (valor - residual) / vida if vida > 0 else 0

    if regime == "lucro_real":
        cred_pis_m = dep_m * 0.0165
        cred_cof_m = dep_m * 0.076
    else:
        cred_pis_m = cred_cof_m = 0

    econ_ir_m = dep_m * 0.15
    econ_cs_m = dep_m * 0.09
    econ_total = (cred_pis_m + cred_cof_m + econ_ir_m + econ_cs_m) * vida

    cenarios = {
        "avista": {
            "nome": "Compra à Vista",
            "saida_inicial": round(valor, 2),
            "total_saidas": round(valor, 2),
            "dep_mensal": round(dep_m, 2),
            "credito_pis_mensal": round(cred_pis_m, 2),
            "credito_cofins_mensal": round(cred_cof_m, 2),
            "economia_irpj_mensal": round(econ_ir_m, 2),
            "economia_csll_mensal": round(econ_cs_m, 2),
            "economia_tributaria_total": round(econ_total, 2),
            "credito_icms": round(icms, 2),
            "custo_liquido": round(valor - econ_total - icms, 2),
        }
    }

    fin = d.get("financiamento")
    if fin:
        entrada = float(fin.get("entrada", 0))
        taxa = float(fin.get("taxa_juros_mensal", 1)) / 100
        prazo = int(fin.get("prazo_meses", 48))
        saldo = valor - entrada
        pmt = (saldo * taxa * (1 + taxa) ** prazo / ((1 + taxa) ** prazo - 1)
               if taxa > 0 else saldo / prazo)
        total = entrada + pmt * prazo
        juros = total - valor
        cenarios["financiado"] = {
            "nome": "Compra Financiada",
            "saida_inicial": round(entrada, 2),
            "parcela_mensal": round(pmt, 2),
            "total_pago": round(total, 2),
            "total_juros": round(juros, 2),
            "economia_tributaria_total": round(econ_total, 2),
            "credito_icms": round(icms, 2),
            "custo_liquido": round(total - econ_total - icms, 2),
        }

    return jsonify({"cenarios": cenarios})


@bp.route("/simular/reforma-tributaria", methods=["POST"])
def api_sim_reforma():
    d = request.json or {}
    valor = float(d.get("valor_aquisicao", 0))
    vida = int(d.get("vida_util_meses", 60))
    regime = d.get("regime_tributario", "lucro_real")
    dep_m = valor / vida if vida > 0 else 0

    if regime == "lucro_real":
        cred_pc_atual = dep_m * 0.0925 * vida
    else:
        cred_pc_atual = 0

    econ_ircsll = dep_m * 0.24 * vida
    total_atual = cred_pc_atual + econ_ircsll

    CBS, IBS = 0.0888, 0.26
    cred_cbs = valor * CBS
    cred_ibs = valor * IBS
    total_reforma = cred_cbs + cred_ibs + econ_ircsll

    delta = total_reforma - total_atual
    return jsonify({
        "modelo_atual": {
            "credito_pis_cofins": round(cred_pc_atual, 2),
            "economia_irpj_csll": round(econ_ircsll, 2),
            "total": round(total_atual, 2),
        },
        "modelo_reforma": {
            "credito_cbs": round(cred_cbs, 2),
            "credito_ibs": round(cred_ibs, 2),
            "economia_irpj_csll": round(econ_ircsll, 2),
            "total": round(total_reforma, 2),
        },
        "diferenca": round(delta, 2),
        "impacto": "positivo" if delta >= 0 else "negativo",
        "variacao_pct": round(delta / total_atual * 100, 2) if total_atual > 0 else 0,
        "aliquotas_referencia": {"CBS": f"{CBS*100:.2f}%", "IBS": f"{IBS*100:.2f}%"},
    })


@bp.route("/simular/comparar-cenarios", methods=["POST"])
def api_comparar():
    """Compara cenários de manter/vender/substituir ativo."""
    d = request.json or {}
    valor = float(d.get("valor_aquisicao", 0))
    vida = int(d.get("vida_util_meses", 60))
    residual = float(d.get("valor_residual", 0))
    regime = d.get("regime_tributario", "lucro_real")
    icms = float(d.get("valor_icms", 0))
    taxa_desc = float(d.get("taxa_desconto_mensal", 1)) / 100

    dep_m = (valor - residual) / vida if vida > 0 else 0
    if regime == "lucro_real":
        cred_m = dep_m * 0.0925 + dep_m * 0.24
    else:
        cred_m = dep_m * 0.24

    # Fluxo de caixa mensal líquido (benefício tributário)
    fluxos = [cred_m] * vida
    vpl = sum(f / (1 + taxa_desc) ** (i + 1) for i, f in enumerate(fluxos))
    payback = valor / cred_m if cred_m > 0 else None

    # TIR aproximada via Newton-Raphson
    def calcular_tir(fluxos_tir):
        r = 0.01
        for _ in range(1000):
            npv = -valor + sum(f / (1 + r) ** (i + 1) for i, f in enumerate(fluxos_tir))
            dnpv = sum(-f * (i + 1) / (1 + r) ** (i + 2) for i, f in enumerate(fluxos_tir))
            if abs(dnpv) < 1e-10:
                break
            r -= npv / dnpv
            if r < -0.99:
                r = 0.0
        return r

    tir = calcular_tir(fluxos)
    roi = (sum(fluxos) - valor) / valor * 100 if valor > 0 else 0

    return jsonify({
        "parametros": {"valor": valor, "vida_meses": vida, "dep_mensal": round(dep_m, 2)},
        "beneficio_mensal": round(cred_m, 2),
        "beneficio_total": round(sum(fluxos), 2),
        "vpl": round(vpl, 2),
        "tir_mensal_pct": round(tir * 100, 4),
        "tir_anual_pct": round(((1 + tir) ** 12 - 1) * 100, 2),
        "payback_meses": round(payback, 1) if payback else None,
        "roi_pct": round(roi, 2),
        "custo_liquido": round(valor - sum(fluxos) - icms, 2),
    })


# ─── dashboard e consolidado ──────────────────────────────────────────────────

@bp.route("/dashboard", methods=["GET"])
def api_dashboard():
    c = db()
    ativos = [dict(r) for r in c.execute("SELECT * FROM patrimonio_ativos").fetchall()]
    fins = [dict(r) for r in c.execute("SELECT * FROM patrimonio_financiamentos").fetchall()]
    c.close()

    pat_bruto = sum(float(a.get("valor_aquisicao") or 0) for a in ativos)
    total_dep, total_cred_pc, total_icms = 0.0, 0.0, 0.0

    resumo = []
    for ativo in ativos:
        soc = dep_societaria(ativo)
        dep_acum = sum(d["depreciacao_mensal"] for d in soc)
        vlc = max(float(ativo.get("valor_aquisicao") or 0) - dep_acum,
                  float(ativo.get("valor_residual_estimado") or 0))
        cred = creditos_pis_cofins(ativo, soc)
        total_dep += dep_acum
        total_cred_pc += cred["total_credito"]
        total_icms += float(ativo.get("valor_icms_aquisicao") or 0)
        resumo.append({
            "id": ativo["id"],
            "codigo": ativo["codigo"],
            "descricao": ativo["descricao"],
            "tipo_ativo": ativo.get("tipo_ativo"),
            "valor_aquisicao": float(ativo.get("valor_aquisicao") or 0),
            "depreciacao_acumulada": round(dep_acum, 2),
            "valor_liquido_contabil": round(vlc, 2),
            "pct_depreciado": round(dep_acum / float(ativo.get("valor_aquisicao") or 1) * 100, 1),
            "status": ativo.get("status"),
        })

    total_financiado = sum(float(f.get("valor_financiado") or 0) for f in fins)
    total_juros = sum(sum(p["juros"] for p in cronograma_financiamento(f)) for f in fins)

    return jsonify({
        "totais": {
            "patrimonio_bruto": round(pat_bruto, 2),
            "depreciacao_acumulada": round(total_dep, 2),
            "patrimonio_liquido": round(pat_bruto - total_dep, 2),
            "credito_pis_cofins_total": round(total_cred_pc, 2),
            "credito_icms_total": round(total_icms, 2),
            "total_financiado": round(total_financiado, 2),
            "total_juros_fins": round(total_juros, 2),
            "qtd_ativos": len(ativos),
            "qtd_ativos": sum(1 for a in ativos if a.get("status") == "ativo"),
        },
        "ativos": sorted(resumo, key=lambda x: x["valor_aquisicao"], reverse=True),
    })


@bp.route("/resultado-consolidado", methods=["GET"])
def api_consolidado():
    c = db()
    ativos = [dict(r) for r in c.execute(
        "SELECT * FROM patrimonio_ativos WHERE status='ativo'").fetchall()]
    fins = [dict(r) for r in c.execute("SELECT * FROM patrimonio_financiamentos").fetchall()]
    c.close()

    g = {"credito_pis": 0, "credito_cofins": 0,
         "beneficio_irpj": 0, "beneficio_csll": 0, "credito_icms": 0}
    p = {"juros_financiamentos": 0}

    for ativo in ativos:
        soc = dep_societaria(ativo)
        cr = creditos_pis_cofins(ativo, soc)
        g["credito_pis"] += cr["total_pis"]
        g["credito_cofins"] += cr["total_cofins"]
        dep_tot = sum(d["depreciacao_mensal"] for d in soc)
        g["beneficio_irpj"] += dep_tot * 0.15
        g["beneficio_csll"] += dep_tot * 0.09
        g["credito_icms"] += float(ativo.get("valor_icms_aquisicao") or 0)

    for fin in fins:
        p["juros_financiamentos"] += sum(x["juros"] for x in cronograma_financiamento(fin))

    tg = sum(g.values())
    tp = sum(p.values())
    return jsonify({
        "ganhos": {k: round(v, 2) for k, v in g.items()},
        "perdas": {k: round(v, 2) for k, v in p.items()},
        "total_ganhos": round(tg, 2),
        "total_perdas": round(tp, 2),
        "resultado_liquido": round(tg - tp, 2),
    })


# ─── exportação ──────────────────────────────────────────────────────────────

@bp.route("/relatorio/excel", methods=["GET"])
def api_excel():
    c = db()
    ativos = [dict(r) for r in c.execute("SELECT * FROM patrimonio_ativos").fetchall()]
    c.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ativos"

    hdr_fill = PatternFill(start_color="1a2744", end_color="1a2744", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)

    colunas = [
        "Código", "Descrição", "Tipo", "Categoria", "Placa", "Chassi",
        "Dt. Aquisição", "Dt. Operação", "V.Útil Fiscal(m)", "V.Útil Soc.(m)",
        "Valor Aquisição", "Valor Residual", "ICMS Aquisição", "Regime", "Status",
        "Filial", "Unid. Negócio",
    ]
    campos = [
        "codigo", "descricao", "tipo_ativo", "categoria", "placa", "chassi",
        "data_aquisicao", "data_entrada_operacao", "vida_util_fiscal_meses",
        "vida_util_societaria_meses", "valor_aquisicao", "valor_residual_estimado",
        "valor_icms_aquisicao", "regime_tributario", "status", "filial", "unidade_negocio",
    ]

    for ci, col in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for ri, a in enumerate(ativos, 2):
        for ci, campo in enumerate(campos, 1):
            ws.cell(row=ri, column=ci, value=a.get(campo))

    for col_cells in ws.columns:
        w = max((len(str(cell.value or "")) for cell in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(w + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"patrimonio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )

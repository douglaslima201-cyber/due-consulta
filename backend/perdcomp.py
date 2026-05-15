"""
PER/DCOMP Analyzer — Backend Blueprint
Extração de PDFs + Motor de Compliance + Vinculação PER↔DCOMP + Download eCAC
"""
import asyncio
import io
import os
import re
import threading
import uuid
from datetime import date
from pathlib import Path
from flask import Blueprint, request, jsonify, send_from_directory, send_file

try:
    import pdfplumber
    PDF_OK = True
except ImportError:
    PDF_OK = False

bp = Blueprint('perdcomp', __name__)

# ─── ECAC DOWNLOAD ────────────────────────────────────────────────────────────
_ECAC_JOBS: dict[str, dict] = {}
_ECAC_DIR  = Path(__file__).parent / "ecac_downloads"
_ECAC_DIR.mkdir(exist_ok=True)
_ECAC_URL  = "https://cav.receita.fazenda.gov.br/eCAC/"
_CHROME_PROFILE = os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\User Data")
_CDP_PORT = 9223
_BACKEND_PORT = int(os.environ.get("PORT", 5000))

_jobs_lock     = threading.Lock()
_capturas_lock = threading.Lock()

def _elog(job_id: str, msg: str):
    _ECAC_JOBS[job_id]["log"].append(msg)
    print(f"[ECAC][{job_id[:8]}] {msg}")

def _extrair_cookies_chrome_windows(dominios: list[str]) -> list[dict]:
    """Extrai e descriptografa cookies do Chrome no Windows via DPAPI + AES-GCM."""
    import sqlite3, json, base64, tempfile
    import win32crypt
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM

    # 1. Ler chave de encriptação do Local State
    local_state_path = Path(_CHROME_PROFILE) / "Local State"
    with open(local_state_path, "r", encoding="utf-8") as f:
        ls = json.load(f)

    enc_key_b64 = ls.get("os_crypt", {}).get("encrypted_key", "")
    if not enc_key_b64:
        raise RuntimeError("Chave de encriptação não encontrada no perfil do Chrome.")

    enc_key = base64.b64decode(enc_key_b64)[5:]  # Remove prefixo "DPAPI"
    key = win32crypt.CryptUnprotectData(enc_key, None, None, None, 0)[1]

    # 2. Copiar banco de cookies — Chrome mantém o arquivo bloqueado,
    #    então usamos ctypes com FILE_SHARE_READ para ler mesmo assim
    import ctypes, ctypes.wintypes

    for caminho_cookies in [
        Path(_CHROME_PROFILE) / "Default" / "Network" / "Cookies",
        Path(_CHROME_PROFILE) / "Default" / "Cookies",
    ]:
        if caminho_cookies.exists():
            break
    else:
        raise RuntimeError("Banco de cookies do Chrome não encontrado.")

    # 3. Consultar cookies diretamente via SQLite imutable (ignora lock do Chrome)
    dominios_limpos = [d.replace("https://", "").replace("http://", "") for d in dominios]
    filtro = " OR ".join(f"host_key LIKE '%{d}%'" for d in dominios_limpos)

    uri = f"file:{caminho_cookies.as_posix()}?mode=ro&immutable=1"
    conn = sqlite3.connect(uri, uri=True)
    rows = conn.execute(
        f"SELECT host_key, name, encrypted_value, path, expires_utc, is_secure, is_httponly "
        f"FROM cookies WHERE {filtro}"
    ).fetchall()
    conn.close()

    # 4. Descriptografar cada cookie
    cookies = []
    for host, name, enc_val, path, expires_utc, is_secure, is_httponly in rows:
        try:
            if enc_val[:3] in (b'v10', b'v11'):
                nonce = enc_val[3:15]
                value = AESGCM(key).decrypt(nonce, enc_val[15:], None).decode('utf-8')
            else:
                value = win32crypt.CryptUnprotectData(enc_val, None, None, None, 0)[1].decode('utf-8')

            cookie: dict = {
                "name": name, "value": value,
                "domain": host, "path": path or "/",
                "secure": bool(is_secure), "httpOnly": bool(is_httponly),
                "sameSite": "Lax",
            }
            # Chrome usa microsegundos desde 1601-01-01 → converter para Unix epoch
            if expires_utc > 0:
                cookie["expires"] = int(expires_utc / 1_000_000) - 11_644_473_600
            cookies.append(cookie)
        except Exception:
            pass

    return cookies


def _encontrar_chrome() -> str | None:
    candidatos = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
    ]
    for c in candidatos:
        if Path(c).exists():
            return c
    return None

def _chrome_bloqueado() -> bool:
    return _encontrar_chrome() is None

async def _ecac_download_async(job_id: str, periodo_ini: str, periodo_fim: str):
    from playwright.async_api import async_playwright, TimeoutError as PwTimeout

    job = _ECAC_JOBS[job_id]
    dest = _ECAC_DIR / "entrada"
    dest.mkdir(exist_ok=True)

    _elog(job_id, "Abrindo Chrome para autenticação no eCAC...")
    job["status"] = "aguardando_login"

    async with async_playwright() as p:
        # Lança Chrome normal — sem debug port, sem flags extras
        # Para A1 (software), o certificado do Windows Certificate Store fica acessível
        browser = await p.chromium.launch(
            channel="chrome",
            headless=False,
            args=[
                "--start-maximized",
                "--disable-blink-features=AutomationControlled",
                "--disable-features=BlockInsecurePrivateNetworkRequests",
                "--disable-features=PrivateNetworkAccessSendPreflights",
            ],
        )
        ctx = await browser.new_context(
            accept_downloads=True,
            viewport={"width": 1280, "height": 900},
        )
        page = await ctx.new_page()

        # Ir para eCAC
        try:
            await page.goto(_ECAC_URL, wait_until="domcontentloaded", timeout=30000)
        except PwTimeout:
            pass

        _elog(job_id, "Chrome aberto. Selecione seu certificado A1 e faça login no eCAC (até 5 min).")

        # Aguardar login — detecta retorno ao domínio do eCAC após autenticação
        try:
            await page.wait_for_url("*cav.receita*", timeout=300_000)
            await asyncio.sleep(3)
            _elog(job_id, "Login detectado. Navegando para PER/DCOMP...")
        except PwTimeout:
            _elog(job_id, "Timeout: login não completado.")
            job["status"] = "erro"; await browser.close(); return

        async def clicar(sels: list[str], label: str, timeout: int = 5000) -> bool:
            for sel in sels:
                try:
                    el = page.locator(sel).first
                    if await el.is_visible(timeout=timeout):
                        await el.click()
                        await asyncio.sleep(1)
                        return True
                except Exception:
                    continue
            return False

        # ── Passo 1: Restituição e Compensação ───────────────────────────────
        _elog(job_id, "Clicando em 'Restituição e Compensação'...")
        ok = await clicar([
            'a:text-matches("Restituição e Compensação", "i")',
            'a:text-matches("Restituição", "i")',
            'span:text-matches("Restituição e Compensação", "i")',
        ], "Restituição e Compensação")
        if ok:
            await page.wait_for_load_state("networkidle", timeout=15000)

        # ── Passo 2: Acessar PERDCOMP Web ────────────────────────────────────
        _elog(job_id, "Clicando em 'Acessar PERDCOMP Web'...")
        ok = await clicar([
            'a:text-matches("Acessar.*PERDCOMP.*Web", "i")',
            'a:text-matches("PERDCOMP Web", "i")',
            'a[href*="perdcomp" i]',
            'button:text-matches("PERDCOMP", "i")',
        ], "PERDCOMP Web")
        if ok:
            await page.wait_for_load_state("networkidle", timeout=15000)

        # ── Passo 3: Visualizar Documentos ───────────────────────────────────
        _elog(job_id, "Clicando em 'Visualizar Documentos'...")
        ok = await clicar([
            'a:text-matches("Visualizar.*Documento", "i")',
            'button:text-matches("Visualizar.*Documento", "i")',
            ':text("Visualizar documentos")',
        ], "Visualizar Documentos")
        if ok:
            await page.wait_for_load_state("networkidle", timeout=15000)

        # ── Passo 4: Aba Documentos Entregues ────────────────────────────────
        _elog(job_id, "Clicando na aba 'Documentos Entregues'...")
        await clicar([
            'a:text-matches("Documentos Entregues", "i")',
            'button:text-matches("Documentos Entregues", "i")',
            '[role="tab"]:text-matches("Documentos Entregues", "i")',
            'li:text-matches("Documentos Entregues", "i")',
        ], "Documentos Entregues")
        await asyncio.sleep(2)

        # ── Passo 5: Preencher datas (DD/MM/AAAA) ────────────────────────────
        _elog(job_id, f"Preenchendo período: {periodo_ini} → {periodo_fim}...")

        async def preencher(sels: list[str], valor: str) -> bool:
            for sel in sels:
                try:
                    el = page.locator(sel).first
                    if await el.is_visible(timeout=2000):
                        await el.triple_click()
                        await el.type(valor, delay=50)
                        return True
                except Exception:
                    continue
            return False

        sels_data_ini = [
            'input[id*="dataInicial" i]', 'input[name*="dataInicial" i]',
            'input[placeholder*="inicial" i]', 'input[placeholder*="início" i]',
            'input[id*="inicio" i]', 'input[id*="dtIni" i]',
        ]
        sels_data_fim = [
            'input[id*="dataFinal" i]', 'input[name*="dataFinal" i]',
            'input[placeholder*="final" i]', 'input[placeholder*="fim" i]',
            'input[id*="fim" i]', 'input[id*="dtFin" i]',
        ]

        await preencher(sels_data_ini, periodo_ini)
        await preencher(sels_data_fim, periodo_fim)

        # ── Passo 6: Pesquisar ────────────────────────────────────────────────
        _elog(job_id, "Pesquisando declarações...")
        for sel in [
            'button:text-matches("Pesquisar|Consultar|Buscar|Filtrar", "i")',
            'input[type="submit"]', 'button[type="submit"]',
        ]:
            try:
                btn = page.locator(sel).first
                if await btn.is_visible(timeout=2000):
                    await btn.click()
                    await page.wait_for_load_state("networkidle", timeout=20000)
                    _elog(job_id, "Pesquisa executada.")
                    break
            except Exception:
                continue

        await asyncio.sleep(2)

        # ── Passo 7: Clicar em "Imprimir Documento" para cada declaração ─────
        job["status"] = "baixando"
        arquivos_baixados: list[str] = []

        icones_imprimir = await page.locator(
            'button[title*="Imprimir" i], a[title*="Imprimir" i], '
            'img[alt*="Imprimir" i], img[title*="Imprimir" i], '
            'button:text-matches("Imprimir", "i"), a:text-matches("Imprimir", "i")'
        ).all()

        if not icones_imprimir:
            _elog(job_id, "Nenhum ícone de impressão encontrado. Navegue manualmente e clique em 'Já estou na lista' quando os documentos aparecerem.")
            job["status"] = "aguardando_navegacao"
            for _ in range(600):
                await asyncio.sleep(1)
                if job.get("usuario_confirmou"):
                    arquivos_baixados = [f.name for f in dest.glob("*.pdf")]
                    break
        else:
            _elog(job_id, f"{len(icones_imprimir)} declaração(ões) encontrada(s). Salvando PDFs...")
            for i, icone in enumerate(icones_imprimir):
                try:
                    # "Imprimir" no eCAC abre nova aba/popup com o documento
                    async with ctx.expect_page() as nova_pg_info:
                        await icone.click()
                    nova_pg = await nova_pg_info.value
                    await nova_pg.wait_for_load_state("networkidle", timeout=20000)
                    await asyncio.sleep(1)

                    nome = f"perdcomp_{i+1:03d}.pdf"
                    # page.pdf() só funciona headless; usa CDP diretamente (funciona em headed)
                    import base64 as _b64
                    cdp = await ctx.new_cdp_session(nova_pg)
                    result = await cdp.send("Page.printToPDF", {
                        "printBackground": True,
                        "paperWidth": 8.27,
                        "paperHeight": 11.69,
                    })
                    await cdp.detach()
                    (dest / nome).write_bytes(_b64.b64decode(result["data"]))
                    await nova_pg.close()

                    arquivos_baixados.append(nome)
                    job["arquivos"] = arquivos_baixados[:]
                    _elog(job_id, f"[{i+1}/{len(icones_imprimir)}] Salvo: {nome}")
                except Exception as exc:
                    _elog(job_id, f"[{i+1}] Falha ao salvar: {exc}")
                await asyncio.sleep(0.5)

        job["arquivos"] = arquivos_baixados
        job["status"] = "concluido" if arquivos_baixados else "sem_arquivos"
        _elog(job_id, f"Concluído. {len(arquivos_baixados)} arquivo(s) na pasta de entrada.")
        await browser.close()

def _ecac_thread(job_id: str, periodo_ini: str, periodo_fim: str):
    asyncio.run(_ecac_download_async(job_id, periodo_ini, periodo_fim))

# ─── UTILITÁRIOS ──────────────────────────────────────────────────────────────

def parse_valor(s: str) -> float:
    if not s:
        return 0.0
    s = re.sub(r'[R$\s\t]', '', str(s)).replace('.', '').replace(',', '.')
    try:
        return float(s)
    except Exception:
        return 0.0

def extrair_texto(file_bytes: bytes) -> str:
    if not PDF_OK:
        return ""
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            paginas = []
            for p in pdf.pages:
                t = p.extract_text(x_tolerance=3, y_tolerance=3)
                if t:
                    paginas.append(t)
            return "\n".join(paginas)
    except Exception as e:
        return f"ERRO_EXTRACAO: {e}"

def primeiro_match(texto: str, padroes: list, flags=re.IGNORECASE) -> str | None:
    for pat in padroes:
        m = re.search(pat, texto, flags)
        if m:
            return m.group(1).strip()
    return None

# ─── DETECÇÃO DE TIPO ─────────────────────────────────────────────────────────

def detectar_tipo(texto: str, tl: str) -> str:
    # Campo explícito do PDF — mais confiável (encoding pode conter '?' no lugar de acentos)
    m = re.search(r'Tipo de Documento\s+([^\n]+)', texto, re.IGNORECASE)
    if m:
        td = m.group(1).strip().lower()
        if 'cancelamento' in td:
            return 'Cancelamento'
        if 'compensa' in td:
            return 'Compensação'
        if 'ressarcimento' in td:
            return 'Ressarcimento'
        if 'restitui' in td:
            return 'Restituição'
    # Fallback: varredura nas primeiras linhas
    primeiras = texto[:800].lower()
    if 'declara' in primeiras and 'compensa' in primeiras:
        return 'Compensação'
    if 'ressarcimento' in primeiras and 'pedido' in primeiras:
        return 'Ressarcimento'
    if 'restitui' in primeiras and 'pedido' in primeiras:
        return 'Restituição'
    if 'compensa' in tl:
        return 'Compensação'
    if 'ressarcimento' in tl:
        return 'Ressarcimento'
    return 'PER/DCOMP'

def detectar_tributo(texto: str) -> str | None:
    padroes = [
        (r'PIS[/\s]Pasep', 'PIS'), (r'\bPIS\b', 'PIS'),
        (r'COFINS\s+N[ãa]o[- ]Cumulativ', 'COFINS'), (r'\bCOFINS\b', 'COFINS'),
        (r'\bCSLL\b', 'CSLL'), (r'\bIRPJ\b', 'IRPJ'), (r'\bIRRF\b', 'IRRF'),
        (r'\bCSRF\b', 'CSRF'), (r'\bIPI\b', 'IPI'), (r'\bINSS\b', 'INSS'),
    ]
    for pat, nome in padroes:
        if re.search(pat, texto, re.IGNORECASE):
            return nome
    return None

# ─── EXTRAÇÃO DE VALORES POR TIPO ────────────────────────────────────────────

_VAL = r'([\d]{1,3}(?:\.[\d]{3})*,\d{2})'

def _val(texto: str, labels: list) -> float:
    """Busca valor monetário após um label — suporta múltiplas linhas e espaços."""
    for label in labels:
        m = re.search(label + r'[\s\t\n:R$]*' + _VAL, texto, re.IGNORECASE | re.MULTILINE)
        if m:
            return parse_valor(m.group(1))
    return 0.0

def _normalizar_competencia(comp: str | None) -> str:
    """Normaliza competência para MM/AAAA para comparação uniforme.
    '3º Trimestre de 2025' → '09/2025' (último mês do trimestre)."""
    if not comp:
        return ""
    m = re.search(r'(\d)[°º\s]*\s*[Tt]rimestre\s+de\s+(\d{4})', comp)
    if m:
        tri, ano = int(m.group(1)), m.group(2)
        return {1: f"03/{ano}", 2: f"06/{ano}", 3: f"09/{ano}", 4: f"12/{ano}"}.get(tri, comp)
    m = re.search(r'(\d{2})/(\d{4})', comp)
    if m:
        return comp
    return comp

def _extrair_competencia_texto(texto: str) -> str | None:
    """Extrai competência — suporta formato eCAC com Ano + Trimestre em linhas separadas."""
    # Formato eCAC: "Ano 2024\nTrimestre 4? Trimestre" (? = º com encoding corrompido)
    m = re.search(r'Ano\s+(\d{4})\s+Trimestre\s+(\d+)', texto, re.IGNORECASE)
    if m:
        ano, tri = m.group(1), int(m.group(2))
        return f"{tri}º Trimestre de {ano}"
    # Formato clássico: "3º Trimestre de 2025"
    m = re.search(r'(\d[°º?]?\s*[Tt]rimestre\s+de\s+\d{4})', texto)
    if m:
        return m.group(1).strip()
    # Formato mensal
    return primeiro_match(texto, [
        r'Per[íi?]odo de Apura[çc?][aã?]o[:\s]+(\d{2}/\d{4})',
        r'Compet[eê?]ncia[:\s]+(\d{2}/\d{4})',
        r'Per[íi?]odo[:\s]+(\d{2}/\d{4})',
    ])

def extrair_referencia_per(texto: str) -> str | None:
    """Extrai o número do PER referenciado dentro de uma DCOMP."""
    # Formato longo Receita Federal: XXXXX.XXXXX.XXXXXX.X.X.XX-XXXX
    pat_num = r'(\d{5}\.\d{5}\.\d{6}\.\d+\.\d+\.\d{2}-\d{4})'
    return primeiro_match(texto, [
        # PERDCOMP 8.3+: "N° do PER/DCOMP Inicial XXXXX..." (° vira ? no encoding)
        r'N[.°?º]\s*do PER/DCOMP Inicial\s+' + pat_num,
        r'N[úu]mero do Processo[:\s]+' + pat_num,
        r'(?:N[úu]mero do )?PER[:\s]+' + pat_num,
        r'Pedido de Ressarcimento[:\s]+' + pat_num,
        r'Processo do Cr[eé]dito[:\s]+' + pat_num,
        r'N[úu]mero do Cr[eé]dito[:\s]+' + pat_num,
        r'Cr[eé]dito[^\n]{0,80}N[úu]mero[:\s]+' + pat_num,
    ])

def extrair_competencia_credito(texto: str) -> str | None:
    """Extrai competência do crédito (para DCOMPs)."""
    m = re.search(r'Per[íi]odo de Apura[çc][aã]o do Cr[eé]dito[:\s]+([^\n]+)', texto, re.IGNORECASE)
    if m:
        val = m.group(1).strip()
        tri = re.search(r'\d[°º]?\s*[Tt]rimestre\s+de\s+\d{4}', val)
        if tri:
            return tri.group(0)
        data = re.search(r'\d{2}/\d{4}', val)
        if data:
            return data.group(0)
    return _extrair_competencia_texto(texto)

def extrair_dcomp(texto: str, tl: str) -> dict:
    valor_credito = _val(texto, [
        # PERDCOMP 8.31+
        r'Valor Original do Cr[eé?]dito Inicial',
        # versões anteriores
        r'Valor do Cr[eé]dito Dispon[íi]vel',
        r'Valor do Cr[eé]dito Original',
        r'Valor Total do Cr[eé]dito',
        r'Cr[eé]dito Dispon[íi]vel',
        r'Valor do Cr[eé]dito',
        r'Total do Cr[eé]dito',
    ])
    credito_na_transmissao = _val(texto, [
        # PERDCOMP 8.31+
        r'Cr[eé?]dito Atualizado',
        r'Cr[eé?]dito Original na Data de Entrega',
        # versões anteriores
        r'Cr[eé?]dito na Data de Transmiss[aã?]o',
    ])
    valor_compensado = _val(texto, [
        # PERDCOMP 8.31+
        r'Total do Cr[eé?]dito Original Utilizado neste Documento',
        # versões anteriores
        r'Valor Utilizado nesta DCOMP',
        r'Valor da Declara[çc?][aã?]o de Compensa[çc?][aã?]o',
        r'Valor Compensado',
        r'Valor da Compensa[çc?][aã?]o',
        r'Valor do D[eé?]bito Compensado',
        r'Valor Objeto da Compensa[çc?][aã?]o',
        r'Valor do D[eé?]bito',
    ])
    saldo = _val(texto, [
        # PERDCOMP 8.31+
        r'Saldo do Cr[eé?]dito Original',
        # versões anteriores
        r'Saldo do Cr[eé?]dito ap[oó?]s',
        r'Saldo do Cr[eé?]dito',
        r'Saldo Remanescente',
        r'Saldo a Compensar',
        r'Saldo Dispon[íi?]vel',
    ])
    if valor_compensado == 0 and valor_credito > 0 and saldo > 0:
        valor_compensado = max(0.0, valor_credito - saldo)
    if saldo == 0 and valor_credito > 0:
        saldo = max(0.0, valor_credito - valor_compensado)
    return {
        "valor_credito":           round(valor_credito, 2),
        "credito_na_transmissao":  round(credito_na_transmissao, 2),
        "valor_compensado":        round(valor_compensado, 2),
        "valor_ressarcido":        0.0,
        "saldo_remanescente":      round(saldo, 2),
    }

def extrair_per_ressarcimento(texto: str, tl: str) -> dict:
    # PIS/COFINS não-cumulativo: labels específicos do eCAC
    valor_credito = _val(texto, [
        r'Valor do Pedido de Ressarcimento',
        r'Cr[eé?]dito Pass[íi?]vel de Ressarcimento',
        r'Valor do Ressarcimento Requerido',
        r'Valor do Ressarcimento',
        r'Cr[eé?]dito a Ressarcir',
        r'Valor do Cr[eé?]dito a Ressarcir',
        r'Total do Cr[eé?]dito do Per[íi?]odo',
        r'Cr[eé?]dito do Per[íi?]odo',
        r'Valor do Cr[eé?]dito Apurado',
        r'Valor Apurado',
        r'Valor Solicitado',
        r'Valor do Cr[eé?]dito',
        r'Total do Cr[eé?]dito',
    ])
    valor_ressarcido = _val(texto, [
        r'Valor Ressarcido', r'Valor Pago', r'Valor Deferido',
        r'Valor Creditado',
    ])
    # Busca case-sensitive no totalizador em maiúsculas (resume todas as parcelas do PER)
    _VAL_PAT = r'([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})'
    m_s = re.search(r'SALDO DO CR.DITO\s+' + _VAL_PAT, texto)
    saldo = parse_valor(m_s.group(1)) if m_s else _val(texto, [
        r'Saldo Remanescente', r'Saldo a Ressarcir',
        r'Saldo do Cr[eé]dito', r'Saldo Dispon[íi]vel',
    ])
    if saldo == 0 and valor_credito > 0:
        saldo = max(0.0, valor_credito - valor_ressarcido)
    return {
        "valor_credito":          round(valor_credito, 2),
        "credito_na_transmissao": 0.0,
        "valor_compensado":       0.0,
        "valor_ressarcido":       round(valor_ressarcido, 2),
        "saldo_remanescente":     round(saldo, 2),
    }

def extrair_per_restituicao(texto: str, tl: str) -> dict:
    valor_credito = _val(texto, [
        r'Valor da Restituição', r'Valor da Restituicao',
        r'Valor Solicitado', r'Valor do Cr[eé]dito', r'Valor a Restituir',
    ])
    valor_ressarcido = _val(texto, [r'Valor Restitu[íi]do', r'Valor Pago', r'Valor Deferido'])
    saldo = _val(texto, [r'Saldo Remanescente', r'Saldo a Restituir'])
    if saldo == 0 and valor_credito > 0:
        saldo = max(0.0, valor_credito - valor_ressarcido)
    return {
        "valor_credito":          round(valor_credito, 2),
        "credito_na_transmissao": 0.0,
        "valor_compensado":       0.0,
        "valor_ressarcido":       round(valor_ressarcido, 2),
        "saldo_remanescente":     round(saldo, 2),
    }

# ─── CAMPOS COMUNS ────────────────────────────────────────────────────────────

def extrair_numero(texto: str) -> str | None:
    # Formato eCAC: XXXXX.XXXXX.XXXXXX.X.X.XX-XXXX
    pat = r'(\d{5}\.\d{5}\.\d{6}\.\d+\.\d+\.\d{2}-\d{4})'
    return primeiro_match(texto, [
        r'N[úu]mero do Processo[:\s]+' + pat,
        r'Processo[:\s]+' + pat,
        r'N[úu]mero[:\s]+' + pat,
        pat,  # fallback: qualquer ocorrência do padrão
    ])

def extrair_data_tx(texto: str) -> str | None:
    return primeiro_match(texto, [
        r'Data de Transmiss[aã]o[:\s]+(\d{2}/\d{2}/\d{4})',
        r'Transmitido em[:\s]+(\d{2}/\d{2}/\d{4})',
        r'Data de Envio[:\s]+(\d{2}/\d{2}/\d{4})',
        r'Data/Hora[:\s]+(\d{2}/\d{2}/\d{4})',
    ])

def extrair_situacao(tl: str) -> str | None:
    for s in ['Homologado', 'Não Homologado', 'Deferido', 'Indeferido',
              'Cancelado', 'Em Análise', 'Em análise', 'Ativo',
              'Pendente', 'Em processamento']:
        if s.lower() in tl:
            return s
    return None

def extrair_empresa(texto: str) -> tuple[str | None, str | None]:
    """Retorna (nome_empresa, cnpj) extraídos do PDF."""
    empresa = primeiro_match(texto, [r'Nome Empresarial\s+([^\n]+)', r'Raz[aã]o Social\s+([^\n]+)'])
    cnpj    = primeiro_match(texto, [r'CNPJ\s+([\d]{2}\.[\d]{3}\.[\d]{3}/[\d]{4}-[\d]{2})'])
    return (empresa.strip() if empresa else None, cnpj)

# ─── MONTAGEM DO REGISTRO ─────────────────────────────────────────────────────

def extrair_registro(texto: str, nome: str) -> dict:
    tl = texto.lower()
    tipo = detectar_tipo(texto, tl)

    if tipo == 'Compensação':
        vals = extrair_dcomp(texto, tl)
        referencia_per = extrair_referencia_per(texto)
        competencia = extrair_competencia_credito(texto)
    elif tipo == 'Ressarcimento':
        vals = extrair_per_ressarcimento(texto, tl)
        referencia_per = None
        competencia = _extrair_competencia_texto(texto)
    else:
        vals = extrair_per_restituicao(texto, tl)
        referencia_per = None
        competencia = _extrair_competencia_texto(texto)

    # Campo "PER/DCOMP Retificador Sim/Não" — verifica o valor, não só a presença da palavra
    m_ret = re.search(r'retificador[a]?\s+(sim|n.o)', tl, re.IGNORECASE)
    retificador = bool(m_ret and m_ret.group(1).lower().startswith('s'))
    numero_original = None
    if retificador:
        numero_original = primeiro_match(texto, [
            r'Processo Original[:\s]+(\d{5}\.\d{6}/\d{4}-\d{2})',
            r'N[úu]mero Original[:\s]+(\d{5}\.\d{6}/\d{4}-\d{2})',
            r'Retifica[çc][aã]o de[:\s]+(\d{5}\.\d{6}/\d{4}-\d{2})',
        ])

    return {
        "arquivo":          nome,
        "numero":           extrair_numero(texto),
        "numero_original":  numero_original,
        "tipo":             tipo,
        "tributo":          detectar_tributo(texto),
        "competencia":      competencia,
        "data_transmissao": extrair_data_tx(texto),
        "situacao":         extrair_situacao(tl),
        "retificador":      retificador,
        "referencia_per":   referencia_per,
        "empresa":          extrair_empresa(texto)[0],
        "cnpj":             extrair_empresa(texto)[1],
        **vals,
    }

# ─── RETIFICADORAS DCOMP ─────────────────────────────────────────────────────

def _marcar_retificadoras_dcomp(registros: list) -> None:
    """Marca DCOMPs substituídas por retificadoras (mesmo tributo + competência)."""
    from collections import defaultdict
    dcomps = [r for r in registros if r["tipo"] == "Compensação"]
    grupos: dict[tuple, list] = defaultdict(list)
    for r in dcomps:
        chave = (r.get("tributo"), _normalizar_competencia(r.get("competencia") or ""))
        grupos[chave].append(r)
    for chave, grupo in grupos.items():
        rets = [r for r in grupo if r.get("retificador")]
        orig = [r for r in grupo if not r.get("retificador")]
        if rets and orig:
            efetivo = rets[-1]
            subs = orig + rets[:-1]
            efetivo.setdefault("_efetivo", True)
            efetivo.setdefault("_substitui", [r.get("arquivo") for r in subs])
            for s in subs:
                s["_efetivo"] = False
                s["_substituido_por"] = efetivo.get("arquivo")
        else:
            for r in grupo:
                r.setdefault("_efetivo", True)

# ─── VINCULAÇÃO PER ↔ DCOMP ──────────────────────────────────────────────────

def _normalizar_num(n: str | None) -> str:
    if not n:
        return ""
    return re.sub(r'[\s\-/\.]', '', n).upper()

def vincular_pers_dcomps(registros: list) -> tuple[list, list]:
    """
    Vincula DCOMPs aos PERs de origem.
    Retificadoras substituem os originais.
    """
    pers_brutos   = [r for r in registros if r["tipo"] in ("Ressarcimento", "Restituição")]
    dcomps        = [r for r in registros if r["tipo"] == "Compensação"]

    # ── Passo 1: resolver retificadoras entre os PERs ────────────────────────
    # Chave de agrupamento: tributo + competencia (identifica o mesmo crédito)
    per_por_chave: dict[tuple, list] = {}
    for p in pers_brutos:
        chave = (p.get("tributo"), p.get("competencia"))
        per_por_chave.setdefault(chave, []).append(p)

    # Reindexar usando competência NORMALIZADA como chave
    per_por_chave_norm: dict[tuple, list] = {}
    for (trib, comp), grupo in per_por_chave.items():
        chave_norm = (trib, _normalizar_competencia(comp))
        per_por_chave_norm.setdefault(chave_norm, []).extend(grupo)
    per_por_chave = per_por_chave_norm

    per_por_numero: dict[str, dict] = {}   # numero normalizado → PER efetivo
    grupos_per: list[dict] = []

    for chave, grupo in per_por_chave.items():
        originais    = [p for p in grupo if not p.get("retificador")]
        retificadoras = [p for p in grupo if p.get("retificador")]

        if retificadoras:
            efetivo = retificadoras[-1]          # mais recente
            substituidos = originais + retificadoras[:-1]
        elif originais:
            efetivo = originais[-1]
            substituidos = originais[:-1]
        else:
            continue

        efetivo["_efetivo"] = True
        for s in substituidos:
            s["_efetivo"] = False
            s["_substituido_por"] = efetivo.get("arquivo")

        grupos_per.append({
            "per_efetivo":   efetivo,
            "substituidos":  substituidos,
            "chave":         chave,
        })
        num = _normalizar_num(efetivo.get("numero"))
        if num:
            per_por_numero[num] = efetivo

    # ── Passo 2: vincular cada DCOMP ao seu PER efetivo ──────────────────────
    dcomps_por_per: dict[int, list] = {id(g["per_efetivo"]): [] for g in grupos_per}
    dcomps_nao_vinculadas: list[dict] = []

    for dcomp in dcomps:
        vinculado_a = None

        # 2a. Por número explícito referenciado no DCOMP
        ref = _normalizar_num(dcomp.get("referencia_per"))
        if ref and ref in per_por_numero:
            vinculado_a = per_por_numero[ref]

        # 2b. Por tributo + competência normalizada (mesmo crédito)
        if not vinculado_a:
            chave_dcomp = (dcomp.get("tributo"), _normalizar_competencia(dcomp.get("competencia")))
            for g in grupos_per:
                if g["chave"] == chave_dcomp:
                    vinculado_a = g["per_efetivo"]
                    break

        if vinculado_a:
            dcomps_por_per[id(vinculado_a)].append(dcomp)
        else:
            dcomps_nao_vinculadas.append(dcomp)

    # ── Passo 3: montar resultado com validação de saldos ───────────────────
    vinculos = []
    for g in grupos_per:
        per = g["per_efetivo"]
        linked = dcomps_por_per.get(id(per), [])

        credito      = per["valor_credito"]
        # Considera apenas DCOMPs efetivas (não substituídas por retificadoras)
        linked_ativos = [d for d in linked if d.get("_efetivo", True)]
        total_comp   = round(sum(d["valor_compensado"] for d in linked_ativos), 2)
        saldo_calc   = round(credito - total_comp, 2)
        saldo_decl   = per.get("saldo_remanescente", 0.0)

        # Validação de status
        # Nota: saldo_decl é o crédito original do PER (estático na data de entrega).
        # A comparação saldo_calc vs saldo_decl sempre diverge quando há DCOMPs — não usar para validação.
        if credito == 0:
            status = "SEM_VALOR"
        elif total_comp > credito * 1.005:
            status = "EXCEDIDO"
        elif not linked_ativos:
            status = "SEM_DCOMPS"
        else:
            status = "OK"

        alertas_vinc = []
        if status == "EXCEDIDO":
            excesso = total_comp - credito
            alertas_vinc.append(f"Total compensado (R$ {total_comp:,.2f}) supera o crédito do PER (R$ {credito:,.2f}) em R$ {excesso:,.2f}.")
        if g["substituidos"]:
            nomes = ", ".join(s["arquivo"] for s in g["substituidos"])
            alertas_vinc.append(f"Documento(s) substituído(s) por retificadora: {nomes}.")

        vinculos.append({
            "per_arquivo":          per["arquivo"],
            "per_numero":           per.get("numero"),
            "per_tributo":          per.get("tributo"),
            "per_competencia":      per.get("competencia"),
            "per_situacao":         per.get("situacao"),
            "per_data_tx":          per.get("data_transmissao"),
            "valor_credito":        credito,
            "tem_retificadora":     bool(g["substituidos"]) or per.get("retificador", False),
            "substituidos":         [{"arquivo": s["arquivo"], "numero": s.get("numero")} for s in g["substituidos"]],
            "dcomps": [
                {
                    "arquivo":          d["arquivo"],
                    "numero":           d.get("numero"),
                    "valor_compensado": d["valor_compensado"],
                    "credito_na_transmissao": d.get("credito_na_transmissao", 0),
                    "data_transmissao": d.get("data_transmissao"),
                    "situacao":         d.get("situacao"),
                    "referencia_per":   d.get("referencia_per"),
                    "substituida":      not d.get("_efetivo", True),
                }
                for d in sorted(linked, key=lambda x: x.get("data_transmissao") or "")
            ],
            "total_compensado":     total_comp,
            "saldo_calculado":      saldo_calc,
            "saldo_declarado":      saldo_decl,
            "percentual_utilizado": round(total_comp / credito * 100, 1) if credito > 0 else 0,
            "status_validacao":     status,
            "alertas_vinculo":      alertas_vinc,
        })

    return vinculos, dcomps_nao_vinculadas

# ─── MOTOR DE COMPLIANCE ──────────────────────────────────────────────────────

def analisar_compliance(registros: list, vinculos: list) -> list:
    alertas = []
    hoje = date.today()

    # R1 — Dupla utilização (mesmo número de processo)
    numeros_vistos: dict[str, dict] = {}
    for r in registros:
        if not r.get("numero"):
            continue
        base = _normalizar_num(r["numero"])
        if base in numeros_vistos:
            alertas.append({
                "nivel": "alto", "tipo": "Possível Dupla Utilização",
                "descricao": f"Número '{r['numero']}' aparece em mais de um documento.",
                "arquivos": [numeros_vistos[base]["arquivo"], r["arquivo"]],
            })
        else:
            numeros_vistos[base] = r

    # R2 — Saldo excedido por vinculação (considera apenas DCOMPs ativas)
    for v in vinculos:
        if v["status_validacao"] == "EXCEDIDO":
            dcomps_ativas = [d for d in v["dcomps"] if not d.get("substituida")]
            alertas.append({
                "nivel": "alto", "tipo": "Crédito do PER Excedido",
                "descricao": (f"PER '{v['per_arquivo']}': total compensado R$ {v['total_compensado']:,.2f} "
                              f"supera o crédito disponível R$ {v['valor_credito']:,.2f}."),
                "arquivos": [v["per_arquivo"]] + [d["arquivo"] for d in dcomps_ativas],
            })

    # R3 — Divergência de saldo
    for v in vinculos:
        if v["status_validacao"] == "DIVERGENCIA":
            alertas.append({
                "nivel": "medio", "tipo": "Divergência de Saldo",
                "descricao": (f"PER '{v['per_arquivo']}': saldo calculado R$ {v['saldo_calculado']:,.2f} "
                              f"diverge do saldo declarado R$ {v['saldo_declarado']:,.2f}."),
                "arquivos": [v["per_arquivo"]],
            })

    # R4 — DCOMPs sem PER vinculado
    for r in registros:
        if r["tipo"] == "Compensação" and not any(
            d["arquivo"] == r["arquivo"] for v in vinculos for d in v["dcomps"]
        ):
            alertas.append({
                "nivel": "medio", "tipo": "Compensação Sem PER Vinculado",
                "descricao": f"{r['arquivo']}: Não foi possível identificar o PER de origem desta compensação.",
                "arquivos": [r["arquivo"]],
            })

    # R5 — Prescrição
    for r in registros:
        comp = r.get("competencia")
        if not comp:
            continue
        try:
            partes = comp.split('/')
            if len(partes) != 2:
                continue
            mes, ano = int(partes[0]), int(partes[1])
            anos = (hoje - date(ano, mes, 1)).days / 365.25
            if anos > 5:
                alertas.append({
                    "nivel": "alto", "tipo": "Risco de Prescrição",
                    "descricao": f"{r['arquivo']}: Crédito de {comp} tem {anos:.1f} anos — possível prescrição.",
                    "arquivos": [r["arquivo"]],
                })
            elif anos > 4:
                alertas.append({
                    "nivel": "medio", "tipo": "Crédito Próximo da Prescrição",
                    "descricao": f"{r['arquivo']}: Crédito de {comp} tem {anos:.1f} anos — monitorar prazo de 5 anos.",
                    "arquivos": [r["arquivo"]],
                })
        except Exception:
            pass

    # R6 — Dados não extraídos
    for r in registros:
        falta = []
        if not r.get("tributo"):    falta.append("tributo")
        if not r.get("competencia"): falta.append("competência")
        if r["valor_credito"] == 0 and r["valor_compensado"] == 0:
            falta.append("valores monetários")
        if falta:
            alertas.append({
                "nivel": "medio", "tipo": "Dados Não Extraídos",
                "descricao": f"{r['arquivo']}: Não identificado: {', '.join(falta)}. Revisão manual necessária.",
                "arquivos": [r["arquivo"]],
            })

    # R7 — Retificadoras
    for r in registros:
        if r.get("retificador"):
            alertas.append({
                "nivel": "info", "tipo": "Documento Retificador",
                "descricao": f"{r['arquivo']}: Retificadora — substituiu documentos anteriores de mesmo tributo/competência.",
                "arquivos": [r["arquivo"]],
            })

    # R8 — Pedidos indeferidos/cancelados
    for r in registros:
        if r.get("situacao") in ("Indeferido", "Cancelado"):
            alertas.append({
                "nivel": "medio", "tipo": f"Pedido {r['situacao']}",
                "descricao": f"{r['arquivo']}: Status '{r['situacao']}' — verificar recurso ou retificação.",
                "arquivos": [r["arquivo"]],
            })

    return alertas

# ─── ROTAS ────────────────────────────────────────────────────────────────────

@bp.route('/api/perdcomp/health')
def health():
    return jsonify({"ok": True, "pdfplumber": PDF_OK,
                    "chrome_disponivel": not _chrome_bloqueado()})

# ─── ROTAS ECAC ──────────────────────────────────────────────────────────────

@bp.route('/api/perdcomp/ecac/iniciar', methods=['POST'])
def ecac_iniciar():
    data = request.json or {}
    periodo_ini = data.get("periodo_ini", "").strip()
    periodo_fim = data.get("periodo_fim", "").strip()
    if not periodo_ini or not periodo_fim:
        return jsonify({"error": "Informe o período inicial e final (MM/AAAA)."}), 400
    job_id = str(uuid.uuid4())
    _ECAC_JOBS[job_id] = {"status": "iniciando", "log": [], "arquivos": []}
    threading.Thread(target=_ecac_thread, args=(job_id, periodo_ini, periodo_fim), daemon=True).start()
    return jsonify({"job_id": job_id})

@bp.route('/api/perdcomp/ecac/status/<job_id>')
def ecac_status(job_id):
    job = _ECAC_JOBS.get(job_id)
    if not job:
        return jsonify({"error": "Job não encontrado"}), 404
    return jsonify(job)

@bp.route('/api/perdcomp/ecac/confirmar/<job_id>', methods=['POST'])
def ecac_confirmar(job_id):
    job = _ECAC_JOBS.get(job_id)
    if not job:
        return jsonify({"error": "Job não encontrado"}), 404
    job["usuario_confirmou"] = True
    return jsonify({"ok": True})

@bp.route('/api/perdcomp/ecac/pasta')
def ecac_pasta():
    pasta = _ECAC_DIR / "entrada"
    pasta.mkdir(exist_ok=True)
    pdfs = list(pasta.glob("*.pdf"))
    return jsonify({
        "pasta": str(pasta.resolve()),
        "arquivos": [f.name for f in pdfs],
        "total": len(pdfs),
    })

_ecac_capturas: list[dict] = []  # HTML capturado via console script

@bp.route('/api/perdcomp/ecac/capturar-html', methods=['POST'])
def ecac_capturar_html():
    """Recebe HTML de uma declaração capturada pelo script do console."""
    import html as html_mod
    import re as _re
    data = request.json or {}
    html_raw = data.get("html", "")
    indice   = data.get("indice", len(_ecac_capturas))

    # Extrair texto puro do HTML
    texto = _re.sub(r'<[^>]+>', ' ', html_raw)
    texto = html_mod.unescape(texto)
    texto = _re.sub(r'\s+', ' ', texto).strip()

    nome = f"ecac_captura_{indice+1:03d}.html"
    reg  = extrair_registro(texto, nome)
    with _capturas_lock:
        _ecac_capturas.append(reg)
        total = len(_ecac_capturas)
    return jsonify({"ok": True, "total": total, "tipo": reg["tipo"]})

@bp.route('/api/perdcomp/ecac/capturas')
def ecac_capturas_status():
    with _capturas_lock:
        snapshot = list(_ecac_capturas)
    return jsonify({"total": len(snapshot),
                    "registros": [{"arquivo": r["arquivo"], "tipo": r["tipo"],
                                   "numero": r["numero"]} for r in snapshot]})

@bp.route('/api/perdcomp/ecac/limpar-capturas', methods=['POST'])
def ecac_limpar_capturas():
    with _capturas_lock:
        _ecac_capturas.clear()
    return jsonify({"ok": True})

@bp.route('/api/perdcomp/ecac/analisar-capturas', methods=['POST'])
def ecac_analisar_capturas():
    with _capturas_lock:
        registros = [r for r in _ecac_capturas if r.get("tipo") != "Cancelamento"]
    if not registros:
        return jsonify({"error": "Nenhuma declaração capturada ainda."}), 400
    _marcar_retificadoras_dcomp(registros)
    vinculos, nv = vincular_pers_dcomps(registros)
    alertas = analisar_compliance(registros, vinculos)
    ativos = [r for r in registros if r.get("_efetivo", True)]
    total_credito    = sum(r["valor_credito"] for r in ativos if r["tipo"] in ("Ressarcimento","Restituição"))
    total_compensado = sum(r["valor_compensado"]   for r in ativos)
    total_ressarcido = sum(r["valor_ressarcido"]   for r in ativos)
    saldo_total      = round(total_credito - total_compensado - total_ressarcido, 2)
    dist_tributos: dict[str, float] = {}
    dist_tipos:    dict[str, int]   = {}
    for r in ativos:
        t  = r["tributo"] or "Não identificado"
        tp = r["tipo"]    or "Não identificado"
        val = r["valor_compensado"] if r["tipo"] == "Compensação" else r["valor_credito"]
        dist_tributos[t]  = round(dist_tributos.get(t, 0) + val, 2)
        dist_tipos[tp]    = dist_tipos.get(tp, 0) + 1
    return jsonify({
        "registros": registros, "vinculos": vinculos,
        "dcomps_nao_vinculadas": [d["arquivo"] for d in nv],
        "alertas": alertas, "erros": [],
        "sumario": {
            "total_arquivos": len(registros),
            "total_credito": round(total_credito, 2),
            "total_compensado": round(total_compensado, 2),
            "total_ressarcido": round(total_ressarcido, 2),
            "saldo_disponivel": round(saldo_total, 2),
            "alertas_alto":   sum(1 for a in alertas if a["nivel"] == "alto"),
            "alertas_medio":  sum(1 for a in alertas if a["nivel"] == "medio"),
            "alertas_info":   sum(1 for a in alertas if a["nivel"] == "info"),
            "dist_tributos": dist_tributos, "dist_tipos": dist_tipos,
            "vinculos_ok":        sum(1 for v in vinculos if v["status_validacao"] == "OK"),
            "vinculos_excedidos": sum(1 for v in vinculos if v["status_validacao"] == "EXCEDIDO"),
            "vinculos_diverg":    sum(1 for v in vinculos if v["status_validacao"] == "DIVERGENCIA"),
            "total_pers":  sum(1 for r in registros if r["tipo"] in ("Ressarcimento","Restituição")),
            "total_dcomps":sum(1 for r in registros if r["tipo"] == "Compensação"),
        }
    })

@bp.route('/api/perdcomp/ecac-script.js')
def ecac_script_js():
    """Serve o script de captura com suporte a paginação — usado pelo bookmarklet."""
    script = r"""
(async function(){
  const API='http://localhost:5000/api/perdcomp';

  // API descoberta: GET rest/api/documento-enviado/copia/{numero_sem_formatacao}
  // Extrai números da tabela e busca direto — sem precisar clicar em nada
  const BASE=window.location.href.split('#')[0].replace(/\/$/,'');
  const API_DOC=BASE+'/rest/api/documento-enviado/copia/';

  const getNumeros=()=>{
    const pat=/\d{5}[.\s]\d{5}[.\s]\d{6}[.\s]\d[.\s]\d+[.\s]\d{2}[-]\d{4}/g;
    const found=new Set();
    let m;
    while((m=pat.exec(document.body.innerText))!==null) found.add(m[0]);
    return[...found];
  };

  const getProxPag=()=>[...document.querySelectorAll('a,button,span,li')].find(el=>{
    const t=el.textContent.trim(),cls=(el.className||'').toLowerCase();
    return(t==='>'||t==='»'||t.toLowerCase()==='próxima'||cls.includes('next')||cls.includes('proximo'))&&
      !el.disabled&&!cls.includes('disabled');
  });

  const capturas=[];
  let total=0,pag=1;
  const processadas=new Set();

  while(true){
    const numeros=getNumeros();
    if(!numeros.length&&pag===1){
      alert('Nenhum número de PER/DCOMP encontrado.\nVá para Documentos Entregues, filtre o período e tente novamente.');
      return;
    }
    console.log('[Analyzer] Pág.'+pag+': '+numeros.length+' documento(s)');

    for(const num of numeros){
      if(processadas.has(num)) continue;
      processadas.add(num);
      const numLimpo=num.replace(/\D/g,'');
      const url=API_DOC+numLimpo;
      console.log('[Analyzer] Buscando: '+num);
      try{
        const resp=await fetch(url,{credentials:'include'});
        if(!resp.ok){console.warn('[Analyzer] HTTP '+resp.status+' para '+num);continue;}
        const html=await resp.text();
        capturas.push({html,indice:total,numero:num});
        total++;console.log('[Analyzer] ✓ Doc '+total+' ('+num+')');
      }catch(ex){console.error('[Analyzer] Erro:',ex.message);}
      await new Promise(r=>setTimeout(r,300));
    }

    const prox=getProxPag();
    if(!prox){console.log('[Analyzer] Última página.');break;}
    const fp=document.body.innerText.slice(0,300);
    prox.click();
    let mudou=false;
    for(let t=0;t<50;t++){await new Promise(r=>setTimeout(r,200));if(document.body.innerText.slice(0,300)!==fp){mudou=true;break;}}
    if(!mudou){console.log('[Analyzer] Última página.');break;}
    pag++;await new Promise(r=>setTimeout(r,1000));
  }

  // Gerar arquivo JSON para download (sem fetch — evita bloqueio CORS/PNA)
  const blob=new Blob([JSON.stringify(capturas)],{type:'application/json'});
  const url=URL.createObjectURL(blob);
  const a=document.createElement('a');
  a.href=url;a.download='perdcomp_capturas_'+new Date().toISOString().slice(0,10)+'.json';
  document.body.appendChild(a);a.click();document.body.removeChild(a);URL.revokeObjectURL(url);

  alert('[PER/DCOMP Analyzer]\n✓ '+total+' declaração(ões) em '+pag+' página(s).\n\nArquivo JSON baixado automaticamente!\nImporte-o na ferramenta PER/DCOMP Analyzer para análise.');
})();
"""
    script = script.replace("http://localhost:5000", f"http://localhost:{_BACKEND_PORT}")
    return script, 200, {
        'Content-Type': 'application/javascript',
        'Cache-Control': 'no-cache',
        'Access-Control-Allow-Origin': '*',
    }

@bp.route('/api/perdcomp/ecac/importar-json', methods=['POST'])
def ecac_importar_json():
    """Importa JSON gerado pelo script do console.
    Suporta pdf_b64 (PDF binário em base64 — formato atual) e html (legado)."""
    import base64 as _b64, json as _json
    import html as html_mod, re as _re

    f = request.files.get('file')
    if not f:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400
    try:
        data = _json.loads(f.read().decode('utf-8'))
    except Exception as e:
        return jsonify({"error": f"JSON inválido: {e}"}), 400

    if not isinstance(data, list) or not data:
        return jsonify({"error": "JSON vazio ou formato inválido"}), 400

    erros = []
    novos = []

    # Subpasta datada para os PDFs desta importação
    from datetime import datetime as _dt
    ts = _dt.now().strftime("%Y-%m-%d_%H-%M-%S")
    pasta_pdfs = _ECAC_DIR / "entrada" / ts
    pasta_pdfs.mkdir(parents=True, exist_ok=True)

    for item in data:
        numero = item.get('numero', '')
        indice = item.get('indice', len(_ecac_capturas))
        nome   = f"{re.sub(r'[^0-9]', '', numero) or str(indice+1)}.pdf"

        # ── Formato atual: PDF em base64 ──────────────────────────────────
        pdf_b64 = item.get('pdf_b64') or item.get('pdf_base64')
        if pdf_b64:
            try:
                pdf_bytes = _b64.b64decode(pdf_b64)
                # Salva PDF individual em subpasta datada
                (pasta_pdfs / nome).write_bytes(pdf_bytes)
                texto = extrair_texto(pdf_bytes)
                if texto.strip() and not texto.startswith("ERRO"):
                    novos.append(extrair_registro(texto, nome))
                    continue
                else:
                    erros.append({"arquivo": nome, "erro": "PDF sem texto extraível"})
                    continue
            except Exception as exc:
                erros.append({"arquivo": nome, "erro": str(exc)[:200]})
                continue

        # ── Formato legado: HTML ──────────────────────────────────────────
        html_raw = item.get('html', '')
        if html_raw:
            texto = _re.sub(r'<[^>]+>', ' ', html_raw)
            texto = html_mod.unescape(texto)
            texto = _re.sub(r'\s+', ' ', texto).strip()
            if texto:
                novos.append(extrair_registro(texto, nome))

    pdfs_salvos = len(list(pasta_pdfs.glob("*.pdf")))
    with _capturas_lock:
        _ecac_capturas.clear()
        _ecac_capturas.extend(novos)
        total = len(_ecac_capturas)

    # Metadados para projeto: empresa, CNPJ e range de competências
    empresa = next((r.get("empresa") for r in novos if r.get("empresa")), None)
    cnpj    = next((r.get("cnpj")    for r in novos if r.get("cnpj")),    None)
    comps   = sorted([r["competencia"] for r in novos if r.get("competencia")])
    comp_ini = comps[0]  if comps else None
    comp_fim = comps[-1] if comps else None

    return jsonify({
        "ok": True, "total": total, "erros": erros,
        "pasta_pdfs": str(pasta_pdfs.resolve()), "pdfs_salvos": pdfs_salvos,
        "empresa": empresa, "cnpj": cnpj,
        "competencia_ini": comp_ini, "competencia_fim": comp_fim,
        "total_pers":  sum(1 for r in novos if r["tipo"] in ("Ressarcimento","Restituição")),
        "total_dcomps": sum(1 for r in novos if r["tipo"] == "Compensação"),
    })

@bp.route('/api/perdcomp/projetos', methods=['GET'])
def listar_projetos():
    import sqlite3 as _sq
    from main import DB_PATH as _DB
    conn = _sq.connect(_DB)
    rows = conn.execute(
        "SELECT id,nome,empresa,cnpj,competencia_ini,competencia_fim,pasta_pdfs,total_docs,total_pers,total_dcomps,criado_em "
        "FROM perdcomp_projetos ORDER BY criado_em DESC"
    ).fetchall()
    conn.close()
    cols = ["id","nome","empresa","cnpj","competencia_ini","competencia_fim","pasta_pdfs","total_docs","total_pers","total_dcomps","criado_em"]
    return jsonify([dict(zip(cols, r)) for r in rows])

@bp.route('/api/perdcomp/projetos', methods=['POST'])
def salvar_projeto():
    import sqlite3 as _sq
    from main import DB_PATH as _DB
    d = request.get_json()
    if not d or not d.get("pasta_pdfs"):
        return jsonify({"error": "Dados insuficientes"}), 400
    pid = str(uuid.uuid4())[:8]
    conn = _sq.connect(_DB)
    conn.execute(
        "INSERT OR REPLACE INTO perdcomp_projetos VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        (pid, d.get("nome","Sem nome"), d.get("empresa",""), d.get("cnpj",""),
         d.get("competencia_ini",""), d.get("competencia_fim",""),
         d.get("pasta_pdfs",""), d.get("total_docs",0),
         d.get("total_pers",0), d.get("total_dcomps",0),
         d.get("criado_em",""))
    )
    conn.commit(); conn.close()
    return jsonify({"ok": True, "id": pid})

@bp.route('/api/perdcomp/projetos/<pid>', methods=['DELETE'])
def deletar_projeto(pid):
    import sqlite3 as _sq
    from main import DB_PATH as _DB
    conn = _sq.connect(_DB)
    conn.execute("DELETE FROM perdcomp_projetos WHERE id=?", (pid,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})

@bp.route('/api/perdcomp/projetos/<pid>/analisar', methods=['POST'])
def analisar_projeto(pid):
    import sqlite3 as _sq
    from main import DB_PATH as _DB
    conn = _sq.connect(_DB)
    row = conn.execute("SELECT pasta_pdfs FROM perdcomp_projetos WHERE id=?", (pid,)).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "Projeto não encontrado"}), 404
    pasta = Path(row[0])
    if not pasta.exists():
        return jsonify({"error": f"Pasta não encontrada: {pasta}"}), 404
    pdfs = list(pasta.glob("*.pdf"))
    if not pdfs:
        return jsonify({"error": "Nenhum PDF na pasta do projeto"}), 400
    registros, erros = [], []
    for pdf in pdfs:
        try:
            texto = extrair_texto(pdf.read_bytes())
            if not texto.strip(): continue
            reg = extrair_registro(texto, pdf.name)
            if reg["tipo"] != "Cancelamento":
                registros.append(reg)
        except Exception as exc:
            erros.append({"arquivo": pdf.name, "erro": str(exc)[:200]})
    _marcar_retificadoras_dcomp(registros)
    vinculos, dnv = vincular_pers_dcomps(registros)
    alertas = analisar_compliance(registros, vinculos)
    ativos = [r for r in registros if r.get("_efetivo", True)]
    total_credito    = sum(r["valor_credito"] for r in ativos if r["tipo"] in ("Ressarcimento","Restituição"))
    total_compensado = sum(r["valor_compensado"] for r in ativos)
    total_ressarcido = sum(r["valor_ressarcido"] for r in ativos)
    saldo_total      = round(total_credito - total_compensado - total_ressarcido, 2)
    dist_tributos: dict = {}; dist_tipos: dict = {}
    for r in ativos:
        t = r["tributo"] or "Não identificado"; tp = r["tipo"] or "Não identificado"
        val = r["valor_compensado"] if r["tipo"] == "Compensação" else r["valor_credito"]
        dist_tributos[t] = round(dist_tributos.get(t, 0) + val, 2)
        dist_tipos[tp]   = dist_tipos.get(tp, 0) + 1
    return jsonify({
        "registros": registros, "vinculos": vinculos,
        "dcomps_nao_vinculadas": [d["arquivo"] for d in dnv],
        "alertas": alertas, "erros": erros,
        "sumario": {
            "total_arquivos": len(ativos),
            "total_pers":  sum(1 for r in ativos if r["tipo"] in ("Ressarcimento","Restituição")),
            "total_dcomps":sum(1 for r in ativos if r["tipo"] == "Compensação"),
            "total_credito": round(total_credito,2), "total_compensado": round(total_compensado,2),
            "total_ressarcido": round(total_ressarcido,2), "saldo_disponivel": round(saldo_total,2),
            "alertas_alto":  sum(1 for a in alertas if a["nivel"]=="alto"),
            "alertas_medio": sum(1 for a in alertas if a["nivel"]=="medio"),
            "alertas_info":  sum(1 for a in alertas if a["nivel"]=="info"),
            "dist_tributos": dist_tributos, "dist_tipos": dist_tipos,
            "vinculos_ok":        sum(1 for v in vinculos if v["status_validacao"]=="OK"),
            "vinculos_excedidos": sum(1 for v in vinculos if v["status_validacao"]=="EXCEDIDO"),
            "vinculos_diverg":    sum(1 for v in vinculos if v["status_validacao"]=="DIVERGENCIA"),
        }
    })

@bp.route('/api/perdcomp/ecac/abrir-pasta')
def ecac_abrir_pasta():
    """Abre a pasta de entrada no Windows Explorer."""
    import subprocess as sp
    pasta = (_ECAC_DIR / "entrada")
    pasta.mkdir(exist_ok=True)
    sp.Popen(["explorer", str(pasta.resolve())])
    return jsonify({"ok": True})

@bp.route('/api/perdcomp/ecac/extrair-zip', methods=['POST'])
def ecac_extrair_zip():
    """Extrai PDFs de ZIPs enviados para a pasta de entrada."""
    import zipfile
    pasta = _ECAC_DIR / "entrada"
    pasta.mkdir(exist_ok=True)
    files = request.files.getlist('files')
    extraidos = 0
    erros = []
    for f in files:
        if not f.filename.lower().endswith('.zip'):
            continue
        try:
            with zipfile.ZipFile(io.BytesIO(f.read())) as z:
                for name in z.namelist():
                    if name.lower().endswith('.pdf'):
                        data = z.read(name)
                        dest = pasta / Path(name).name
                        dest.write_bytes(data)
                        extraidos += 1
        except Exception as exc:
            erros.append({"arquivo": f.filename, "erro": str(exc)})
    pdfs = list(pasta.glob("*.pdf"))
    return jsonify({"extraidos": extraidos, "total_pasta": len(pdfs), "erros": erros})


@bp.route('/api/perdcomp/ecac/limpar', methods=['POST'])
def ecac_limpar():
    """Remove todos os PDFs da pasta de entrada."""
    pasta = _ECAC_DIR / "entrada"
    pasta.mkdir(exist_ok=True)
    removidos = 0
    for f in pasta.glob("*.pdf"):
        try:
            f.unlink()
            removidos += 1
        except Exception:
            pass
    return jsonify({"removidos": removidos})

@bp.route('/api/perdcomp/ecac/analisar', methods=['POST'])
def ecac_analisar():
    """Analisa todos os PDFs da pasta de entrada (ou pasta específica via JSON body)."""
    if not PDF_OK:
        return jsonify({"error": "pdfplumber não instalado"}), 500
    # Aceita pasta específica via body JSON (usado pelo upload direto de PDFs)
    body = request.get_json(silent=True) or {}
    pasta_especifica = body.get("pasta_pdfs")
    if pasta_especifica:
        dest = Path(pasta_especifica)
    else:
        dest = _ECAC_DIR / "entrada"
    dest.mkdir(exist_ok=True)

    pdfs = list(dest.glob("*.pdf"))
    if not pdfs:
        return jsonify({"error": "Nenhum PDF na pasta de downloads"}), 400

    registros, erros = [], []
    for pdf in pdfs:
        try:
            texto = extrair_texto(pdf.read_bytes())
            if not texto.strip():
                erros.append({"arquivo": pdf.name, "erro": "PDF sem texto extraível"})
                continue
            reg = extrair_registro(texto, pdf.name)
            if reg["tipo"] == "Cancelamento":
                erros.append({"arquivo": pdf.name, "erro": "Pedido de Cancelamento — ignorado na análise"})
                continue
            registros.append(reg)
        except Exception as exc:
            erros.append({"arquivo": pdf.name, "erro": str(exc)[:200]})

    _marcar_retificadoras_dcomp(registros)
    vinculos, dcomps_nao_vinculadas = vincular_pers_dcomps(registros)
    alertas = analisar_compliance(registros, vinculos)

    ativos = [r for r in registros if r.get("_efetivo", True)]
    total_credito    = sum(r["valor_credito"] for r in ativos if r["tipo"] in ("Ressarcimento","Restituição"))
    total_compensado = sum(r["valor_compensado"]   for r in ativos)
    total_ressarcido = sum(r["valor_ressarcido"]   for r in ativos)
    saldo_total      = round(total_credito - total_compensado - total_ressarcido, 2)

    dist_tributos: dict[str, float] = {}
    dist_tipos:    dict[str, int]   = {}
    for r in ativos:
        t  = r["tributo"] or "Não identificado"
        tp = r["tipo"]    or "Não identificado"
        val = r["valor_compensado"] if r["tipo"] == "Compensação" else r["valor_credito"]
        dist_tributos[t]  = round(dist_tributos.get(t, 0) + val, 2)
        dist_tipos[tp]    = dist_tipos.get(tp, 0) + 1

    return jsonify({
        "registros": registros, "vinculos": vinculos,
        "dcomps_nao_vinculadas": [d["arquivo"] for d in dcomps_nao_vinculadas],
        "alertas": alertas, "erros": erros,
        "sumario": {
            "total_arquivos": len(ativos), "total_credito": round(total_credito, 2),
            "total_compensado": round(total_compensado, 2),
            "total_ressarcido": round(total_ressarcido, 2),
            "saldo_disponivel": round(saldo_total, 2),
            "alertas_alto":   sum(1 for a in alertas if a["nivel"] == "alto"),
            "alertas_medio":  sum(1 for a in alertas if a["nivel"] == "medio"),
            "alertas_info":   sum(1 for a in alertas if a["nivel"] == "info"),
            "dist_tributos": dist_tributos, "dist_tipos": dist_tipos,
            "vinculos_ok":        sum(1 for v in vinculos if v["status_validacao"] == "OK"),
            "vinculos_excedidos": sum(1 for v in vinculos if v["status_validacao"] == "EXCEDIDO"),
            "vinculos_diverg":    sum(1 for v in vinculos if v["status_validacao"] == "DIVERGENCIA"),
            "total_pers":  sum(1 for r in registros if r["tipo"] in ("Ressarcimento","Restituição")),
            "total_dcomps":sum(1 for r in registros if r["tipo"] == "Compensação"),
        }
    })

@bp.route('/api/perdcomp/capturas-debug')
def capturas_debug():
    """Retorna o que foi extraído de cada declaração capturada — para diagnóstico."""
    return jsonify([
        {
            "arquivo":    r["arquivo"],
            "tipo":       r["tipo"],
            "tributo":    r["tributo"],
            "numero":     r["numero"],
            "competencia": r["competencia"],
            "valor_credito":     r["valor_credito"],
            "valor_compensado":  r["valor_compensado"],
            "valor_ressarcido":  r["valor_ressarcido"],
            "saldo_remanescente":r["saldo_remanescente"],
            "referencia_per":    r.get("referencia_per"),
            "retificador":       r["retificador"],
        }
        for r in _ecac_capturas
    ])

@bp.route('/api/perdcomp/debug', methods=['POST'])
def debug():
    if not PDF_OK:
        return jsonify({"error": "pdfplumber não instalado"}), 500
    files = request.files.getlist('files')
    resultado = []
    for f in files:
        if not f.filename:
            continue
        texto = extrair_texto(f.read())
        resultado.append({
            "arquivo":              f.filename,
            "chars":                len(texto),
            "texto":                texto[:3000],
            "tipo_detectado":       detectar_tipo(texto, texto.lower()),
            "tributo_detectado":    detectar_tributo(texto),
            "numero_detectado":     extrair_numero(texto),
            "competencia_detectada":extrair_competencia_credito(texto),
            "referencia_per":       extrair_referencia_per(texto),
        })
    return jsonify(resultado)

@bp.route('/api/perdcomp/upload', methods=['POST'])
def upload():
    if not PDF_OK:
        return jsonify({"error": "pdfplumber não instalado. Execute: pip install pdfplumber"}), 500

    files = request.files.getlist('files')
    if not files or all(not f.filename for f in files):
        return jsonify({"error": "Nenhum arquivo enviado"}), 400

    registros, erros = [], []
    for f in files:
        if not f.filename:
            continue
        try:
            texto = extrair_texto(f.read())
            if not texto.strip() or texto.startswith("ERRO_EXTRACAO"):
                erros.append({"arquivo": f.filename,
                              "erro": texto if texto.startswith("ERRO") else
                              "PDF sem texto extraível — pode ser escaneado"})
                continue
            registros.append(extrair_registro(texto, f.filename))
        except Exception as e:
            erros.append({"arquivo": f.filename, "erro": str(e)[:300]})

    if not registros and erros:
        return jsonify({"error": "Nenhum arquivo processado", "detalhes": erros}), 400

    _marcar_retificadoras_dcomp(registros)
    vinculos, dcomps_nao_vinculadas = vincular_pers_dcomps(registros)
    alertas = analisar_compliance(registros, vinculos)

    total_credito    = sum(r["valor_credito"] for r in registros if r["tipo"] in ("Ressarcimento","Restituição"))
    total_compensado = sum(r["valor_compensado"]   for r in registros)
    total_ressarcido = sum(r["valor_ressarcido"]   for r in registros)
    saldo_total      = round(total_credito - total_compensado - total_ressarcido, 2)

    dist_tributos: dict[str, float] = {}
    dist_tipos:    dict[str, int]   = {}
    for r in registros:
        t  = r["tributo"] or "Não identificado"
        tp = r["tipo"]    or "Não identificado"
        val = r["valor_compensado"] if r["tipo"] == "Compensação" else r["valor_credito"]
        dist_tributos[t]  = round(dist_tributos.get(t, 0) + val, 2)
        dist_tipos[tp]    = dist_tipos.get(tp, 0) + 1

    return jsonify({
        "registros":              registros,
        "vinculos":               vinculos,
        "dcomps_nao_vinculadas":  [d["arquivo"] for d in dcomps_nao_vinculadas],
        "alertas":                alertas,
        "erros":                  erros,
        "sumario": {
            "total_arquivos":    len(registros),
            "total_pers":        sum(1 for r in registros if r["tipo"] in ("Ressarcimento", "Restituição")),
            "total_dcomps":      sum(1 for r in registros if r["tipo"] == "Compensação"),
            "total_credito":     round(total_credito, 2),
            "total_compensado":  round(total_compensado, 2),
            "total_ressarcido":  round(total_ressarcido, 2),
            "saldo_disponivel":  round(saldo_total, 2),
            "alertas_alto":      sum(1 for a in alertas if a["nivel"] == "alto"),
            "alertas_medio":     sum(1 for a in alertas if a["nivel"] == "medio"),
            "alertas_info":      sum(1 for a in alertas if a["nivel"] == "info"),
            "dist_tributos":     dist_tributos,
            "dist_tipos":        dist_tipos,
            "vinculos_ok":       sum(1 for v in vinculos if v["status_validacao"] == "OK"),
            "vinculos_excedidos":sum(1 for v in vinculos if v["status_validacao"] == "EXCEDIDO"),
            "vinculos_diverg":   sum(1 for v in vinculos if v["status_validacao"] == "DIVERGENCIA"),
        }
    })

@bp.route('/api/perdcomp/ecac/exportar-excel', methods=['POST'])
def ecac_exportar_excel():
    """Gera planilha Excel com registros, alertas e análise PER/DCOMP."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime as _dt
    import io as _io

    dados = request.get_json()
    if not dados:
        return jsonify({"error": "Dados não enviados"}), 400

    registros = dados.get("registros", [])
    alertas   = dados.get("alertas", [])
    vinculos  = dados.get("vinculos", [])
    sumario   = dados.get("sumario", {})

    wb = openpyxl.Workbook()

    # ── Estilos ──────────────────────────────────────────────────────────────
    def hdr(ws, row, cols, bg="1E293B", fg="F97316", bold=True):
        for c, v in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = Font(bold=bold, color=fg, name="Calibri", size=10)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="thin", color="334155"))
        ws.row_dimensions[row].height = 28

    def brl_fmt(v):
        return round(float(v or 0), 2)

    fmt_brl  = '#.##0,00'
    fmt_date = 'DD/MM/AAAA'

    # ── Aba 1: Registros ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Registros"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A2"

    colunas = ["Número","Arquivo","Tipo","Tributo","Competência",
               "Valor Crédito","Créd. Transmissão","Compensado","Ressarcido",
               "Saldo","Situação","Retificador","Status"]
    hdr(ws1, 1, colunas)

    for r in registros:
        ef = r.get("_efetivo", True)
        tipo = r.get("tipo","")
        row = [
            r.get("numero",""),
            r.get("arquivo",""),
            tipo,
            r.get("tributo",""),
            r.get("competencia",""),
            brl_fmt(r.get("valor_credito",0)) if tipo != "Compensação" else None,
            brl_fmt(r.get("credito_na_transmissao",0)) or None,
            brl_fmt(r.get("valor_compensado",0)) or None,
            brl_fmt(r.get("valor_ressarcido",0)) or None,
            brl_fmt(r.get("saldo_remanescente",0)),
            r.get("situacao",""),
            "Sim" if r.get("retificador") else "Não",
            "Substituído" if not ef else ("Retificadora" if r.get("retificador") else "Ativo"),
        ]
        ws1.append(row)
        ri = ws1.max_row
        for ci in [6,7,8,9,10]:
            cell = ws1.cell(ri, ci)
            if cell.value is not None:
                cell.number_format = fmt_brl
        if not ef:
            for ci in range(1, 14):
                ws1.cell(ri, ci).font = Font(color="94A3B8", italic=True, name="Calibri", size=9)

    widths = [32,36,15,8,22,16,18,14,12,14,14,12,12]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Aba 2: Alertas ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Alertas")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A2"
    hdr(ws2, 1, ["Nível","Tipo","Descrição","Documentos Relacionados"])
    cores = {"alto":"EF4444","medio":"F59E0B","info":"F97316"}
    for a in sorted(alertas, key=lambda x: {"alto":0,"medio":1,"info":2}.get(x.get("nivel",""),3)):
        arqs = ", ".join(a.get("arquivos",[]))
        ws2.append([a.get("nivel","").upper(), a.get("tipo",""), a.get("descricao",""), arqs])
        ri = ws2.max_row
        cor = cores.get(a.get("nivel",""), "94A3B8")
        ws2.cell(ri, 1).font = Font(bold=True, color=cor, name="Calibri", size=9)
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 60
    ws2.column_dimensions["D"].width = 50

    # ── Aba 3: Análise PER/DCOMP ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Análise PER-DCOMP")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A2"
    hdr(ws3, 1, ["PER Número","Tributo","Competência","Crédito Original",
                  "Total Compensado","Saldo Atual","% Utilizado","Status"])
    for v in vinculos:
        ws3.append([
            v.get("per_numero",""), v.get("per_tributo",""), v.get("per_competencia",""),
            brl_fmt(v.get("valor_credito",0)),
            brl_fmt(v.get("total_compensado",0)),
            brl_fmt(v.get("valor_credito",0)) - brl_fmt(v.get("total_compensado",0)),
            f"{v.get('percentual_utilizado',0):.1f}%",
            v.get("status_validacao",""),
        ])
        ri = ws3.max_row
        for ci in [4,5,6]:
            ws3.cell(ri, ci).number_format = fmt_brl
    for w, col in zip([32,10,22,16,16,14,12,14], range(1,9)):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # ── Aba 4: Resumo ────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Resumo")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 36
    ws4.column_dimensions["B"].width = 22

    def row4(label, value, bold=False):
        r = ws4.max_row + 1
        ws4.cell(r, 1, label).font = Font(bold=bold, name="Calibri", size=10, color="CBD5E1")
        cell = ws4.cell(r, 2, value)
        cell.font = Font(bold=True, name="Calibri", size=10, color="F97316" if bold else "F8FAFC")
        if isinstance(value, float):
            cell.number_format = fmt_brl

    ws4.cell(1, 1, "PER/DCOMP Analyzer — Resumo Executivo").font = Font(bold=True, size=13, color="F97316", name="Calibri")
    ws4.cell(2, 1, f"Gerado em {_dt.now().strftime('%d/%m/%Y %H:%M')}").font = Font(size=9, color="64748B", name="Calibri", italic=True)
    ws4.append([])
    row4("Total de documentos analisados", sumario.get("total_arquivos",0))
    row4("Pedidos de Ressarcimento/Restituição (PER)", sumario.get("total_pers",0))
    row4("Declarações de Compensação (DCOMP)", sumario.get("total_dcomps",0))
    ws4.append([])
    row4("Crédito original (PERs)", brl_fmt(sumario.get("total_credito",0)), bold=True)
    row4("Total compensado", brl_fmt(sumario.get("total_compensado",0)))
    row4("Total ressarcido", brl_fmt(sumario.get("total_ressarcido",0)))
    row4("Saldo disponível", brl_fmt(sumario.get("saldo_disponivel",0)), bold=True)
    ws4.append([])
    row4("Alertas de alto risco", sumario.get("alertas_alto",0))
    row4("Alertas de médio risco", sumario.get("alertas_medio",0))
    row4("Informativos", sumario.get("alertas_info",0))

    out = _io.BytesIO()
    wb.save(out)
    out.seek(0)
    nome = f"perdcomp_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=nome)

"""
Blueprint — Análise de SPED Contribuições (EFD-PIS/COFINS) para
transportadoras.

Recebe um ou mais arquivos SPED Contribuições (.txt), faz o parse e roda o
motor de regras G1-G5 (ver sped_contribuicoes_regras.py), retornando os
achados agrupados e um resumo executivo. Quando 2+ arquivos são enviados, os
achados de transposição de saldo (G4 — registros 1100/1500) são calculados
entre períodos consecutivos.
"""

import io
import uuid
from datetime import datetime

import pandas as pd
from flask import Blueprint, jsonify, request, send_file

from sped_contribuicoes_parser import parse_sped_file, extract_header
from sped_contribuicoes_regras import gerar_achados, gerar_achados_transposicao

bp = Blueprint("sped_contribuicoes", __name__, url_prefix="/api/sped-contribuicoes")
_analyses: dict[str, dict] = {}

NOTA_REVISAO = (
    "Esta análise é automatizada e tem caráter indicativo, com base no "
    "conteúdo do(s) arquivo(s) SPED Contribuições enviado(s), na legislação "
    "do PIS/COFINS não-cumulativo (Leis 10.637/2002 e 10.833/2003), no Guia "
    "Prático EFD-Contribuições e na IN RFB 2.121/2022. Todos os achados são "
    "pontos de atenção para revisão por especialista tributário, não "
    "constituindo posição definitiva sobre o aproveitamento de créditos."
)


def _competencia_key(header: dict) -> tuple[int, int]:
    comp = header.get("competencia", "")
    try:
        mes, ano = comp.split("/")
        return (int(ano), int(mes))
    except (ValueError, AttributeError):
        return (0, 0)


def _resumo(periodos: list[dict], achados: list[dict]) -> dict:
    contagem_tipo: dict[str, int] = {}
    contagem_severidade: dict[str, int] = {}
    valor_oportunidades = 0.0
    valor_riscos = 0.0
    valor_inconsistencias = 0.0

    for a in achados:
        contagem_tipo[a["tipo"]] = contagem_tipo.get(a["tipo"], 0) + 1
        contagem_severidade[a["severidade"]] = contagem_severidade.get(a["severidade"], 0) + 1
        if a["tipo"] == "OPORTUNIDADE":
            valor_oportunidades += abs(a["valor_envolvido"])
        elif a["tipo"] == "RISCO":
            valor_riscos += abs(a["valor_envolvido"])
        elif a["tipo"] == "INCONSISTENCIA":
            valor_inconsistencias += abs(a["valor_envolvido"])

    primeiro = periodos[0] if periodos else {}
    return {
        "empresa": primeiro.get("razao_social", ""),
        "cnpj": primeiro.get("cnpj", ""),
        "periodos_analisados": [p.get("competencia", "") for p in periodos],
        "total_achados": len(achados),
        "contagem_por_tipo": contagem_tipo,
        "contagem_por_severidade": contagem_severidade,
        "valor_total_oportunidades": round(valor_oportunidades, 2),
        "valor_total_riscos": round(valor_riscos, 2),
        "valor_total_inconsistencias": round(valor_inconsistencias, 2),
    }


@bp.route("/upload", methods=["POST"])
def upload():
    files = request.files.getlist("files")
    if not files or all(not f.filename for f in files):
        return jsonify({"error": "Nenhum arquivo enviado"}), 400

    periodos_dfs: list[tuple[dict, dict[str, pd.DataFrame]]] = []
    erros: list[str] = []

    for f in files:
        if not f.filename:
            continue
        try:
            conteudo = f.read()
            dfs = parse_sped_file(conteudo)
            header = extract_header(dfs)
            if not header.get("competencia"):
                erros.append(f"{f.filename}: não foi possível identificar o registro 0000 (cabeçalho).")
                continue
            header["arquivo"] = f.filename
            periodos_dfs.append((header, dfs))
        except Exception as e:
            erros.append(f"{f.filename}: erro ao processar — {e}")

    if not periodos_dfs:
        return jsonify({"error": "Nenhum arquivo válido processado", "detalhes": erros}), 400

    periodos_dfs.sort(key=lambda t: _competencia_key(t[0]))

    todos_achados: list[dict] = []
    por_periodo: list[dict] = []
    for header, dfs in periodos_dfs:
        achados = gerar_achados(dfs, header)
        todos_achados.extend(achados)
        por_periodo.append({
            "competencia": header.get("competencia", ""),
            "arquivo": header.get("arquivo", ""),
            "razao_social": header.get("razao_social", ""),
            "cnpj": header.get("cnpj", ""),
            "receita_bruta": header.get("receita_bruta"),
            "total_achados": len(achados),
        })

    if len(periodos_dfs) >= 2:
        todos_achados.extend(gerar_achados_transposicao(periodos_dfs))

    headers = [h for h, _ in periodos_dfs]
    resumo = _resumo(headers, todos_achados)

    analysis_id = str(uuid.uuid4())
    _analyses[analysis_id] = {
        "created_at": datetime.now().isoformat(),
        "resumo": resumo,
        "por_periodo": por_periodo,
        "achados": todos_achados,
    }

    return jsonify({
        "analysis_id": analysis_id,
        "aviso": NOTA_REVISAO,
        "erros": erros,
        "resumo": resumo,
        "por_periodo": por_periodo,
        "achados": todos_achados,
    })


@bp.route("/download/excel/<analysis_id>")
def download_excel(analysis_id):
    if analysis_id not in _analyses:
        return jsonify({"error": "Análise não encontrada. Refaça o upload."}), 404

    result = _analyses[analysis_id]
    resumo = result["resumo"]
    por_periodo = result["por_periodo"]
    achados = result["achados"]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Aba 1 — Resumo executivo
        ws_sum = writer.book.create_sheet("Resumo Executivo")
        for rd in [
            ["ANÁLISE SPED CONTRIBUIÇÕES (EFD-PIS/COFINS) — TRANSPORTADORAS"],
            ["Fundamentação: Leis 10.637/2002 e 10.833/2003, Guia Prático EFD-Contribuições, IN RFB 2.121/2022"],
            ["Empresa:", resumo.get("empresa", "")],
            ["CNPJ:", resumo.get("cnpj", "")],
            ["Períodos analisados:", ", ".join(resumo.get("periodos_analisados", []))],
            ["Gerado em:", datetime.now().strftime("%d/%m/%Y %H:%M")], [""],
            ["RESULTADO"],
            ["Total de achados:", resumo["total_achados"]],
            ["Oportunidades (possível crédito não aproveitado) — valor (R$):", resumo["valor_total_oportunidades"]],
            ["Riscos identificados — valor (R$):", resumo["valor_total_riscos"]],
            ["Inconsistências — valor (R$):", resumo["valor_total_inconsistencias"]], [""],
            ["DISTRIBUIÇÃO POR TIPO"],
            *[[k, v] for k, v in resumo["contagem_por_tipo"].items()], [""],
            ["DISTRIBUIÇÃO POR SEVERIDADE"],
            *[[k, v] for k, v in resumo["contagem_por_severidade"].items()], [""],
            [NOTA_REVISAO],
        ]:
            ws_sum.append(rd)
        ws_sum.column_dimensions["A"].width = 70
        ws_sum.column_dimensions["B"].width = 22

        # Aba 2 — Por período
        if por_periodo:
            df_per = pd.DataFrame([{
                "Competência": p["competencia"],
                "Arquivo": p["arquivo"],
                "Razão Social": p["razao_social"],
                "CNPJ": p["cnpj"],
                "Receita Trib. MI (R$)": (p["receita_bruta"] or {}).get("trib_mi", 0),
                "Receita Não Trib. MI (R$)": (p["receita_bruta"] or {}).get("nt_mi", 0),
                "Receita Exportação (R$)": (p["receita_bruta"] or {}).get("exp", 0),
                "Receita Total (R$)": (p["receita_bruta"] or {}).get("total", 0),
                "Achados no Período": p["total_achados"],
            } for p in por_periodo])
            df_per.to_excel(writer, sheet_name="Períodos", index=False)

        # Abas 3-7 — Achados por grupo de regra
        grupos = {
            "G1": "Créditos A-C-F (Insumos)",
            "G2": "Ativo Imobilizado-Frota",
            "G3": "Reconciliação Bloco M",
            "G4": "Transposição 1100-1500",
            "G5": "Rateio Proporcional 0111",
        }
        for grupo, nome_aba in grupos.items():
            linhas = [a for a in achados if a["grupo"] == grupo]
            if not linhas:
                continue
            df_g = pd.DataFrame([{
                "Tipo": a["tipo"],
                "Severidade": a["severidade"],
                "Competência": a["competencia"],
                "Bloco": a["bloco"],
                "Registro": a["registro"],
                "Descrição": a["descricao"],
                "Valor Envolvido (R$)": a["valor_envolvido"],
                "Base Legal": a["base_legal"],
                "Recomendação": a["recomendacao"],
            } for a in linhas])
            df_g.to_excel(writer, sheet_name=nome_aba, index=False)

            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            ws = writer.sheets[nome_aba]
            hf = PatternFill("solid", start_color="1B3A5C")
            hfont = Font(bold=True, color="FFFFFF", size=10)
            for c in range(1, len(df_g.columns) + 1):
                cell = ws.cell(row=1, column=c)
                cell.fill = hf
                cell.font = hfont
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            cmap = {
                "OPORTUNIDADE": PatternFill("solid", start_color="C6EFCE"),
                "RISCO": PatternFill("solid", start_color="FFC7CE"),
                "INCONSISTENCIA": PatternFill("solid", start_color="FFEB9C"),
                "INFORMATIVO": PatternFill("solid", start_color="D9D9D9"),
            }
            for ri, a in enumerate(linhas, start=2):
                fill = cmap.get(a["tipo"], PatternFill("solid", start_color="FFFFFF"))
                for c in range(1, len(df_g.columns) + 1):
                    ws.cell(row=ri, column=c).fill = fill

            widths = [16, 12, 12, 8, 10, 70, 16, 55, 70]
            for i, w in enumerate(widths[:len(df_g.columns)], 1):
                ws.column_dimensions[get_column_letter(i)].width = w

    output.seek(0)
    dl_name = f"analise_sped_contribuicoes_{analysis_id[:8]}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=dl_name,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/health")
def health():
    return jsonify({"status": "ok", "analyses_cached": len(_analyses)})
